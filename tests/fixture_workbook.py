"""Programmatic POPS miniature used by the end-to-end tests."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


GREEN = "08775F"
NAVY = "0B315E"
BLUE = "0088A8"
PURPLE = "5B3F92"
ORANGE = "C95712"
WHITE = "FFFFFF"
GRID = Side(style="thin", color="59697A")
BORDER = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)


def _style_grid(ws, ref: str) -> None:
    for row in ws[ref]:
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _header(cell, color: str = GREEN) -> None:
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def create_pops_mini(path: Path) -> Path:
    workbook = Workbook()
    synthesis = workbook.active
    synthesis.title = "SYNTHESIS -->"
    synthesis["A1"] = "Navigation"

    obs = workbook.create_sheet("OBS KPI")
    obs["B1"] = "Copy table below to PowerPoint"
    obs.merge_cells("B3:F3")
    obs["B3"] = "OBS KPIs EVOLUTION"
    _header(obs["B3"])
    obs.merge_cells("G3:H3")
    obs["G3"] = "Δ"
    _header(obs["G3"], NAVY)
    headers = [
        "KPI",
        "Actual 2025",
        "F3 2026",
        "PLAN 2026",
        "B 2027",
        "B2027 vs F3 2026",
        "B2027 vs PLAN 2026",
    ]
    for column, label in enumerate(headers, 2):
        obs.cell(4, column, label)
        _header(obs.cell(4, column), [GREEN, GREEN, ORANGE, BLUE, PURPLE, NAVY, NAVY][column - 2])
    obs.merge_cells("B5:H5")
    obs["B5"] = "TRANSFORMATION KPI'S"
    _header(obs["B5"])
    values = [
        ("Selfcare rate (%)", 0.628, 0.695, "=D6-C6", 0.72, "=F6-D6", "=F6-E6"),
        ("Global Productivity Index", 139, "='Calc'!B2", 150, 155, "=F7-D7", "=F7-E7"),
        ("Desk Sharing Ratio (%)", 0.3, "-", "ns", 0.4, None, None),
    ]
    for row, values_row in enumerate(values, 6):
        for column, value in enumerate(values_row, 2):
            obs.cell(row, column, value)
    obs.merge_cells("B9:H9")
    obs["B9"] = "COST & PRODUCTIVITY"
    _header(obs["B9"])
    obs["B10"] = "Customer Services VC OPEX (M€)"
    obs["C10"] = -10.5
    obs["D10"] = 0
    obs["E10"] = -8.6
    obs["F10"] = -9.2
    obs["G10"] = "=F10-D10"
    obs["H10"] = "=F10-E10"
    obs["B11"] = "TOTAL OPEX"
    obs["C11"] = "=C10"
    obs["D11"] = "=D10"
    obs["E11"] = "=E10"
    obs["F11"] = "=F10"
    obs["G11"] = "=G10"
    obs["H11"] = "=H10"
    for cell in obs[11]:
        cell.font = Font(bold=True)
    _style_grid(obs, "B3:H11")
    for cell in obs[6][2:]:
        cell.number_format = "0.0%"
    obs["B14"] = "Note: formula caches are intentionally absent in this generated fixture."
    obs["B14"].font = Font(italic=True, color="FFFFFF")
    obs["B14"].fill = PatternFill("solid", fgColor=NAVY)
    obs.column_dimensions["D"].hidden = True
    obs.column_dimensions["D"].outlineLevel = 1
    obs.row_dimensions[7].hidden = True
    obs.row_dimensions[7].outlineLevel = 1
    obs.freeze_panes = "C5"
    obs["B6"].comment = Comment("New definition; country input expected.", "POPS")
    validation = DataValidation(type="decimal", operator="between", formula1="0", formula2="1")
    validation.promptTitle = "Rate"
    validation.prompt = "Enter a rate between 0 and 1"
    obs.add_data_validation(validation)
    validation.add("C6:F6")
    obs.conditional_formatting.add(
        "C6:F6", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFCCCC"))
    )
    obs.sheet_properties.outlinePr.summaryBelow = True

    mbs = workbook.create_sheet("MBS (OPEX)")
    mbs["B1"] = "Copy table below to PowerPoint"
    mbs.merge_cells("B3:C3")
    mbs["B3"] = "BUY & SHARE"
    _header(mbs["B3"])
    for ref, label in (("D3:G3", "OPEX (M€)"), ("H3:K3", "FTE Average"), ("L3:O3", "FTE EoP")):
        mbs.merge_cells(ref)
        anchor = mbs[ref.split(":")[0]]
        anchor.value = label
        _header(anchor)
    period_labels = ["2025", "F3 2026", "PLAN 2026", "B 2027"] * 3
    for column, label in enumerate(period_labels, 4):
        mbs.cell(4, column, label)
        _header(mbs.cell(4, column), BLUE if "PLAN" in label else PURPLE)
    for row, label in enumerate(
        ["GRANTING", "CUSTOMER SERVICE", "COLLECTION", "RECOVERY", "TOTAL"], 5
    ):
        mbs["B" + str(row)] = "BUSINESS OUTSOURCING" if row == 5 else None
        mbs["C" + str(row)] = label
        for column in range(4, 16):
            mbs.cell(row, column, (row - 4) * (column - 2))
    mbs.merge_cells("B5:B9")
    mbs["B5"].alignment = Alignment(text_rotation=90, horizontal="center", vertical="center")
    _style_grid(mbs, "B3:O9")
    mbs.merge_cells("B13:O13")
    mbs["B13"] = "BUSINESS OUTSOURCING - DETAIL BY PROVIDER"
    _header(mbs["B13"])
    detail_headers = ["Department", "Provider", "Main mission", "Contract expiry"]
    for column, label in enumerate(detail_headers, 2):
        mbs.cell(14, column, label)
        _header(mbs.cell(14, column), NAVY)
    for row in range(15, 19):
        mbs.cell(row, 2, "GRANTING" if row == 15 else "")
        mbs.cell(row, 3, "xxx" if row < 18 else "TOTAL")
        mbs.cell(row, 4, "xxx" if row < 18 else "")
        mbs.cell(row, 5, "yyyy" if row < 18 else "")
        for column in range(6, 16):
            mbs.cell(row, column, None if row < 18 else 0)
    _style_grid(mbs, "B13:O18")
    mbs["Q4"] = "Please ensure details correspond to totals"
    mbs["Q4"].font = Font(color="C00000", bold=True)
    mbs["Q4"].alignment = Alignment(wrap_text=True)

    kpi = workbook.create_sheet("KPI")
    kpi.append(["KPI", "Source", "Unit", "Annual Reference calculation", "Actual 2024", "Forecast 2026"])
    kpi.append(["Gross Production", "ILON", "€", "annual sum", 6_087_743, 6_800_937])
    kpi.append(["Opened accounts", "ILON", "nb", "annual sum", 836_383, 866_440])
    kpi.append(["Active accounts EoP", "ILON", "nb", "december reference", 2_126_197, 2_093_927])
    kpi.append(["TOTAL", "", "", "", "=SUM(E2:E4)", "=SUM(F2:F4)"])
    native = Table(displayName="KPIReference", ref="A1:F5")
    native.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    kpi.add_table(native)
    kpi["A2"].hyperlink = "https://example.invalid/source"
    kpi["A2"].comment = Comment("Source system: ILON", "POPS")
    source_validation = DataValidation(type="list", formula1='"ILON,PERSEUS"', allow_blank=True)
    kpi.add_data_validation(source_validation)
    source_validation.add("B2:B5")
    kpi.column_dimensions["F"].hidden = True
    kpi.row_dimensions[4].hidden = True

    calc = workbook.create_sheet("Calc")
    calc["A1"] = "Global Productivity Index"
    calc["B2"] = 145
    calc.sheet_state = "hidden"
    internal = workbook.create_sheet("Internal")
    internal["A1"] = "Do not edit"
    internal.sheet_state = "veryHidden"

    local_name = DefinedName("OBS_Input", attr_text="'OBS KPI'!$C$6:$F$10")
    workbook.defined_names.add(local_name)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)
    return path
