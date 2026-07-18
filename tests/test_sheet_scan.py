from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, PatternFill, Side

from pops_ingest.config import ExtractionConfig
from pops_ingest.ooxml import OOXMLIndex
from pops_ingest.sheet_scan import StyleRegistry, cell_snapshot, scan_sheet
from pops_ingest.utils import canonical_json
from tests.fixture_workbook import create_pops_mini


class SheetScanTests(unittest.TestCase):
    def test_semantic_styles_and_local_style_ids_are_both_preserved(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = create_pops_mini(Path(directory) / "mini.xlsx")
            config = ExtractionConfig()
            index = OOXMLIndex.open(source, config)
            formula_book = load_workbook(source, data_only=False, rich_text=True)
            cached_book = load_workbook(source, data_only=True, rich_text=True)
            try:
                ws = formula_book["OBS KPI"]
                cached_ws = cached_book["OBS KPI"]
                registry = StyleRegistry()
                profile = scan_sheet(
                    ws,
                    cached_ws,
                    index.formula_records["OBS KPI"],
                    config,
                    registry,
                )
                self.assertEqual(
                    registry.fingerprint(ws["C6"]), registry.fingerprint(ws["D6"])
                )
                self.assertNotEqual(
                    registry.fingerprint(ws["C6"]), registry.fingerprint(ws["C4"])
                )
                snapshot = cell_snapshot(
                    ws,
                    cached_ws,
                    6,
                    3,
                    profile,
                    index.formula_records["OBS KPI"],
                    registry,
                )
                self.assertIsInstance(snapshot["style_id"], int)
                self.assertTrue(snapshot["style_hash"])
                self.assertNotIn("Values must be of type", canonical_json(registry.to_dict()))
            finally:
                formula_book.close()
                cached_book.close()

    def test_remote_fill_only_cell_does_not_inflate_detection_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "remote_style.xlsx"
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Data"
            ws["A1"] = "Metric"
            ws["B1"] = "2026"
            ws["A2"] = "KPI"
            ws["B2"] = 10
            side = Side(style="thin", color="000000")
            for row in ws["A1:B2"]:
                for cell in row:
                    cell.border = Border(left=side, right=side, top=side, bottom=side)
            ws["Z1000"].fill = PatternFill("solid", fgColor="FF0000")
            workbook.save(source)

            formula_book = load_workbook(source, data_only=False)
            cached_book = load_workbook(source, data_only=True)
            try:
                registry = StyleRegistry()
                profile = scan_sheet(
                    formula_book["Data"],
                    cached_book["Data"],
                    {},
                    ExtractionConfig(),
                    registry,
                )
                self.assertEqual(profile.style_bounds.ref, "A1:Z1000")
                self.assertEqual(profile.evidence_bounds.ref, "A1:B2")
                self.assertIn((1000, 26), profile.material_coordinates)
            finally:
                formula_book.close()
                cached_book.close()

    def test_malformed_formula_is_preserved_with_tokenizer_diagnostic(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "broken_formula.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "Metric"
            worksheet["B1"] = "2026"
            worksheet["A2"] = "Broken formula"
            worksheet["B2"] = "=#REF!:#REF!"
            workbook.save(source)

            config = ExtractionConfig()
            index = OOXMLIndex.open(source, config)
            formula_book = load_workbook(source, data_only=False)
            cached_book = load_workbook(source, data_only=True)
            try:
                registry = StyleRegistry()
                profile = scan_sheet(
                    formula_book["Data"],
                    cached_book["Data"],
                    index.formula_records["Data"],
                    config,
                    registry,
                )
                snapshot = cell_snapshot(
                    formula_book["Data"],
                    cached_book["Data"],
                    2,
                    2,
                    profile,
                    index.formula_records["Data"],
                    registry,
                )
                formula = snapshot["formula"]
                self.assertEqual(formula["exact"], "=#REF!:#REF!")
                self.assertEqual(formula["dependencies"], [])
                self.assertEqual(
                    formula["normalized_relative_signature"], "=#REF!:#REF!"
                )
                self.assertEqual(formula["tokenization_status"], "failed")
                self.assertEqual(
                    formula["tokenization_error"]["type"], "TokenizerError"
                )
                self.assertTrue(
                    any(
                        warning.code == "FORMULA_TOKENIZATION_FAILED"
                        and warning.coordinate == "B2"
                        for warning in profile.warnings
                    )
                )
            finally:
                formula_book.close()
                cached_book.close()


if __name__ == "__main__":
    unittest.main()
