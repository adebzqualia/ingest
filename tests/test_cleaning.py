from __future__ import annotations

import re
import unittest

from openpyxl.utils import get_column_letter

from pops_ingest.cleaning import build_clean_table


def _snapshot(
    row: int,
    column: int,
    literal: object = None,
    *,
    formula: str | None = None,
    cached: object = None,
    cache_status: str = "not_applicable",
    number_format: str = "General",
    merge_anchor: str | None = None,
    merge_ref: str | None = None,
) -> dict[str, object]:
    coordinate = f"{get_column_letter(column)}{row}"
    if formula and cache_status == "not_applicable":
        cache_status = "missing" if cached is None else "present"
    return {
        "coordinate": coordinate,
        "row": row,
        "column": column,
        "column_letter": get_column_letter(column),
        "literal_value": literal,
        "literal_value_type": (
            "blank"
            if literal is None
            else "number"
            if isinstance(literal, (int, float)) and not isinstance(literal, bool)
            else "text"
        ),
        "formula": formula,
        "cached_value": cached,
        "cached_value_type": (
            "blank"
            if cached is None
            else "number"
            if isinstance(cached, (int, float)) and not isinstance(cached, bool)
            else "text"
        ),
        "cached_value_status": cache_status,
        "number_format": number_format,
        "merge_anchor": merge_anchor,
        "merge_ref": merge_ref,
    }


def _dropped_indexes(items: list[object], key: str) -> set[int]:
    result: set[int] = set()
    for item in items:
        if isinstance(item, int):
            result.add(item)
        elif isinstance(item, dict):
            value = item.get(key)
            if value is None:
                value = item.get("row" if key == "source_row" else "column")
            if value is not None:
                result.add(int(value))
    return result


def _leaf_label(label: object) -> str:
    parts = re.split(r"\s*(?:\||/|>)\s*", str(label))
    return parts[-1].strip()


def _messy_table_fixture() -> tuple[
    dict[str, object],
    dict[str, object],
    dict[tuple[int, int], dict[str, object]],
]:
    table: dict[str, object] = {
        "table_id": "tbl_cleaning_fixture",
        "sheet": "OBS KPI",
        "range": "A1:H10",
        "title": "ENTITY KEY FIGURES",
    }
    structure: dict[str, object] = {
        "title": "ENTITY KEY FIGURES",
        # Row 1 is a PowerPoint-copy banner that an upstream heuristic treated
        # as a header. Rows 2 and 3 are the actual parent and leaf headers.
        "header_rows": [1, 2, 3],
        "column_descriptors": [
            {
                "column": 1,
                "letter": "A",
                "role": "dimension",
                "header_path": ["ENTITY KEY FIGURES", "Entity"],
                "header_sources": ["A2", "A3"],
                "key": "entity-0a1b2c3d",
            },
            {
                "column": 2,
                "letter": "B",
                "role": "row_group",
                "header_path": ["ENTITY KEY FIGURES", "Activity"],
                "header_sources": ["A2", "B3"],
                "key": "activity-1a2b3c4d",
            },
            {
                "column": 3,
                "letter": "C",
                "role": "row_label",
                "header_path": ["ENTITY KEY FIGURES", "Metric"],
                "header_sources": ["A2", "C3"],
                "key": "metric-2a3b4c5d",
            },
            {
                "column": 4,
                "letter": "D",
                "role": "unknown",
                "header_path": [],
                "header_sources": [],
                "key": "column-4-3a4b5c6d",
            },
            {
                "column": 5,
                "letter": "E",
                "role": "measure",
                "header_path": ["Baseline", "2024"],
                "header_sources": ["E2", "E3"],
                "key": "baseline-2024-4a5b6c7d",
                "scenario": "baseline",
                "year": 2024,
            },
            {
                "column": 6,
                "letter": "F",
                "role": "measure",
                "header_path": ["Baseline", "2025"],
                "header_sources": ["E2", "F3"],
                "key": "baseline-2025-5a6b7c8d",
                "scenario": "baseline",
                "year": 2025,
            },
            {
                "column": 7,
                "letter": "G",
                "role": "unknown",
                "header_path": [],
                "header_sources": [],
                "key": "column-7-6a7b8c9d",
            },
            {
                "column": 8,
                "letter": "H",
                "role": "formula_measure",
                "header_path": ["Forecast", "F1 2026"],
                "header_sources": ["H2", "H3"],
                "key": "forecast-f1-2026-7a8b9c0d",
                "scenario": "forecast",
                "year": 2026,
            },
        ],
        "row_descriptors": [
            {"row": 1, "role": "title", "section_path": []},
            {"row": 4, "role": "spacer", "section_path": []},
            {
                "row": 5,
                "role": "section_header",
                "metric_label": "ACTIVITY",
                "label_path": ["ACTIVITY"],
                "section_path": [],
            },
            {
                "row": 6,
                "role": "data",
                "metric_label": "Outstanding",
                "label_path": ["Europe", "Granting", "Outstanding"],
                "section_path": [],
            },
            {
                "row": 7,
                "role": "note",
                "metric_label": "Note: values are in millions",
                "label_path": ["Note: values are in millions"],
                "section_path": [],
            },
            {
                "row": 8,
                "role": "instruction",
                "metric_label": "Please complete all yellow cells",
                "label_path": ["Please complete all yellow cells"],
                "section_path": [],
            },
            {
                "row": 9,
                "role": "data",
                "metric_label": "Active Accounts",
                "label_path": ["Customer Service", "Active Accounts"],
                "section_path": [],
            },
            {"row": 10, "role": "spacer", "section_path": []},
        ],
        "sections": [
            {"title": "ACTIVITY", "start_row": 5, "end_row": 9}
        ],
    }

    cells = [
        _snapshot(1, 1, "Copy table below to PowerPoint", merge_anchor="A1", merge_ref="A1:H1"),
        _snapshot(2, 1, "ENTITY KEY FIGURES", merge_anchor="A2", merge_ref="A2:C2"),
        _snapshot(2, 5, "Baseline", merge_anchor="E2", merge_ref="E2:F2"),
        _snapshot(2, 6, None, merge_anchor="E2", merge_ref="E2:F2"),
        _snapshot(2, 8, "Forecast"),
        _snapshot(3, 1, "Entity"),
        _snapshot(3, 2, "Activity"),
        _snapshot(3, 3, "Metric"),
        _snapshot(3, 4, None),
        _snapshot(3, 5, "2024"),
        _snapshot(3, 6, "2025"),
        _snapshot(3, 7, None),
        _snapshot(3, 8, "F1 2026"),
        _snapshot(4, 1, None),
        _snapshot(4, 8, None),
        _snapshot(5, 1, "ACTIVITY", merge_anchor="A5", merge_ref="A5:H5"),
        _snapshot(6, 1, "Europe", merge_anchor="A6", merge_ref="A6:A9"),
        _snapshot(6, 2, "Granting"),
        _snapshot(6, 3, "Outstanding"),
        _snapshot(6, 4, None),
        _snapshot(6, 5, 100, number_format="# ##0"),
        _snapshot(6, 6, 110, number_format="# ##0"),
        _snapshot(6, 7, None),
        _snapshot(6, 8, 120, number_format="# ##0"),
        _snapshot(7, 3, "Note: values are in millions"),
        _snapshot(8, 3, "Please complete all yellow cells"),
        _snapshot(9, 1, None, merge_anchor="A6", merge_ref="A6:A9"),
        _snapshot(9, 2, "Customer Service"),
        _snapshot(9, 3, "Active Accounts"),
        _snapshot(9, 4, None),
        _snapshot(9, 5, 200, number_format="# ##0"),
        _snapshot(9, 6, 210, number_format="# ##0"),
        _snapshot(9, 7, None),
        _snapshot(
            9,
            8,
            formula="=F9+10",
            cached=None,
            cache_status="missing",
            number_format="# ##0",
        ),
        _snapshot(10, 1, None),
        _snapshot(10, 8, None),
    ]
    snapshots = {
        (int(cell["row"]), int(cell["column"])): cell for cell in cells
    }
    return table, structure, snapshots


class CleanTableTests(unittest.TestCase):
    def test_recovers_physical_headers_when_semantic_guess_only_saw_banner(self) -> None:
        table, structure, snapshots = _messy_table_fixture()
        table["title"] = "A1:H10"
        structure["title"] = "A1:H10"
        structure["header_rows"] = [1]
        for descriptor in structure["column_descriptors"]:
            descriptor["header_path"] = []
            descriptor["header_sources"] = ["A1"] if descriptor["column"] == 1 else []

        clean = build_clean_table(table, structure, snapshots)

        self.assertEqual(clean["status"], "ready")
        self.assertEqual(clean["title"], "ENTITY KEY FIGURES")
        self.assertEqual(clean["header_rows"], [2, 3])
        self.assertEqual(
            [_leaf_label(column["label"]) for column in clean["columns"]],
            ["Entity", "Activity", "Metric", "2024", "2025", "F1 2026"],
        )
        self.assertNotIn("deadbeef", repr(clean))

    def test_uses_real_headers_and_removes_banner_and_blank_gutters(self) -> None:
        table, structure, snapshots = _messy_table_fixture()
        clean = build_clean_table(table, structure, snapshots)

        self.assertEqual(clean["status"], "ready")
        self.assertEqual(clean["title"], "ENTITY KEY FIGURES")
        self.assertEqual(clean["source_range"], "A1:H10")
        self.assertEqual(clean["header_rows"], [2, 3])

        columns = clean["columns"]
        self.assertEqual(
            [column["source_column"] for column in columns], [1, 2, 3, 5, 6, 8]
        )
        self.assertEqual(
            [column["source_letter"] for column in columns],
            ["A", "B", "C", "E", "F", "H"],
        )
        self.assertEqual(
            [_leaf_label(column["label"]) for column in columns],
            ["Entity", "Activity", "Metric", "2024", "2025", "F1 2026"],
        )
        self.assertEqual(len({column["key"] for column in columns}), 6)
        for column in columns:
            self.assertFalse(
                re.search(r"-[0-9a-f]{8}$", str(column["key"]), re.IGNORECASE),
                column,
            )
            self.assertFalse(bool(column.get("synthetic", False)), column)
            self.assertTrue(column["header_path"])

        self.assertEqual(clean["counts"]["columns"], 6)
        self.assertEqual(
            _dropped_indexes(clean["dropped_columns"], "source_column"), {4, 7}
        )

    def test_drops_non_data_rows_and_propagates_section_and_merged_anchor(self) -> None:
        table, structure, snapshots = _messy_table_fixture()
        clean = build_clean_table(table, structure, snapshots)

        rows = clean["rows"]
        self.assertEqual([row["source_row"] for row in rows], [6, 9])
        self.assertEqual([row["role"] for row in rows], ["data", "data"])
        self.assertEqual(
            [row["section_path"] for row in rows],
            [["ACTIVITY"], ["ACTIVITY"]],
        )
        self.assertEqual(clean["counts"]["rows"], 2)
        self.assertTrue(
            {1, 4, 5, 7, 8, 10}.issubset(
                _dropped_indexes(clean["dropped_rows"], "source_row")
            )
        )

        # The second data row's entity cell is a merged continuation. Its clean
        # value must come from A6 while retaining A9 as the source coordinate.
        second_entity = rows[1]["cells"][0]
        self.assertEqual(second_entity["coordinate"], "A9")
        self.assertEqual(second_entity["value"], "Europe")
        self.assertIsNone(second_entity["formula"])
        self.assertIsNone(second_entity["cached"])
        self.assertEqual(second_entity["literal"], "Europe")
        self.assertEqual(second_entity["value_source_coordinate"], "A6")
        self.assertTrue(second_entity["merged_inherited"])

        output_text = repr(clean)
        self.assertNotIn("Copy table below to PowerPoint", output_text)
        self.assertNotIn("Note: values are in millions", output_text)
        self.assertNotIn("Please complete all yellow cells", output_text)

    def test_formula_without_cache_remains_auditable_and_readable(self) -> None:
        table, structure, snapshots = _messy_table_fixture()
        clean = build_clean_table(table, structure, snapshots)

        formula_cell = next(
            cell
            for row in clean["rows"]
            for cell in row["cells"]
            if cell["coordinate"] == "H9"
        )
        self.assertEqual(formula_cell["formula"], "=F9+10")
        self.assertIsNone(formula_cell["cached"])
        self.assertIsNone(formula_cell["literal"])
        self.assertEqual(formula_cell["number_format"], "# ##0")
        # With no stored Excel result, a human-facing clean value must fall back
        # to the formula instead of silently displaying zero or an empty cell.
        self.assertEqual(formula_cell["value"], "=F9+10")

    def test_keeps_business_terms_and_drops_detached_annotation_column(self) -> None:
        table, structure, snapshots = _messy_table_fixture()
        table["range"] = "A1:I10"
        structure["column_descriptors"].append(
            {
                "column": 9,
                "letter": "I",
                "role": "annotation",
                "header_path": ["Comments"],
                "header_sources": ["I3"],
                "key": "comments-deadbeef",
            }
        )
        snapshots[(3, 9)] = _snapshot(3, 9, "Comments")
        snapshots[(6, 2)] = _snapshot(6, 2, "ILON")
        snapshots[(6, 3)] = _snapshot(6, 3, "Complete Incoming Files")
        snapshots[(6, 9)] = _snapshot(
            6,
            9,
            "Please ensure manually entered details correspond to totals",
            merge_anchor="I6",
            merge_ref="I6:I9",
        )
        snapshots[(9, 9)] = _snapshot(
            9,
            9,
            None,
            merge_anchor="I6",
            merge_ref="I6:I9",
        )

        clean = build_clean_table(table, structure, snapshots)

        self.assertEqual([row["source_row"] for row in clean["rows"]], [6, 9])
        self.assertEqual(clean["rows"][0]["cells"][1]["value"], "ILON")
        self.assertEqual(
            clean["rows"][0]["cells"][2]["value"],
            "Complete Incoming Files",
        )
        self.assertNotIn(9, [column["source_column"] for column in clean["columns"]])
        annotation_drop = next(
            item for item in clean["dropped_columns"] if item["source_column"] == 9
        )
        self.assertEqual(annotation_drop["reason"], "annotation")
        self.assertNotIn("Please ensure", repr(clean))

    def test_horizontal_body_merges_do_not_create_duplicate_columns(self) -> None:
        table = {
            "table_id": "tbl_horizontal_body_merge",
            "sheet": "Merged Labels",
            "range": "A1:E3",
            "title": "A1:E3",
        }
        structure = {
            "title": "A1:E3",
            "header_rows": [1],
            "column_descriptors": [
                {
                    "column": column,
                    "letter": get_column_letter(column),
                    "role": (
                        "row_group"
                        if column <= 3
                        else "measure"
                        if column >= 4
                        else "unknown"
                    ),
                    "header_path": ["Metric"] if column <= 3 else [str(2021 + column)],
                    "header_sources": (
                        ["A1"]
                        if column <= 3
                        else [f"{get_column_letter(column)}1"]
                    ),
                    "key": f"column-{column}-deadbeef",
                }
                for column in range(1, 6)
            ],
            "row_descriptors": [
                {
                    "row": 2,
                    "role": "data",
                    "metric_label": "Outstanding",
                    "label_path": ["Outstanding"],
                    "section_path": [],
                },
                {
                    "row": 3,
                    "role": "data",
                    "metric_label": "Production",
                    "label_path": ["Production"],
                    "section_path": [],
                },
            ],
            "sections": [],
        }
        snapshots = {
            (1, 1): _snapshot(1, 1, "Metric", merge_anchor="A1", merge_ref="A1:C1"),
            (1, 2): _snapshot(1, 2, None, merge_anchor="A1", merge_ref="A1:C1"),
            (1, 3): _snapshot(1, 3, None, merge_anchor="A1", merge_ref="A1:C1"),
            (1, 4): _snapshot(1, 4, "2025"),
            (1, 5): _snapshot(1, 5, "2026"),
            (2, 1): _snapshot(2, 1, "Outstanding", merge_anchor="A2", merge_ref="A2:C2"),
            (2, 2): _snapshot(2, 2, None, merge_anchor="A2", merge_ref="A2:C2"),
            (2, 3): _snapshot(2, 3, None, merge_anchor="A2", merge_ref="A2:C2"),
            (2, 4): _snapshot(2, 4, 10),
            (2, 5): _snapshot(2, 5, 11),
            (3, 1): _snapshot(3, 1, "Production", merge_anchor="A3", merge_ref="A3:C3"),
            (3, 2): _snapshot(3, 2, None, merge_anchor="A3", merge_ref="A3:C3"),
            (3, 3): _snapshot(3, 3, None, merge_anchor="A3", merge_ref="A3:C3"),
            (3, 4): _snapshot(3, 4, 20),
            (3, 5): _snapshot(3, 5, 21),
        }

        clean = build_clean_table(table, structure, snapshots)

        self.assertEqual(
            [column["source_column"] for column in clean["columns"]], [1, 4, 5]
        )
        self.assertEqual(
            [cell["value"] for cell in clean["rows"][0]["cells"]],
            ["Outstanding", 10, 11],
        )
        self.assertEqual(
            {
                item["source_column"]
                for item in clean["dropped_columns"]
                if item["reason"] == "body_empty"
            },
            {2, 3},
        )

    def test_banner_only_false_candidate_is_empty(self) -> None:
        table = {
            "table_id": "tbl_banner_only",
            "sheet": "Empty Candidate",
            "range": "A1:H2",
            "title": "Copy table below to PowerPoint",
        }
        structure = {
            "title": "Copy table below to PowerPoint",
            "header_rows": [1],
            "column_descriptors": [
                {
                    "column": column,
                    "letter": get_column_letter(column),
                    "role": "unknown",
                    "header_path": (
                        ["Copy table below to PowerPoint"] if column == 1 else []
                    ),
                    "header_sources": ["A1"] if column == 1 else [],
                    "key": f"column-{column}-deadbeef",
                }
                for column in range(1, 9)
            ],
            "row_descriptors": [
                {"row": 1, "role": "title", "section_path": []},
                {"row": 2, "role": "spacer", "section_path": []},
            ],
            "sections": [],
        }
        snapshots = {
            (1, 1): _snapshot(
                1,
                1,
                "Copy table below to PowerPoint",
                merge_anchor="A1",
                merge_ref="A1:H1",
            )
        }

        clean = build_clean_table(table, structure, snapshots)

        self.assertEqual(clean["status"], "empty")
        self.assertEqual(clean["source_range"], "A1:H2")
        self.assertEqual(clean["rows"], [])
        self.assertEqual(clean["columns"], [])
        self.assertEqual(clean["counts"]["rows"], 0)
        self.assertEqual(clean["counts"]["columns"], 0)


if __name__ == "__main__":
    unittest.main()
