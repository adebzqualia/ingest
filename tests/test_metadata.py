from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from pops_ingest.metadata import (
    defined_name_diagnostics,
    defined_names_metadata,
    static_named_ranges_for_sheet,
)


def _records_by_name(records: list[dict[str, object]]) -> dict[str, dict[str, object]]:
    return {str(record["name"]): record for record in records}


def _records_by_name_and_scope(
    records: list[dict[str, object]],
) -> dict[tuple[str, str], dict[str, object]]:
    return {
        (str(record["name"]), str(record["scope"])): record
        for record in records
    }


class DefinedNameMetadataTests(unittest.TestCase):
    def test_malformed_broken_and_non_range_names_are_preserved_without_evaluation(
        self,
    ) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"

        definitions = {
            "BrokenCell": "#REF!",
            "BrokenArea": "=#REF!:#REF!",
            "ExternalInput": "'[Country.xlsx]Input'!$A$1:$B$2",
            "ConstantThreshold": "42",
            "FormulaDefinition": "=SUM(Data!$A$1:$A$3)",
            "StaticInput": "Data!$A$1:$A$3",
            "LargeStaticInput": "Data!$A$1:$Z$1000",
            "WholeColumns": "Data!$A:$B",
            "_xlnm.Print_Area": "Data!$A$1:$B$2",
        }
        for name, attr_text in definitions.items():
            workbook.defined_names.add(DefinedName(name, attr_text=attr_text))

        # In particular, openpyxl's ``destinations`` iterator may raise a
        # TokenizerError for BrokenArea.  Metadata collection must retain the
        # raw definition rather than aborting the workbook ingestion.
        metadata_records = defined_names_metadata(workbook)
        records = _records_by_name(metadata_records)

        self.assertEqual(set(records), set(definitions))
        for name, attr_text in definitions.items():
            self.assertEqual(records[name]["definition"], attr_text)
            self.assertEqual(records[name]["scope"], "workbook")

        self.assertEqual(records["BrokenCell"]["status"], "broken")
        self.assertEqual(records["BrokenArea"]["status"], "broken")
        self.assertEqual(records["BrokenCell"]["destinations"], [])
        self.assertEqual(records["BrokenArea"]["destinations"], [])
        self.assertEqual(
            records["BrokenArea"]["destination_parse_status"], "failed"
        )
        self.assertEqual(
            records["BrokenArea"]["destination_parse_error"]["type"],
            "TokenizerError",
        )

        self.assertEqual(records["ExternalInput"]["status"], "external_unresolved")
        self.assertNotEqual(records["ExternalInput"]["status"], "static")

        self.assertEqual(records["ConstantThreshold"]["status"], "raw")
        self.assertEqual(records["ConstantThreshold"]["destinations"], [])
        self.assertEqual(records["FormulaDefinition"]["status"], "raw")
        self.assertEqual(records["FormulaDefinition"]["destinations"], [])

        self.assertEqual(records["StaticInput"]["status"], "static")
        self.assertEqual(
            records["StaticInput"]["destinations"],
            [{"sheet": "Data", "range": "$A$1:$A$3"}],
        )
        self.assertEqual(records["WholeColumns"]["status"], "invalid_unparsed")
        self.assertEqual(
            records["WholeColumns"]["destination_parse_status"],
            "invalid_destinations",
        )
        self.assertTrue(records["_xlnm.Print_Area"]["built_in"])

        # Built-ins, non-finite ranges, and ranges above the configured table
        # cap remain metadata but never become table-detector seeds.
        self.assertEqual(
            static_named_ranges_for_sheet(
                metadata_records, "Data", max_area=100
            ),
            [
                {
                    "name": "StaticInput",
                    "ref": "$A$1:$A$3",
                    "scope": "workbook",
                }
            ],
        )

        diagnostics = defined_name_diagnostics(metadata_records)
        self.assertCountEqual(
            [item["defined_name"] for item in diagnostics],
            ["BrokenCell", "BrokenArea", "WholeColumns"],
        )
        broken_area_warning = next(
            item for item in diagnostics if item["defined_name"] == "BrokenArea"
        )
        self.assertEqual(
            broken_area_warning["code"], "DEFINED_NAME_BROKEN_REFERENCE"
        )

    def test_sheet_scoped_names_round_trip_and_remain_distinct(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "sheet_scoped_names.xlsx"
            workbook = Workbook()
            data_sheet = workbook.active
            data_sheet.title = "Data Sheet"
            other_sheet = workbook.create_sheet("Other")

            # The same identifier is valid at workbook scope and independently
            # at each worksheet scope.  The serializer must not collapse these.
            workbook.defined_names.add(
                DefinedName("SharedInput", attr_text="'Data Sheet'!$A$1:$A$2")
            )
            data_sheet.defined_names.add(
                DefinedName("SharedInput", attr_text="'Data Sheet'!$B$1:$B$2")
            )
            other_sheet.defined_names.add(
                DefinedName("SharedInput", attr_text="Other!$C$1:$C$2")
            )

            data_sheet.defined_names.add(
                DefinedName("LocalBroken", attr_text="=#REF!:#REF!")
            )
            data_sheet.defined_names.add(
                DefinedName(
                    "LocalExternal",
                    attr_text="'[Country.xlsx]Input'!$D$1",
                )
            )
            data_sheet.defined_names.add(
                DefinedName(
                    "LocalFormula",
                    attr_text="=SUM('Data Sheet'!$A$1:$A$2)",
                )
            )
            data_sheet.defined_names.add(
                DefinedName("LocalConstant", attr_text="0.95")
            )
            workbook.save(source)

            loaded = load_workbook(source, data_only=False)
            try:
                records = defined_names_metadata(loaded)
            finally:
                loaded.close()

        indexed = _records_by_name_and_scope(records)
        self.assertEqual(
            indexed[("SharedInput", "workbook")]["destinations"],
            [{"sheet": "Data Sheet", "range": "$A$1:$A$2"}],
        )
        self.assertEqual(
            indexed[("SharedInput", "sheet:Data Sheet")]["destinations"],
            [{"sheet": "Data Sheet", "range": "$B$1:$B$2"}],
        )
        self.assertEqual(
            indexed[("SharedInput", "sheet:Other")]["destinations"],
            [{"sheet": "Other", "range": "$C$1:$C$2"}],
        )

        local_broken = indexed[("LocalBroken", "sheet:Data Sheet")]
        self.assertEqual(local_broken["definition"], "=#REF!:#REF!")
        self.assertEqual(local_broken["status"], "broken")
        self.assertEqual(local_broken["destinations"], [])

        local_external = indexed[("LocalExternal", "sheet:Data Sheet")]
        self.assertEqual(local_external["status"], "external_unresolved")
        self.assertEqual(
            local_external["definition"], "'[Country.xlsx]Input'!$D$1"
        )

        local_formula = indexed[("LocalFormula", "sheet:Data Sheet")]
        self.assertEqual(local_formula["status"], "raw")
        self.assertEqual(
            local_formula["definition"], "=SUM('Data Sheet'!$A$1:$A$2)"
        )
        self.assertEqual(local_formula["destinations"], [])

        local_constant = indexed[("LocalConstant", "sheet:Data Sheet")]
        self.assertEqual(local_constant["status"], "raw")
        self.assertEqual(local_constant["definition"], "0.95")

        data_ranges = static_named_ranges_for_sheet(records, "Data Sheet")
        self.assertCountEqual(
            data_ranges,
            [
                {
                    "name": "SharedInput",
                    "ref": "$A$1:$A$2",
                    "scope": "workbook",
                },
                {
                    "name": "SharedInput",
                    "ref": "$B$1:$B$2",
                    "scope": "sheet:Data Sheet",
                },
            ],
        )
        self.assertEqual(
            static_named_ranges_for_sheet(records, "Other"),
            [
                {
                    "name": "SharedInput",
                    "ref": "$C$1:$C$2",
                    "scope": "sheet:Other",
                }
            ],
        )


if __name__ == "__main__":
    unittest.main()
