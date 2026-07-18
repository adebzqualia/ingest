"""Sparse, semantics-preserving worksheet scanning.

The scanner deliberately separates worksheet discovery from table detection.  It
collects enough cell and layout evidence for the detector without walking an
inflated ``max_row * max_column`` rectangle.  Full-column data-validation and
conditional-formatting ranges are retained as interval metadata and are never
expanded into one million synthetic cells.
"""

from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict, dataclass, field, is_dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal
import math
import re
from typing import Any

from openpyxl.formula import Tokenizer
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
    range_boundaries,
)

from .config import ExtractionConfig
from .models import Bounds, CellFeature, WarningRecord
from .utils import encode_typed_value, normalize_whitespace, stable_hash


_EXCEL_MAX_ROW = 1_048_576
_EXCEL_MAX_COLUMN = 16_384
_CELL_REFERENCE_RE = re.compile(
    r"(?<![A-Z0-9_.])(?P<abs_col>\$?)(?P<col>[A-Z]{1,3})"
    r"(?P<abs_row>\$?)(?P<row>[1-9][0-9]*)(?![A-Z0-9_])",
    re.IGNORECASE,
)
_COORDINATE_RE = re.compile(r"^\$?([A-Z]{1,3})\$?([1-9][0-9]*)$", re.IGNORECASE)


def _safe_get(obj: object, name: str, default: Any = None) -> Any:
    """Return an attribute without allowing a malformed OOXML object to fail a scan."""

    try:
        return getattr(obj, name, default)
    except (AttributeError, TypeError, ValueError):
        return default


def _primitive(value: Any, *, _depth: int = 0) -> Any:
    """Convert common openpyxl/dataclass values to deterministic JSON primitives."""

    if _depth > 12:
        return str(value)
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return "NaN"
        if math.isinf(value):
            return "Infinity" if value > 0 else "-Infinity"
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return value.total_seconds()
    if isinstance(value, Decimal):
        return str(value)
    if is_dataclass(value):
        return _primitive(asdict(value), _depth=_depth + 1)
    if isinstance(value, Mapping):
        return {
            str(key): _primitive(item, _depth=_depth + 1)
            for key, item in sorted(value.items(), key=lambda pair: str(pair[0]))
        }
    if isinstance(value, (list, tuple)):
        return [_primitive(item, _depth=_depth + 1) for item in value]
    if isinstance(value, (set, frozenset)):
        items = [_primitive(item, _depth=_depth + 1) for item in value]
        return sorted(items, key=lambda item: repr(item))
    if hasattr(value, "text"):
        text_value = _safe_get(value, "text")
        if text_value is not None:
            return _primitive(text_value, _depth=_depth + 1)
    return str(value)


def _color_dict(color: Any) -> dict[str, Any] | None:
    """Serialize an openpyxl color using semantic fields rather than style IDs."""

    if color is None:
        return None
    color_type = _safe_get(color, "type")
    payload: dict[str, Any] = {"type": color_type}
    # Openpyxl's inactive color descriptors intentionally return diagnostic
    # strings such as "Values must be of type ...". Serializing only the active
    # value avoids leaking those implementation messages into stable hashes.
    active_value = _safe_get(color, "value")
    if active_value is not None:
        payload["value"] = _primitive(active_value)
        if color_type in {"rgb", "indexed", "theme", "auto"}:
            payload[str(color_type)] = _primitive(active_value)
    tint = _safe_get(color, "tint")
    if tint not in (None, 0, 0.0):
        payload["tint"] = _primitive(tint)
    return payload


def _side_dict(side: Any) -> dict[str, Any]:
    """Serialize one border side."""

    if side is None:
        return {"style": None, "color": None}
    return {
        "style": _safe_get(side, "style"),
        "color": _color_dict(_safe_get(side, "color")),
    }


class StyleRegistry:
    """Deduplicate cell styles by a canonical semantic SHA-256 fingerprint.

    Workbook-local ``style_id`` values are intentionally excluded.  Two cells
    with equivalent visual/number/protection semantics therefore receive the
    same hash even if Excel assigned them different internal style records.
    """

    algorithm = "sha256-canonical-json-v1"

    def __init__(self) -> None:
        self._styles: dict[str, dict[str, Any]] = {}

    @staticmethod
    def canonical_style(cell: Any) -> dict[str, Any]:
        """Return the canonical, JSON-compatible semantic style for ``cell``."""

        font = _safe_get(cell, "font")
        fill = _safe_get(cell, "fill")
        border = _safe_get(cell, "border")
        alignment = _safe_get(cell, "alignment")
        protection = _safe_get(cell, "protection")

        font_payload = {
            "name": _safe_get(font, "name"),
            "size": _primitive(_safe_get(font, "sz")),
            "bold": bool(_safe_get(font, "b", False)),
            "italic": bool(_safe_get(font, "i", False)),
            "underline": _safe_get(font, "u"),
            "strike": bool(_safe_get(font, "strike", False)),
            "outline": bool(_safe_get(font, "outline", False)),
            "shadow": bool(_safe_get(font, "shadow", False)),
            "condense": _safe_get(font, "condense"),
            "extend": _safe_get(font, "extend"),
            "vertical_align": _safe_get(font, "vertAlign"),
            "charset": _safe_get(font, "charset"),
            "family": _primitive(_safe_get(font, "family")),
            "scheme": _safe_get(font, "scheme"),
            "color": _color_dict(_safe_get(font, "color")),
        }

        stops: list[dict[str, Any]] = []
        for stop in _safe_get(fill, "stop", ()) or ():
            stops.append(
                {
                    "position": _primitive(_safe_get(stop, "position")),
                    "color": _color_dict(_safe_get(stop, "color")),
                }
            )
        fill_payload = {
            "fill_type": _safe_get(fill, "fill_type"),
            "pattern_type": _safe_get(fill, "patternType"),
            "foreground": _color_dict(_safe_get(fill, "fgColor")),
            "background": _color_dict(_safe_get(fill, "bgColor")),
            "gradient_type": _safe_get(fill, "type"),
            "degree": _primitive(_safe_get(fill, "degree")),
            "left": _primitive(_safe_get(fill, "left")),
            "right": _primitive(_safe_get(fill, "right")),
            "top": _primitive(_safe_get(fill, "top")),
            "bottom": _primitive(_safe_get(fill, "bottom")),
            "stops": stops,
        }

        border_payload = {
            name: _side_dict(_safe_get(border, name))
            for name in (
                "left",
                "right",
                "top",
                "bottom",
                "diagonal",
                "vertical",
                "horizontal",
                "start",
                "end",
            )
        }
        border_payload.update(
            {
                "diagonal_up": bool(_safe_get(border, "diagonalUp", False)),
                "diagonal_down": bool(_safe_get(border, "diagonalDown", False)),
                "outline": bool(_safe_get(border, "outline", False)),
            }
        )

        alignment_payload = {
            "horizontal": _safe_get(alignment, "horizontal"),
            "vertical": _safe_get(alignment, "vertical"),
            "text_rotation": _primitive(_safe_get(alignment, "textRotation")),
            "wrap_text": _safe_get(alignment, "wrapText"),
            "shrink_to_fit": _safe_get(alignment, "shrinkToFit"),
            "indent": _primitive(_safe_get(alignment, "indent")),
            "relative_indent": _primitive(_safe_get(alignment, "relativeIndent")),
            "justify_last_line": _safe_get(alignment, "justifyLastLine"),
            "reading_order": _primitive(_safe_get(alignment, "readingOrder")),
        }

        return {
            "font": font_payload,
            "fill": fill_payload,
            "border": border_payload,
            "alignment": alignment_payload,
            "number_format": _safe_get(cell, "number_format", "General"),
            "protection": {
                "locked": _safe_get(protection, "locked"),
                "hidden": _safe_get(protection, "hidden"),
            },
            "quote_prefix": bool(_safe_get(cell, "quotePrefix", False)),
            "pivot_button": bool(_safe_get(cell, "pivotButton", False)),
        }

    def fingerprint(self, cell: Any) -> str:
        """Return the semantic style fingerprint without relying on ``style_id``."""

        return stable_hash(self.canonical_style(cell))

    def register(self, cell: Any) -> str:
        """Register ``cell``'s semantic style and return its stable hash."""

        details = self.canonical_style(cell)
        style_hash = stable_hash(details)
        self._styles.setdefault(style_hash, details)
        return style_hash

    def details_for(self, style_hash: str) -> dict[str, Any] | None:
        """Return canonical style details for ``style_hash`` if registered."""

        return self._styles.get(style_hash)

    def to_dict(self) -> dict[str, Any]:
        """Return a deterministically ordered registry suitable for JSON output."""

        return {
            "algorithm": self.algorithm,
            "styles": {
                style_hash: self._styles[style_hash]
                for style_hash in sorted(self._styles)
            },
        }

    def __len__(self) -> int:
        return len(self._styles)


@dataclass(slots=True)
class SheetProfile:
    """Sparse worksheet facts and structural evidence used by later stages."""

    sheet_name: str
    sheet_state: str
    features: dict[tuple[int, int], CellFeature] = field(default_factory=dict)
    material_coordinates: set[tuple[int, int]] = field(default_factory=set)
    content_bounds: Bounds | None = None
    style_bounds: Bounds | None = None
    evidence_bounds: Bounds | None = None
    merge_index: dict[tuple[int, int], dict[str, Any]] = field(default_factory=dict)
    merged_ranges: list[dict[str, Any]] = field(default_factory=list)
    validations: list[dict[str, Any]] = field(default_factory=list)
    conditional_formats: list[dict[str, Any]] = field(default_factory=list)
    explicit_tables: list[dict[str, Any]] = field(default_factory=list)
    row_dimensions: dict[int, dict[str, Any]] = field(default_factory=dict)
    column_dimensions: dict[str, dict[str, Any]] = field(default_factory=dict)
    sheet_format: dict[str, Any] = field(default_factory=dict)
    counts: dict[str, Any] = field(default_factory=dict)
    warnings: list[WarningRecord] = field(default_factory=list)
    config_fingerprint: str | None = None
    include_cached_values: bool = True
    validation_bounds: dict[str, tuple[Bounds, ...]] = field(
        default_factory=dict, repr=False
    )
    conditional_format_bounds: dict[str, tuple[Bounds, ...]] = field(
        default_factory=dict, repr=False
    )
    column_dimension_index: dict[int, dict[str, Any]] = field(
        default_factory=dict, repr=False
    )
    formula_record_index: dict[str, Any] = field(default_factory=dict, repr=False)
    forced_evidence_coordinates: set[tuple[int, int]] = field(
        default_factory=set, repr=False
    )

    @property
    def title(self) -> str:
        """Alias used by callers that think in openpyxl worksheet terms."""

        return self.sheet_name

    def feature_at(self, row: int, col: int) -> CellFeature | None:
        """Return a sparse feature record without instantiating a worksheet cell."""

        return self.features.get((row, col))

    def merge_at(self, row: int, col: int) -> dict[str, Any] | None:
        """Return merge metadata, falling back to non-expanded large ranges."""

        direct = self.merge_index.get((row, col))
        if direct is not None:
            return direct
        for item in self.merged_ranges:
            bounds = item.get("_bounds")
            if isinstance(bounds, Bounds) and bounds.contains(row, col):
                return {
                    "range": item["range"],
                    "anchor": item["anchor"],
                    "is_anchor": row == bounds.min_row and col == bounds.min_col,
                }
        return None

    def validation_memberships(self, row: int, col: int) -> list[dict[str, Any]]:
        """Return validations applying to a coordinate without expanding ranges."""

        matching_ids = {
            item_id
            for item_id, bounds_list in self.validation_bounds.items()
            if any(bounds.contains(row, col) for bounds in bounds_list)
        }
        return [item for item in self.validations if item["id"] in matching_ids]

    def conditional_format_memberships(
        self, row: int, col: int
    ) -> list[dict[str, Any]]:
        """Return conditional formats applying to a coordinate."""

        matching_ids = {
            item_id
            for item_id, bounds_list in self.conditional_format_bounds.items()
            if any(bounds.contains(row, col) for bounds in bounds_list)
        }
        return [
            item for item in self.conditional_formats if item["id"] in matching_ids
        ]

    def to_dict(self) -> dict[str, Any]:
        """Serialize profile-level metadata; cell features remain a sparse list."""

        return {
            "sheet_name": self.sheet_name,
            "sheet_state": self.sheet_state,
            "content_bounds": self.content_bounds.ref if self.content_bounds else None,
            "style_bounds": self.style_bounds.ref if self.style_bounds else None,
            "evidence_bounds": self.evidence_bounds.ref if self.evidence_bounds else None,
            "material_coordinates": [
                f"{get_column_letter(col)}{row}"
                for row, col in sorted(self.material_coordinates)
            ],
            "features": [
                asdict(self.features[key]) for key in sorted(self.features)
            ],
            "merged_ranges": [
                {key: value for key, value in item.items() if key != "_bounds"}
                for item in self.merged_ranges
            ],
            "validations": self.validations,
            "conditional_formats": self.conditional_formats,
            "explicit_tables": self.explicit_tables,
            "row_dimensions": {
                str(key): value for key, value in sorted(self.row_dimensions.items())
            },
            "column_dimensions": {
                key: self.column_dimensions[key]
                for key in sorted(self.column_dimensions)
            },
            "sheet_format": self.sheet_format,
            "counts": self.counts,
            "warnings": [warning.to_dict() for warning in self.warnings],
            "config_fingerprint": self.config_fingerprint,
        }


def _bounds_from_ref(ref: str) -> Bounds:
    """Parse normal, whole-column, or whole-row A1 ranges into bounded geometry."""

    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return Bounds(
        int(min_row or 1),
        int(min_col or 1),
        int(max_row or _EXCEL_MAX_ROW),
        int(max_col or _EXCEL_MAX_COLUMN),
    )


def _range_refs(sqref: Any) -> list[str]:
    """Return deterministic individual refs from a MultiCellRange-like value."""

    ranges = _safe_get(sqref, "ranges")
    if ranges is not None:
        return sorted(
            (str(item) for item in ranges),
            key=lambda ref: (
                _bounds_from_ref(ref).min_row,
                _bounds_from_ref(ref).min_col,
                _bounds_from_ref(ref).max_row,
                _bounds_from_ref(ref).max_col,
            ),
        )
    return [token for token in str(sqref or "").split() if token]


def _bounds_for_refs(
    refs: Iterable[str],
    *,
    warnings: list[WarningRecord],
    warning_code: str,
    sheet_name: str,
) -> tuple[Bounds, ...]:
    """Parse a collection of refs while recording malformed OOXML ranges."""

    result: list[Bounds] = []
    for ref in refs:
        try:
            result.append(_bounds_from_ref(ref))
        except (TypeError, ValueError):
            warnings.append(
                WarningRecord(
                    code=warning_code,
                    message=f"Could not parse worksheet range {ref!r}",
                    sheet=sheet_name,
                    details={"range": ref},
                )
            )
    return tuple(result)


def _formula_object_text(value: Any) -> str | None:
    """Extract formula text from normal, array, and data-table formula objects."""

    if isinstance(value, str):
        return value
    text_value = _safe_get(value, "text")
    if text_value is not None:
        return str(text_value)
    return None


def _is_formula_cell(cell: Any) -> bool:
    """Return whether openpyxl identifies ``cell`` as a formula cell."""

    return _safe_get(cell, "data_type") == "f" or (
        not isinstance(_safe_get(cell, "value"), (str, bytes))
        and _formula_object_text(_safe_get(cell, "value")) is not None
    )


def _record_mapping(record: Any) -> Mapping[str, Any]:
    if isinstance(record, Mapping):
        return record
    if is_dataclass(record):
        value = asdict(record)
        return value if isinstance(value, Mapping) else {}
    data = _safe_get(record, "__dict__", {})
    return data if isinstance(data, Mapping) else {}


def _record_value(record: Any, names: Sequence[str]) -> Any:
    mapping = _record_mapping(record)
    for name in names:
        if name in mapping and mapping[name] is not None:
            return mapping[name]
        value = _safe_get(record, name)
        if value is not None:
            return value
    return None


def _normalize_coordinate(value: Any) -> str | None:
    """Normalize an A1 coordinate, accepting an optional sheet qualifier."""

    if value is None:
        return None
    text = str(value).strip()
    if "!" in text:
        text = text.rsplit("!", 1)[1]
    text = text.strip("'")
    match = _COORDINATE_RE.fullmatch(text)
    if not match:
        return None
    return f"{match.group(1).upper()}{int(match.group(2))}"


def _record_sheet(record: Any) -> str | None:
    value = _record_value(record, ("sheet", "sheet_name", "worksheet", "title"))
    return str(value) if value is not None else None


def _record_coordinate(record: Any) -> str | None:
    return _normalize_coordinate(
        _record_value(record, ("coordinate", "cell", "cell_ref", "ref", "address"))
    )


def _index_raw_formula_records(records: Any, sheet_name: str) -> dict[str, Any]:
    """Index several common raw-formula-record layouts by local coordinate."""

    result: dict[str, Any] = {}
    if records is None:
        return result

    def add(record: Any, coordinate_hint: Any = None, sheet_hint: Any = None) -> None:
        record_sheet = _record_sheet(record) or (
            str(sheet_hint) if sheet_hint is not None else None
        )
        if record_sheet is not None and record_sheet != sheet_name:
            return
        coordinate = _record_coordinate(record) or _normalize_coordinate(coordinate_hint)
        if coordinate is not None:
            result[coordinate] = record

    if isinstance(records, Mapping):
        nested = records.get(sheet_name)
        if isinstance(nested, Mapping):
            for key, record in nested.items():
                add(record, key, sheet_name)
        elif isinstance(nested, Sequence) and not isinstance(nested, (str, bytes)):
            for record in nested:
                add(record, sheet_hint=sheet_name)

        for key, record in records.items():
            if key == sheet_name:
                continue
            if isinstance(key, tuple) and len(key) >= 2:
                add(record, key[1], key[0])
                continue
            key_text = str(key)
            if "!" in key_text:
                key_sheet, key_coordinate = key_text.rsplit("!", 1)
                add(record, key_coordinate, key_sheet.strip("'"))
            else:
                add(record, key)
        return result

    if isinstance(records, Sequence) and not isinstance(records, (str, bytes)):
        for record in records:
            add(record)
    else:
        add(records)
    return result


def _dependencies_from_formula(
    formula: str | None,
) -> tuple[list[Any], dict[str, str] | None]:
    """Extract raw range/name operands without attempting workbook evaluation."""

    if not formula:
        return [], None
    expression = formula if formula.startswith("=") else "=" + formula
    try:
        tokens = Tokenizer(expression).items
    except Exception as exc:  # noqa: BLE001 - untrusted workbook parser boundary
        return [], {"type": type(exc).__name__, "message": str(exc)}
    result: list[str] = []
    for token in tokens:
        if token.type == "OPERAND" and token.subtype == "RANGE":
            dependency = str(token.value)
            if dependency not in result:
                result.append(dependency)
    return result, None


def _relative_reference_signature(
    formula: str | None, origin_row: int, origin_col: int
) -> tuple[str | None, dict[str, str] | None]:
    """Normalize A1 operands to R1C1-like relative offsets.

    Tokenization prevents strings such as ``"A1"`` from being rewritten.  Named
    ranges and whole-column/whole-row references are retained verbatim.
    """

    if not formula:
        return None, None
    expression = formula if formula.startswith("=") else "=" + formula
    try:
        tokens = Tokenizer(expression).items
    except Exception as exc:  # noqa: BLE001 - untrusted workbook parser boundary
        return normalize_whitespace(expression), {
            "type": type(exc).__name__,
            "message": str(exc),
        }

    def replace_reference(match: re.Match[str]) -> str:
        column = column_index_from_string(match.group("col"))
        row = int(match.group("row"))
        if match.group("abs_row"):
            row_part = f"R{row}"
        else:
            row_part = f"R[{row - origin_row}]"
        if match.group("abs_col"):
            column_part = f"C{column}"
        else:
            column_part = f"C[{column - origin_col}]"
        return row_part + column_part

    output: list[str] = []
    for token in tokens:
        value = str(token.value)
        if token.type == "OPERAND" and token.subtype == "RANGE":
            if "!" in value:
                prefix, local_ref = value.rsplit("!", 1)
                value = prefix + "!" + _CELL_REFERENCE_RE.sub(
                    replace_reference, local_ref
                )
            else:
                value = _CELL_REFERENCE_RE.sub(replace_reference, value)
        output.append(value)
    return "=" + "".join(output), None


def _formula_info(cell: Any, record: Any, row: int, col: int) -> dict[str, Any] | None:
    """Reconcile openpyxl's effective formula with raw OOXML formula metadata."""

    cell_formula = _formula_object_text(_safe_get(cell, "value")) if _is_formula_cell(cell) else None
    raw_formula = _record_value(
        record,
        ("raw_formula", "formula_raw", "raw", "xml_formula", "formula_text"),
    )
    resolved_formula = _record_value(
        record,
        (
            "resolved_formula",
            "formula_resolved",
            "resolved",
            "translated_formula",
            "effective_formula",
        ),
    )
    exact_formula = _record_value(record, ("exact_formula", "formula_exact", "exact"))
    generic_formula = _record_value(record, ("formula",))

    if cell_formula is None and all(
        value is None
        for value in (raw_formula, resolved_formula, exact_formula, generic_formula)
    ):
        return None

    exact = cell_formula or exact_formula or resolved_formula or generic_formula or raw_formula
    raw = raw_formula if raw_formula is not None else generic_formula or exact
    resolved = resolved_formula if resolved_formula is not None else exact
    exact_text = str(exact) if exact is not None else None
    raw_text = str(raw) if raw is not None else None
    resolved_text = str(resolved) if resolved is not None else exact_text

    signature = _record_value(
        record,
        (
            "normalized_relative_signature",
            "relative_signature",
            "normalized_signature",
            "signature",
        ),
    )
    tokenization_attempted = False
    tokenization_error: dict[str, str] | None = None
    if signature is None:
        tokenization_attempted = True
        signature, tokenization_error = _relative_reference_signature(
            resolved_text, row, col
        )

    dependencies = _record_value(record, ("dependencies", "references", "refs"))
    if dependencies is None:
        tokenization_attempted = True
        dependencies, dependency_error = _dependencies_from_formula(resolved_text)
        if tokenization_error is None:
            tokenization_error = dependency_error

    formula_kind = _record_value(record, ("kind", "formula_kind", "formula_type"))
    if formula_kind is None:
        value = _safe_get(cell, "value")
        formula_kind = "cell" if isinstance(value, str) else type(value).__name__

    metadata = _primitive(record) if record is not None else None
    return {
        "exact": exact_text,
        "raw": raw_text,
        "resolved": resolved_text,
        "kind": str(formula_kind),
        "normalized_relative_signature": _primitive(signature),
        "dependencies": _primitive(dependencies),
        "tokenization_status": (
            "failed"
            if tokenization_error
            else "parsed"
            if tokenization_attempted
            else "provided"
        ),
        "tokenization_error": tokenization_error,
        "raw_record": metadata,
    }


def _typed_cell_value(value: Any, data_type: str | None) -> tuple[Any, str]:
    """Encode a cell value while retaining Excel error and rich-text semantics."""

    encoded, kind = encode_typed_value(value)
    if data_type == "e" and value is not None:
        return encoded, "error"
    if data_type in {"s", "str", "inlineStr"} and value is not None:
        return str(value), "text"
    return encoded, kind


def _row_dimension_metadata(ws: Any) -> dict[int, dict[str, Any]]:
    result: dict[int, dict[str, Any]] = {}
    for key, dimension in sorted(ws.row_dimensions.items(), key=lambda item: int(item[0])):
        row = int(key)
        result[row] = {
            "row": row,
            "height": _primitive(_safe_get(dimension, "height")),
            "hidden": bool(_safe_get(dimension, "hidden", False)),
            "outline_level": int(_safe_get(dimension, "outlineLevel", 0) or 0),
            "collapsed": bool(_safe_get(dimension, "collapsed", False)),
            "thick_top": bool(_safe_get(dimension, "thickTop", False)),
            "thick_bottom": bool(_safe_get(dimension, "thickBottom", False)),
            "custom_format": bool(_safe_get(dimension, "customFormat", False)),
            "custom_height": bool(_safe_get(dimension, "customHeight", False)),
            "style_id": int(_safe_get(dimension, "style_id", 0) or 0),
        }
    return result


def _column_dimension_metadata(
    ws: Any,
) -> tuple[dict[str, dict[str, Any]], dict[int, dict[str, Any]]]:
    result: dict[str, dict[str, Any]] = {}
    index: dict[int, dict[str, Any]] = {}
    for key, dimension in sorted(ws.column_dimensions.items(), key=lambda item: str(item[0])):
        min_col = int(_safe_get(dimension, "min", 0) or 0)
        max_col = int(_safe_get(dimension, "max", 0) or 0)
        if min_col < 1:
            try:
                min_col = column_index_from_string(str(key))
            except ValueError:
                continue
        if max_col < min_col:
            max_col = min_col
        metadata = {
            "index": str(key),
            "min_column": min_col,
            "max_column": max_col,
            "width": _primitive(_safe_get(dimension, "width")),
            "hidden": bool(_safe_get(dimension, "hidden", False)),
            "outline_level": int(_safe_get(dimension, "outlineLevel", 0) or 0),
            "collapsed": bool(_safe_get(dimension, "collapsed", False)),
            "best_fit": bool(_safe_get(dimension, "bestFit", False)),
            "custom_width": bool(_safe_get(dimension, "customWidth", False)),
            "custom_format": bool(_safe_get(dimension, "customFormat", False)),
            "style_id": int(_safe_get(dimension, "style_id", 0) or 0),
        }
        result[str(key)] = metadata
        for col in range(min_col, min(max_col, _EXCEL_MAX_COLUMN) + 1):
            index[col] = metadata
    return result, index


def _sheet_format_metadata(ws: Any) -> dict[str, Any]:
    properties = _safe_get(ws, "sheet_format")
    return {
        "base_column_width": _primitive(_safe_get(properties, "baseColWidth")),
        "default_column_width": _primitive(_safe_get(properties, "defaultColWidth")),
        "default_row_height": _primitive(_safe_get(properties, "defaultRowHeight")),
        "custom_height": bool(_safe_get(properties, "customHeight", False)),
        "zero_height": bool(_safe_get(properties, "zeroHeight", False)),
        "outline_level_row": int(_safe_get(properties, "outlineLevelRow", 0) or 0),
        "outline_level_column": int(
            _safe_get(properties, "outlineLevelCol", 0) or 0
        ),
    }


def _serialize_validations(
    ws: Any, warnings: list[WarningRecord]
) -> tuple[list[dict[str, Any]], dict[str, tuple[Bounds, ...]]]:
    records: list[dict[str, Any]] = []
    bounds_index: dict[str, tuple[Bounds, ...]] = {}
    container = _safe_get(ws, "data_validations")
    items = _safe_get(container, "dataValidation", ()) or ()
    for position, validation in enumerate(items, start=1):
        item_id = f"dv{position:04d}"
        refs = _range_refs(_safe_get(validation, "sqref"))
        bounds_index[item_id] = _bounds_for_refs(
            refs,
            warnings=warnings,
            warning_code="INVALID_DATA_VALIDATION_RANGE",
            sheet_name=ws.title,
        )
        records.append(
            {
                "id": item_id,
                "sqref": " ".join(refs),
                "ranges": refs,
                "type": _safe_get(validation, "type"),
                "operator": _safe_get(validation, "operator"),
                "formula1": _primitive(_safe_get(validation, "formula1")),
                "formula2": _primitive(_safe_get(validation, "formula2")),
                "allow_blank": _safe_get(validation, "allowBlank"),
                "show_drop_down": _safe_get(validation, "showDropDown"),
                "show_input_message": _safe_get(validation, "showInputMessage"),
                "show_error_message": _safe_get(validation, "showErrorMessage"),
                "error_style": _safe_get(validation, "errorStyle"),
                "error_title": _safe_get(validation, "errorTitle"),
                "error": _safe_get(validation, "error"),
                "prompt_title": _safe_get(validation, "promptTitle"),
                "prompt": _safe_get(validation, "prompt"),
                "ime_mode": _safe_get(validation, "imeMode"),
            }
        )
    return records, bounds_index


def _serialize_cf_rule(rule: Any) -> dict[str, Any]:
    """Serialize stable conditional-formatting rule fields."""

    payload = {
        "type": _safe_get(rule, "type"),
        "priority": _primitive(_safe_get(rule, "priority")),
        "stop_if_true": _safe_get(rule, "stopIfTrue"),
        "operator": _safe_get(rule, "operator"),
        "formula": _primitive(_safe_get(rule, "formula")),
        "text": _safe_get(rule, "text"),
        "time_period": _safe_get(rule, "timePeriod"),
        "rank": _primitive(_safe_get(rule, "rank")),
        "percent": _safe_get(rule, "percent"),
        "bottom": _safe_get(rule, "bottom"),
        "equal_average": _safe_get(rule, "equalAverage"),
        "above_average": _safe_get(rule, "aboveAverage"),
        "dxf_id": _primitive(_safe_get(rule, "dxfId")),
    }
    for name, output_name in (
        ("colorScale", "color_scale"),
        ("dataBar", "data_bar"),
        ("iconSet", "icon_set"),
    ):
        value = _safe_get(rule, name)
        if value is not None:
            payload[output_name] = str(value)
    return payload


def _serialize_conditional_formats(
    ws: Any, warnings: list[WarningRecord]
) -> tuple[list[dict[str, Any]], dict[str, tuple[Bounds, ...]]]:
    records: list[dict[str, Any]] = []
    bounds_index: dict[str, tuple[Bounds, ...]] = {}
    collection = _safe_get(ws, "conditional_formatting")
    if collection is None:
        return records, bounds_index
    try:
        items = list(collection)
    except (AttributeError, TypeError, ValueError):
        items = []
    for position, conditional_format in enumerate(items, start=1):
        item_id = f"cf{position:04d}"
        refs = _range_refs(_safe_get(conditional_format, "sqref"))
        bounds_index[item_id] = _bounds_for_refs(
            refs,
            warnings=warnings,
            warning_code="INVALID_CONDITIONAL_FORMAT_RANGE",
            sheet_name=ws.title,
        )
        rules = [_serialize_cf_rule(rule) for rule in _safe_get(conditional_format, "rules", ()) or ()]
        records.append(
            {
                "id": item_id,
                "sqref": " ".join(refs),
                "ranges": refs,
                "pivot": bool(_safe_get(conditional_format, "pivot", False)),
                "rules": rules,
            }
        )
    return records, bounds_index


def _table_formula(value: Any) -> str | None:
    if value is None:
        return None
    text_value = _safe_get(value, "text")
    return str(text_value if text_value is not None else value)


def _serialize_explicit_tables(
    ws: Any, warnings: list[WarningRecord]
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    tables = _safe_get(ws, "tables")
    if tables is None:
        return records
    try:
        table_values = list(tables.values())
    except (AttributeError, TypeError, ValueError):
        table_values = []
    for position, table in enumerate(table_values, start=1):
        ref = str(_safe_get(table, "ref", ""))
        try:
            bounds = _bounds_from_ref(ref)
        except (TypeError, ValueError):
            warnings.append(
                WarningRecord(
                    code="INVALID_NATIVE_TABLE_RANGE",
                    message=f"Could not parse native table range {ref!r}",
                    sheet=ws.title,
                    details={"table": _safe_get(table, "displayName"), "range": ref},
                )
            )
            continue

        columns: list[dict[str, Any]] = []
        for table_column in _safe_get(table, "tableColumns", ()) or ():
            columns.append(
                {
                    "id": _primitive(_safe_get(table_column, "id")),
                    "name": _safe_get(table_column, "name"),
                    "unique_name": _safe_get(table_column, "uniqueName"),
                    "totals_row_label": _safe_get(table_column, "totalsRowLabel"),
                    "totals_row_function": _safe_get(
                        table_column, "totalsRowFunction"
                    ),
                    "calculated_column_formula": _table_formula(
                        _safe_get(table_column, "calculatedColumnFormula")
                    ),
                    "totals_row_formula": _table_formula(
                        _safe_get(table_column, "totalsRowFormula")
                    ),
                    "data_dxf_id": _primitive(_safe_get(table_column, "dataDxfId")),
                    "header_row_dxf_id": _primitive(
                        _safe_get(table_column, "headerRowDxfId")
                    ),
                    "totals_row_dxf_id": _primitive(
                        _safe_get(table_column, "totalsRowDxfId")
                    ),
                    "query_table_field_id": _primitive(
                        _safe_get(table_column, "queryTableFieldId")
                    ),
                }
            )

        style = _safe_get(table, "tableStyleInfo")
        auto_filter = _safe_get(table, "autoFilter")
        records.append(
            {
                "id": f"table{position:04d}",
                "name": _safe_get(table, "name"),
                "display_name": _safe_get(table, "displayName"),
                "range": bounds.ref,
                "table_type": _safe_get(table, "tableType"),
                "header_row_count": _primitive(_safe_get(table, "headerRowCount")),
                "totals_row_count": _primitive(_safe_get(table, "totalsRowCount")),
                "totals_row_shown": _safe_get(table, "totalsRowShown"),
                "insert_row": _safe_get(table, "insertRow"),
                "published": _safe_get(table, "published"),
                "comment": _safe_get(table, "comment"),
                "connection_id": _primitive(_safe_get(table, "connectionId")),
                "auto_filter": {
                    "range": _safe_get(auto_filter, "ref"),
                    "filter_column_count": len(
                        _safe_get(auto_filter, "filterColumn", ()) or ()
                    ),
                    "sort_state": _primitive(_safe_get(auto_filter, "sortState")),
                }
                if auto_filter is not None
                else None,
                "style": {
                    "name": _safe_get(style, "name"),
                    "show_first_column": _safe_get(style, "showFirstColumn"),
                    "show_last_column": _safe_get(style, "showLastColumn"),
                    "show_row_stripes": _safe_get(style, "showRowStripes"),
                    "show_column_stripes": _safe_get(style, "showColumnStripes"),
                }
                if style is not None
                else None,
                "columns": columns,
            }
        )
    return sorted(records, key=lambda item: (_bounds_from_ref(item["range"]), item["id"]))


def _bounds_from_coordinates(coordinates: Iterable[tuple[int, int]]) -> Bounds | None:
    iterator = iter(coordinates)
    try:
        first_row, first_col = next(iterator)
    except StopIteration:
        return None
    min_row = max_row = first_row
    min_col = max_col = first_col
    for row, col in iterator:
        min_row = min(min_row, row)
        min_col = min(min_col, col)
        max_row = max(max_row, row)
        max_col = max(max_col, col)
    return Bounds(min_row, min_col, max_row, max_col)


def _iter_bounds_coordinates(bounds: Bounds) -> Iterable[tuple[int, int]]:
    for row in range(bounds.min_row, bounds.max_row + 1):
        for col in range(bounds.min_col, bounds.max_col + 1):
            yield row, col


def _is_full_axis_range(bounds: Bounds) -> bool:
    return bounds.height == _EXCEL_MAX_ROW or bounds.width == _EXCEL_MAX_COLUMN


def _instantiated_coordinates(ws: Any) -> set[tuple[int, int]]:
    """Return coordinates already represented in worksheet XML/openpyxl state."""

    cells = _safe_get(ws, "_cells")
    if not isinstance(cells, Mapping):
        raise TypeError(
            "scan_sheet requires a normal openpyxl Worksheet; read-only worksheets "
            "do not expose the semantic cell/style/comment model"
        )
    return {
        (int(row), int(col))
        for row, col in cells.keys()
        if int(row) >= 1 and int(col) >= 1
    }


def _merge_metadata(
    ws: Any,
    material_coordinates: set[tuple[int, int]],
    config: ExtractionConfig,
    warnings: list[WarningRecord],
) -> tuple[dict[tuple[int, int], dict[str, Any]], list[dict[str, Any]]]:
    merge_index: dict[tuple[int, int], dict[str, Any]] = {}
    merged_ranges: list[dict[str, Any]] = []
    remaining = max(0, config.max_material_cells_per_sheet - len(material_coordinates))
    for merged in sorted(
        ws.merged_cells.ranges,
        key=lambda item: (item.min_row, item.min_col, item.max_row, item.max_col),
    ):
        bounds = Bounds(merged.min_row, merged.min_col, merged.max_row, merged.max_col)
        anchor = f"{get_column_letter(bounds.min_col)}{bounds.min_row}"
        merged_ranges.append(
            {"range": bounds.ref, "anchor": anchor, "area": bounds.area, "_bounds": bounds}
        )
        # Normal POPS merges are small.  A pathological full-sheet merge remains
        # queryable through SheetProfile.merge_at without materializing it.
        if bounds.area <= remaining and bounds.area <= 100_000:
            for row, col in _iter_bounds_coordinates(bounds):
                material_coordinates.add((row, col))
                merge_index[(row, col)] = {
                    "range": bounds.ref,
                    "anchor": anchor,
                    "is_anchor": row == bounds.min_row and col == bounds.min_col,
                }
            remaining = max(0, config.max_material_cells_per_sheet - len(material_coordinates))
        else:
            merge_index[(bounds.min_row, bounds.min_col)] = {
                "range": bounds.ref,
                "anchor": anchor,
                "is_anchor": True,
            }
            material_coordinates.add((bounds.min_row, bounds.min_col))
            warnings.append(
                WarningRecord(
                    code="MERGE_RANGE_NOT_EXPANDED",
                    message="Large merged range was retained as interval metadata",
                    sheet=ws.title,
                    coordinate=anchor,
                    details={"range": bounds.ref, "area": bounds.area},
                )
            )
    return merge_index, merged_ranges


def _bounded_membership_materialization(
    bounds_indexes: Iterable[dict[str, tuple[Bounds, ...]]],
    material_coordinates: set[tuple[int, int]],
    forced_evidence: set[tuple[int, int]],
    config: ExtractionConfig,
) -> None:
    """Materialize only small finite validation/CF regions for blank input cells."""

    per_range_limit = min(50_000, max(1_000, config.max_material_cells_per_sheet // 20))
    for bounds_index in bounds_indexes:
        for bounds_list in bounds_index.values():
            for bounds in bounds_list:
                if _is_full_axis_range(bounds) or bounds.area > per_range_limit:
                    continue
                missing_budget = config.max_material_cells_per_sheet - len(
                    material_coordinates
                )
                if bounds.area > missing_budget:
                    continue
                for coordinate in _iter_bounds_coordinates(bounds):
                    material_coordinates.add(coordinate)
                    forced_evidence.add(coordinate)


def _cell_for_coordinate(ws: Any, row: int, col: int) -> Any:
    cells = _safe_get(ws, "_cells", {})
    if isinstance(cells, Mapping) and (row, col) in cells:
        return cells[(row, col)]
    return ws.cell(row=row, column=col)


def _comment_dict(comment: Any) -> dict[str, Any] | None:
    if comment is None:
        return None
    return {
        "text": str(_safe_get(comment, "text", "")),
        "author": _safe_get(comment, "author"),
        "width": _primitive(_safe_get(comment, "width")),
        "height": _primitive(_safe_get(comment, "height")),
    }


def _hyperlink_dict(hyperlink: Any) -> dict[str, Any] | None:
    if hyperlink is None:
        return None
    if isinstance(hyperlink, str):
        return {"target": hyperlink, "location": None, "display": None, "tooltip": None}
    return {
        "target": _safe_get(hyperlink, "target"),
        "location": _safe_get(hyperlink, "location"),
        "display": _safe_get(hyperlink, "display"),
        "tooltip": _safe_get(hyperlink, "tooltip"),
        "relationship_id": _safe_get(hyperlink, "id"),
    }


def cell_snapshot(
    ws: Any,
    cached_ws: Any,
    row: int,
    col: int,
    profile: SheetProfile,
    raw_formula_records: Any,
    style_registry: StyleRegistry,
) -> dict[str, Any]:
    """Return a loss-aware snapshot for one worksheet coordinate.

    Literal values, formula text/metadata, and workbook-cached formula results
    are deliberately separate.  The function never evaluates a formula.
    """

    cell = _cell_for_coordinate(ws, row, col)
    coordinate = f"{get_column_letter(col)}{row}"
    record = profile.formula_record_index.get(coordinate)
    if record is None and raw_formula_records is not None:
        record = _index_raw_formula_records(raw_formula_records, ws.title).get(coordinate)
    formula = _formula_info(cell, record, row, col)

    cell_value = _safe_get(cell, "value")
    cell_data_type = _safe_get(cell, "data_type")
    if formula is None:
        literal_value, literal_type = _typed_cell_value(cell_value, cell_data_type)
    else:
        literal_value, literal_type = None, "blank"

    cached_value: Any = None
    cached_type = "blank"
    cache_status = "not_applicable"
    if formula is not None:
        if not profile.include_cached_values:
            cache_status = "disabled"
        elif cached_ws is None:
            cache_status = "not_loaded"
        else:
            cached_cell = _cell_for_coordinate(cached_ws, row, col)
            if _is_formula_cell(cached_cell):
                cache_status = "not_data_only"
            else:
                cached_value, cached_type = _typed_cell_value(
                    _safe_get(cached_cell, "value"),
                    _safe_get(cached_cell, "data_type"),
                )
                cache_status = "present" if cached_value is not None else "missing"

    style_hash = style_registry.register(cell)
    style_details = style_registry.details_for(style_hash)
    merge = profile.merge_at(row, col)
    row_dimension = profile.row_dimensions.get(row)
    column_dimension = profile.column_dimension_index.get(col)
    row_hidden = bool(row_dimension and row_dimension.get("hidden"))
    column_hidden = bool(column_dimension and column_dimension.get("hidden"))
    validations = profile.validation_memberships(row, col)
    conditional_formats = profile.conditional_format_memberships(row, col)
    protection = _safe_get(cell, "protection")

    return {
        "coordinate": coordinate,
        "row": row,
        "column": col,
        "column_letter": get_column_letter(col),
        "excel_data_type": cell_data_type,
        "literal_value": literal_value,
        "literal_type": literal_type,
        "formula": formula,
        "cached_value": cached_value,
        "cached_type": cached_type,
        "cache_status": cache_status,
        "cache_source": "workbook_cache" if cache_status == "present" else None,
        "style_id": int(_safe_get(cell, "style_id", 0) or 0),
        "style_hash": style_hash,
        "style": style_details,
        "number_format": _safe_get(cell, "number_format", "General"),
        "merge_range": merge.get("range") if merge else None,
        "merge_anchor": merge.get("anchor") if merge else None,
        "is_merge_anchor": merge.get("is_anchor") if merge else False,
        "row_dimension": row_dimension
        or {
            "row": row,
            "height": profile.sheet_format.get("default_row_height"),
            "hidden": False,
            "outline_level": 0,
            "collapsed": False,
        },
        "column_dimension": column_dimension
        or {
            "index": get_column_letter(col),
            "min_column": col,
            "max_column": col,
            "width": profile.sheet_format.get("default_column_width"),
            "hidden": False,
            "outline_level": 0,
            "collapsed": False,
        },
        "row_hidden": row_hidden,
        "column_hidden": column_hidden,
        "comment": _comment_dict(_safe_get(cell, "comment")),
        "hyperlink": _hyperlink_dict(_safe_get(cell, "hyperlink")),
        "validation_memberships": [
            {key: item[key] for key in ("id", "sqref", "type", "operator")}
            for item in validations
        ],
        "conditional_format_memberships": [
            {
                "id": item["id"],
                "sqref": item["sqref"],
                "rule_count": len(item["rules"]),
            }
            for item in conditional_formats
        ],
        "protection": {
            "locked": _safe_get(protection, "locked"),
            "hidden": _safe_get(protection, "hidden"),
        },
    }


def _border_edge_count(style: Mapping[str, Any]) -> int:
    border = style.get("border") or {}
    return sum(
        1
        for name in ("left", "right", "top", "bottom")
        if isinstance(border.get(name), Mapping) and border[name].get("style")
    )


def _feature_from_snapshot(cell: Any, snapshot: Mapping[str, Any]) -> CellFeature:
    formula = snapshot.get("formula")
    value = snapshot.get("literal_value")
    value_kind = str(snapshot.get("literal_type") or "blank")
    cached_value = snapshot.get("cached_value")
    cached_type = str(snapshot.get("cached_type") or "blank")
    if formula is not None:
        value_kind = "formula"

    text: str | None = None
    textual_value = cached_value if formula is not None else value
    textual_kind = cached_type if formula is not None else value_kind
    if textual_value is not None and textual_kind in {"text", "error"}:
        text = normalize_whitespace(textual_value)

    numeric_kinds = {"integer", "number", "decimal", "duration_seconds"}
    style = snapshot.get("style") or {}
    font = style.get("font") or {}
    fill = style.get("fill") or {}
    alignment = style.get("alignment") or {}
    validation_refs = tuple(
        item["sqref"] for item in snapshot.get("validation_memberships", ())
    )
    conditional_refs = tuple(
        item["sqref"]
        for item in snapshot.get("conditional_format_memberships", ())
    )
    return CellFeature(
        row=int(snapshot["row"]),
        col=int(snapshot["column"]),
        coordinate=str(snapshot["coordinate"]),
        has_value=_safe_get(cell, "value") is not None or cached_value is not None,
        has_formula=formula is not None,
        value_kind=value_kind,
        text=text,
        numeric=(cached_type if formula is not None else value_kind) in numeric_kinds,
        border_edges=_border_edge_count(style),
        has_fill=bool(fill.get("fill_type") or fill.get("gradient_type")),
        bold=bool(font.get("bold")),
        italic=bool(font.get("italic")),
        non_general_format=snapshot.get("number_format") not in (None, "General"),
        indent=float(alignment.get("indent") or 0.0),
        merged_range=snapshot.get("merge_range"),
        merge_anchor=snapshot.get("merge_anchor"),
        validation_refs=validation_refs,
        conditional_format_refs=conditional_refs,
        has_comment=snapshot.get("comment") is not None,
        has_hyperlink=snapshot.get("hyperlink") is not None,
        style_hash=snapshot.get("style_hash"),
        row_hidden=bool(snapshot.get("row_hidden")),
        column_hidden=bool(snapshot.get("column_hidden")),
    )


def _coordinate_from_a1(coordinate: str) -> tuple[int, int] | None:
    match = _COORDINATE_RE.fullmatch(coordinate)
    if not match:
        return None
    return int(match.group(2)), column_index_from_string(match.group(1))


def scan_sheet(
    ws: Any,
    cached_ws: Any,
    raw_formula_records: Any,
    config: ExtractionConfig,
    style_registry: StyleRegistry,
) -> SheetProfile:
    """Scan ``ws`` into a sparse :class:`SheetProfile`.

    ``cached_ws`` should be the corresponding worksheet from a workbook loaded
    with ``data_only=True``.  It may be ``None``; missing caches are reported and
    are never confused with literal zero/blank values.
    """

    warnings: list[WarningRecord] = []
    row_dimensions = _row_dimension_metadata(ws)
    column_dimensions, column_dimension_index = _column_dimension_metadata(ws)
    validations, validation_bounds = _serialize_validations(ws, warnings)
    conditional_formats, conditional_format_bounds = _serialize_conditional_formats(
        ws, warnings
    )
    explicit_tables = _serialize_explicit_tables(ws, warnings)
    formula_record_index = _index_raw_formula_records(raw_formula_records, ws.title)

    material_coordinates = _instantiated_coordinates(ws)
    instantiated_count = len(material_coordinates)
    if instantiated_count > config.max_material_cells_per_sheet:
        raise ValueError(
            f"Worksheet {ws.title!r} contains {instantiated_count:,} instantiated "
            f"cells, exceeding limit {config.max_material_cells_per_sheet:,}"
        )

    merge_index, merged_ranges = _merge_metadata(
        ws, material_coordinates, config, warnings
    )
    forced_evidence: set[tuple[int, int]] = set(merge_index)

    for coordinate in formula_record_index:
        parsed = _coordinate_from_a1(coordinate)
        if parsed is not None:
            material_coordinates.add(parsed)
            forced_evidence.add(parsed)

    _bounded_membership_materialization(
        (validation_bounds, conditional_format_bounds),
        material_coordinates,
        forced_evidence,
        config,
    )
    if len(material_coordinates) > config.max_material_cells_per_sheet:
        raise ValueError(
            f"Worksheet {ws.title!r} structural materialization exceeds limit "
            f"{config.max_material_cells_per_sheet:,}"
        )

    profile = SheetProfile(
        sheet_name=ws.title,
        sheet_state=str(_safe_get(ws, "sheet_state", "visible")),
        material_coordinates=material_coordinates,
        merge_index=merge_index,
        merged_ranges=merged_ranges,
        validations=validations,
        conditional_formats=conditional_formats,
        explicit_tables=explicit_tables,
        row_dimensions=row_dimensions,
        column_dimensions=column_dimensions,
        sheet_format=_sheet_format_metadata(ws),
        warnings=warnings,
        config_fingerprint=config.fingerprint,
        include_cached_values=config.include_cached_values,
        validation_bounds=validation_bounds,
        conditional_format_bounds=conditional_format_bounds,
        column_dimension_index=column_dimension_index,
        formula_record_index=formula_record_index,
        forced_evidence_coordinates=forced_evidence,
    )

    # Existing coordinates within native table ranges are unequivocal evidence,
    # while the full native range remains available without creating blank cells.
    table_bounds = [_bounds_from_ref(item["range"]) for item in explicit_tables]
    for coordinate in material_coordinates:
        if any(bounds.contains(*coordinate) for bounds in table_bounds):
            forced_evidence.add(coordinate)

    content_coordinates: set[tuple[int, int]] = set()
    style_coordinates: set[tuple[int, int]] = set()
    strong_evidence: set[tuple[int, int]] = set(forced_evidence)
    worksheet_formula_coordinates: set[str] = set()
    formula_coordinates: set[str] = set()
    cached_formula_count = 0
    missing_cached_formula_count = 0
    comments = 0
    hyperlinks = 0
    value_kind_counts: dict[str, int] = {}

    for row, col in sorted(material_coordinates):
        cell = _cell_for_coordinate(ws, row, col)
        snapshot = cell_snapshot(
            ws,
            cached_ws,
            row,
            col,
            profile,
            raw_formula_records,
            style_registry,
        )
        feature = _feature_from_snapshot(cell, snapshot)
        profile.features[(row, col)] = feature
        value_kind_counts[feature.value_kind] = value_kind_counts.get(feature.value_kind, 0) + 1

        if _safe_get(cell, "value") is not None or feature.has_formula:
            content_coordinates.add((row, col))
        if bool(_safe_get(cell, "has_style", False)):
            style_coordinates.add((row, col))
        if feature.structural:
            strong_evidence.add((row, col))
        if feature.has_formula:
            formula_coordinates.add(feature.coordinate)
            formula_metadata = snapshot.get("formula")
            if (
                isinstance(formula_metadata, Mapping)
                and formula_metadata.get("tokenization_status") == "failed"
            ):
                warnings.append(
                    WarningRecord(
                        code="FORMULA_TOKENIZATION_FAILED",
                        message=(
                            "Formula text was preserved, but dependency and relative-"
                            "signature parsing could not be completed."
                        ),
                        sheet=ws.title,
                        coordinate=feature.coordinate,
                        details={
                            "formula": formula_metadata.get("exact"),
                            "error": formula_metadata.get("tokenization_error"),
                        },
                    )
                )
            if snapshot["cache_status"] == "present":
                cached_formula_count += 1
            elif snapshot["cache_status"] == "missing":
                missing_cached_formula_count += 1
        if _is_formula_cell(cell):
            worksheet_formula_coordinates.add(feature.coordinate)
        if feature.has_comment:
            comments += 1
        if feature.has_hyperlink:
            hyperlinks += 1

    # Filled/bold/non-general blank cells are weak evidence.  Include them only
    # when directly adjacent to strong evidence, preventing one remote formatted
    # cell from inflating the detector's working bounds.
    weak_style_coordinates = {
        coordinate
        for coordinate, feature in profile.features.items()
        if feature.has_fill
        or feature.bold
        or feature.italic
        or feature.non_general_format
    }
    for row, col in weak_style_coordinates:
        if any(
            (row + row_delta, col + col_delta) in strong_evidence
            for row_delta, col_delta in ((0, 0), (-1, 0), (1, 0), (0, -1), (0, 1))
        ):
            strong_evidence.add((row, col))

    profile.content_bounds = _bounds_from_coordinates(content_coordinates)
    profile.style_bounds = _bounds_from_coordinates(style_coordinates)
    profile.evidence_bounds = _bounds_from_coordinates(strong_evidence)

    raw_coordinates = set(formula_record_index)
    worksheet_without_raw = sorted(worksheet_formula_coordinates - raw_coordinates)
    raw_without_worksheet = sorted(raw_coordinates - worksheet_formula_coordinates)
    if raw_formula_records is not None and (worksheet_without_raw or raw_without_worksheet):
        warnings.append(
            WarningRecord(
                code="FORMULA_RECORD_RECONCILIATION",
                message="Raw OOXML formula records and worksheet formula cells differ",
                sheet=ws.title,
                details={
                    "worksheet_without_raw_count": len(worksheet_without_raw),
                    "worksheet_without_raw": worksheet_without_raw[:100],
                    "raw_without_worksheet_count": len(raw_without_worksheet),
                    "raw_without_worksheet": raw_without_worksheet[:100],
                },
            )
        )
    if formula_coordinates and config.include_cached_values and cached_ws is None:
        warnings.append(
            WarningRecord(
                code="FORMULA_CACHE_NOT_LOADED",
                message="Formula cells were scanned without a data-only worksheet",
                sheet=ws.title,
                details={"formula_count": len(formula_coordinates)},
            )
        )
    elif missing_cached_formula_count:
        warnings.append(
            WarningRecord(
                code="FORMULA_CACHE_MISSING",
                message="Some formulas have no workbook-cached result",
                sheet=ws.title,
                details={"missing_cache_count": missing_cached_formula_count},
            )
        )

    declared_bounds: Bounds | None = None
    try:
        declared_bounds = _bounds_from_ref(str(ws.calculate_dimension()))
    except (AttributeError, TypeError, ValueError):
        pass
    if (
        declared_bounds is not None
        and profile.evidence_bounds is not None
        and declared_bounds.area > max(10_000, profile.evidence_bounds.area * 100)
    ):
        warnings.append(
            WarningRecord(
                code="INFLATED_WORKSHEET_DIMENSION",
                message="Declared worksheet dimensions greatly exceed structural evidence",
                sheet=ws.title,
                details={
                    "declared_range": declared_bounds.ref,
                    "evidence_range": profile.evidence_bounds.ref,
                },
            )
        )

    profile.counts = {
        "instantiated_cells": instantiated_count,
        "material_cells": len(material_coordinates),
        "content_cells": len(content_coordinates),
        "styled_cells": len(style_coordinates),
        "evidence_cells": len(strong_evidence),
        "formula_cells": len(formula_coordinates),
        "worksheet_formula_cells": len(worksheet_formula_coordinates),
        "raw_formula_records": len(raw_coordinates),
        "formula_records_matched": len(worksheet_formula_coordinates & raw_coordinates),
        "formula_cells_with_cache": cached_formula_count,
        "formula_cells_missing_cache": missing_cached_formula_count,
        "merged_ranges": len(merged_ranges),
        "merged_index_cells": len(merge_index),
        "validations": len(validations),
        "conditional_formats": len(conditional_formats),
        "native_tables": len(explicit_tables),
        "row_dimensions": len(row_dimensions),
        "column_dimensions": len(column_dimensions),
        "hidden_rows": sum(1 for item in row_dimensions.values() if item["hidden"]),
        "hidden_column_ranges": sum(
            1 for item in column_dimensions.values() if item["hidden"]
        ),
        "comments": comments,
        "hyperlinks": hyperlinks,
        "value_kinds": {key: value_kind_counts[key] for key in sorted(value_kind_counts)},
        "warnings": len(warnings),
    }
    return profile


__all__ = ["SheetProfile", "StyleRegistry", "cell_snapshot", "scan_sheet"]
