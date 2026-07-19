"""Build concise, business-readable tables from lossless POPS extraction data.

The detector and semantic inference layers intentionally retain the original
worksheet rectangle, including presentation banners, merged section bands,
notes, blank grid scaffolding, and formula metadata.  This module creates a
second, clean view for people and downstream analytics without mutating or
discarding the raw evidence.

``build_clean_table`` is deliberately tolerant of both internal dataclasses
and their JSON/dict representations.  Every retained value remains traceable
to an Excel coordinate; formulas, cached and literal values, number formats,
and style references are kept separate.  An uncached formula is displayed as
its formula text rather than being confused with a blank cell.
"""

from __future__ import annotations

from collections import Counter
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict, is_dataclass
import re
from typing import Any

from openpyxl.utils import coordinate_to_tuple, get_column_letter

from .models import Bounds, TableCandidate
from .utils import normalize_whitespace, unique_preserving_order


_BANNER_RE = re.compile(
    r"(?:copy\s+(?:the\s+)?table|copy\s+.*(?:powerpoint|\bppt\b)|"
    r"do\s+not\s+copy|not\s+(?:to\s+be\s+)?cop(?:y|ied))",
    re.IGNORECASE,
)
_INSTRUCTION_RE = re.compile(
    r"(?:^\s*(?:please|ensure|must|manually|enter|fill\s+in)\b|"
    r"^\s*complete\s+(?:all|the|these|yellow|required|missing|cells?)\b|"
    r"please\s+ensure)",
    re.IGNORECASE,
)
_SOURCE_RE = re.compile(
    r"(?:^\s*(?:source|sourcing)\b|https?://|equal\s+to\s+figures)",
    re.IGNORECASE,
)
_NOTE_RE = re.compile(
    r"(?:^\s*(?:note|definition|where)\b|^\s*(?:\*|\(\*))",
    re.IGNORECASE,
)
_HEADER_TOKEN_RE = re.compile(
    r"(?:\b20\d{2}\b|\b(?:actual|baseline|budget|forecast|plan|target)\b|"
    r"\bf\d{1,2}\b|\bb\s*20\d{2}\b|\b(?:fte|opex|kpi)\b|"
    r"\b(?:activity|category|contract|department|description|entity|expiry|"
    r"function|metric|mission|name|provider|role|source|unit)\b|"
    r"\b(?:m|k)?(?:eur|€)\b|%|\b(?:nb|number|count)\b|"
    r"(?:\bvs\.?\b|\bversus\b|\bvariance\b|\bdelta\b|[Δ∆]))",
    re.IGNORECASE,
)
_RANGE_LIKE_RE = re.compile(
    r"^(?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+"
    r"(?::\$?[A-Z]{1,3}\$?\d+)?$",
    re.IGNORECASE,
)
_PERIOD_ONLY_RE = re.compile(
    r"^(?:actual|baseline|budget|forecast|plan|target|f\d{1,2}|b\s*20\d{2}|"
    r"20\d{2}|deep[-\s]*dive(?:\s*20\d{2}\s*(?:-|->|â†’|to)\s*20\d{2})?)$",
    re.IGNORECASE,
)
_STRUCTURAL_ROLES = {"row_group", "row_label", "dimension"}
_VALUE_ROLES = {"measure", "formula_measure", "variance", "unknown_measure"}
_REMOVED_ROLES = {
    "title",
    "header",
    "section_header",
    "group_header",
    "spacer",
    "note",
    "footnote",
    "source",
    "instruction",
    "banner",
}


def _plain(value: object) -> dict[str, Any]:
    """Return a shallow plain-dict representation of a descriptor."""

    if isinstance(value, Mapping):
        return dict(value)
    if is_dataclass(value):
        return asdict(value)
    if hasattr(value, "__dict__"):
        return dict(vars(value))
    return {}


def _coerce_bounds(value: object) -> Bounds:
    if isinstance(value, Bounds):
        return value
    if isinstance(value, TableCandidate):
        return value.bounds
    if isinstance(value, str):
        ref = value.rsplit("!", 1)[-1].replace("$", "").strip().strip("'")
        return Bounds.from_a1(ref)
    if isinstance(value, Mapping):
        for key in ("bounds", "range", "source_range", "ref"):
            if value.get(key) is not None:
                return _coerce_bounds(value[key])
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes)):
        if len(value) == 4 and all(isinstance(item, int) for item in value):
            return Bounds(*value)
    if hasattr(value, "bounds"):
        return _coerce_bounds(getattr(value, "bounds"))
    if all(
        hasattr(value, name)
        for name in ("min_row", "min_col", "max_row", "max_col")
    ):
        return Bounds(
            int(getattr(value, "min_row")),
            int(getattr(value, "min_col")),
            int(getattr(value, "max_row")),
            int(getattr(value, "max_col")),
        )
    raise TypeError("table must expose Bounds or an A1 range")


def _snapshot_items(snapshots: object) -> Iterable[tuple[object, object]]:
    if isinstance(snapshots, Mapping):
        yield from snapshots.items()
        return
    if isinstance(snapshots, Iterable) and not isinstance(snapshots, (str, bytes)):
        for item in snapshots:
            yield None, item


def _coordinate_parts(key: object, record: Mapping[str, Any]) -> tuple[int, int] | None:
    row = record.get("row")
    col = record.get("column")
    if row is not None and col is not None:
        try:
            return int(row), int(col)
        except (TypeError, ValueError):
            pass
    coordinate = record.get("coordinate")
    if coordinate is None and isinstance(key, str):
        coordinate = key
    if isinstance(coordinate, str):
        try:
            parsed_row, parsed_col = coordinate_to_tuple(coordinate.replace("$", ""))
            return parsed_row, parsed_col
        except (TypeError, ValueError):
            pass
    if (
        isinstance(key, Sequence)
        and not isinstance(key, (str, bytes))
        and len(key) >= 2
    ):
        try:
            return int(key[0]), int(key[1])
        except (TypeError, ValueError):
            return None
    return None


def _snapshot_index(snapshots: object) -> dict[tuple[int, int], dict[str, Any]]:
    """Normalize coordinate-, tuple-, and iterable-keyed snapshot containers."""

    result: dict[tuple[int, int], dict[str, Any]] = {}
    for key, value in _snapshot_items(snapshots):
        record = _plain(value)
        coordinate = _coordinate_parts(key, record)
        if coordinate is None:
            continue
        row, col = coordinate
        record.setdefault("row", row)
        record.setdefault("column", col)
        record.setdefault("coordinate", f"{get_column_letter(col)}{row}")
        result[(row, col)] = record
    return result


def _formula(snapshot: Mapping[str, Any]) -> str | None:
    value = snapshot.get("formula")
    if isinstance(value, Mapping):
        value = value.get("exact") or value.get("resolved") or value.get("raw")
    if value in (None, ""):
        return None
    return str(value)


def _cache_status(snapshot: Mapping[str, Any]) -> str:
    return str(
        snapshot.get("cached_value_status")
        or snapshot.get("cache_status")
        or "not_applicable"
    )


def _display_value(snapshot: Mapping[str, Any]) -> Any:
    """Return a readable value without hiding an uncached formula."""

    formula = _formula(snapshot)
    cached = snapshot.get("cached_value")
    if formula:
        if _cache_status(snapshot) == "present" and cached is not None:
            return cached
        return formula
    return snapshot.get("literal_value")


def _meaningful(snapshot: Mapping[str, Any]) -> bool:
    if _formula(snapshot):
        return True
    value = snapshot.get("literal_value")
    if value is None:
        return False
    return not (isinstance(value, str) and not normalize_whitespace(value))


def _merge_catalog(
    snapshots: Mapping[tuple[int, int], Mapping[str, Any]],
) -> list[tuple[Bounds, tuple[int, int]]]:
    result: dict[str, tuple[Bounds, tuple[int, int]]] = {}
    for (row, col), snapshot in snapshots.items():
        ref = snapshot.get("merge_ref") or snapshot.get("merge_range")
        if not isinstance(ref, str) or ":" not in ref:
            continue
        try:
            bounds = Bounds.from_a1(ref.replace("$", ""))
        except ValueError:
            continue
        anchor = snapshot.get("merge_anchor")
        if isinstance(anchor, str):
            try:
                anchor_row, anchor_col = coordinate_to_tuple(anchor.replace("$", ""))
            except (TypeError, ValueError):
                anchor_row, anchor_col = bounds.min_row, bounds.min_col
        else:
            anchor_row, anchor_col = bounds.min_row, bounds.min_col
        result[bounds.ref] = (bounds, (anchor_row, anchor_col))
    return [result[key] for key in sorted(result)]


def _resolved_snapshot(
    row: int,
    col: int,
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
    *,
    propagate_horizontal: bool = True,
) -> tuple[dict[str, Any], str | None, bool]:
    """Resolve merged followers to their anchor while retaining their coordinate.

    Header labels need horizontal propagation so a group heading applies to
    every child column.  Body cells do not: a label merged across layout columns
    represents one logical field and would otherwise create duplicate columns.
    Vertical propagation remains useful for row-group labels.
    """

    coordinate = f"{get_column_letter(col)}{row}"
    own = dict(snapshots.get((row, col), {}))
    own.setdefault("coordinate", coordinate)
    own.setdefault("row", row)
    own.setdefault("column", col)
    if _meaningful(own):
        return own, coordinate, False
    for bounds, anchor in merges:
        if not bounds.contains(row, col):
            continue
        if not propagate_horizontal and col != anchor[1]:
            continue
        anchor_snapshot = dict(snapshots.get(anchor, {}))
        if _meaningful(anchor_snapshot):
            anchor_coordinate = f"{get_column_letter(anchor[1])}{anchor[0]}"
            anchor_snapshot["coordinate"] = coordinate
            anchor_snapshot["row"] = row
            anchor_snapshot["column"] = col
            return anchor_snapshot, anchor_coordinate, True
    return own, coordinate, False


def _row_text(
    row: int,
    bounds: Bounds,
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
) -> list[str]:
    values: list[str] = []
    for col in range(bounds.min_col, bounds.max_col + 1):
        snapshot, _, _ = _resolved_snapshot(
            row,
            col,
            snapshots,
            merges,
        )
        value = _display_value(snapshot)
        if isinstance(value, str) and not value.startswith("="):
            text = normalize_whitespace(value)
            if text and text not in values:
                values.append(text)
    return values


def _row_measure_count(
    row: int,
    bounds: Bounds,
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
) -> int:
    count = 0
    for col in range(bounds.min_col, bounds.max_col + 1):
        snapshot, _, _ = _resolved_snapshot(
            row,
            col,
            snapshots,
            merges,
            propagate_horizontal=False,
        )
        if _formula(snapshot):
            count += 1
            continue
        value = snapshot.get("literal_value")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            count += 1
    return count


def _noise_kind(text: str) -> str | None:
    if _BANNER_RE.search(text):
        return "banner"
    if _SOURCE_RE.search(text):
        return "source"
    if _INSTRUCTION_RE.search(text):
        return "instruction"
    if _NOTE_RE.search(text):
        return "note"
    return None


def _row_reason(
    role: str,
    texts: Sequence[str],
    measure_count: int = 0,
) -> str | None:
    joined = " | ".join(texts)
    folded_role = role.casefold()
    if folded_role in {"section_header"}:
        return "section"
    if folded_role == "group_header":
        return "group"
    if folded_role in {"note", "footnote", "source", "instruction", "banner"}:
        return folded_role
    if folded_role == "spacer":
        return "spacer"
    if folded_role == "title":
        return "title"

    noise = _noise_kind(joined)
    if noise is None:
        return None

    # A detached comment can occupy a merged cell beside otherwise legitimate
    # business rows (a common POPS layout).  Do not discard those rows merely
    # because the comment is repeated through the merge.  Its annotation
    # column is removed independently below.
    non_noise_texts = [text for text in texts if _noise_kind(text) is None]
    if (measure_count or folded_role in {"data", "total"}) and non_noise_texts:
        return None
    return noise


def _header_value(snapshot: Mapping[str, Any]) -> str:
    if _formula(snapshot):
        return ""
    value = snapshot.get("literal_value")
    if isinstance(value, bool) or value is None:
        return ""
    if isinstance(value, (int, float)):
        if float(value).is_integer() and 1900 <= int(value) <= 2200:
            return str(int(value))
        return ""
    return normalize_whitespace(value)


def _row_header_texts(
    row: int,
    bounds: Bounds,
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
) -> list[str]:
    values: list[str] = []
    for col in range(bounds.min_col, bounds.max_col + 1):
        snapshot, _, _ = _resolved_snapshot(
            row,
            col,
            snapshots,
            merges,
            propagate_horizontal=False,
        )
        text = _header_value(snapshot)
        if text and text not in values:
            values.append(text)
    return values


def _usable_title(value: object) -> str | None:
    text = normalize_whitespace(value or "")
    if not text:
        return None
    if (
        _BANNER_RE.search(text)
        or _NOTE_RE.search(text)
        or _INSTRUCTION_RE.search(text)
        or _SOURCE_RE.search(text)
        or _RANGE_LIKE_RE.fullmatch(text)
        or _PERIOD_ONLY_RE.fullmatch(text)
    ):
        return None
    return text


def _merge_width_at(
    row: int,
    col: int,
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
) -> int:
    for bounds, _ in merges:
        if bounds.contains(row, col):
            return bounds.width
    return 1


def _infer_title_from_headers(
    bounds: Bounds,
    header_rows: Sequence[int],
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
) -> str | None:
    generic = {
        "activity",
        "entity",
        "group",
        "kpi",
        "metric",
        "source",
        "unit",
        "value",
    }
    candidates: list[tuple[int, int, int, str]] = []
    seen: set[tuple[int, int]] = set()
    for row in sorted(header_rows):
        for col in range(bounds.min_col, bounds.max_col + 1):
            snapshot, value_source, _ = _resolved_snapshot(
                row, col, snapshots, merges
            )
            text = _usable_title(_header_value(snapshot))
            if text is None or text.casefold() in generic:
                continue
            source = coordinate_to_tuple(value_source) if value_source else (row, col)
            if source in seen:
                continue
            seen.add(source)
            letters = [character for character in text if character.isalpha()]
            upper_ratio = (
                sum(character.isupper() for character in letters) / len(letters)
                if letters
                else 0.0
            )
            merge_width = _merge_width_at(row, col, merges)
            word_count = len(re.findall(r"[A-Za-z]+", text))
            if word_count < 2 or (upper_ratio < 0.65 and merge_width == 1):
                continue
            candidates.append((row, col, -merge_width, text))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][3]


def _source_row_from_coordinate(value: object) -> int | None:
    if not isinstance(value, str):
        return None
    try:
        row, _ = coordinate_to_tuple(value.replace("$", ""))
        return row
    except (TypeError, ValueError):
        return None


def _real_header_rows(
    bounds: Bounds,
    structure: Mapping[str, Any],
    columns: Sequence[Mapping[str, Any]],
    row_map: Mapping[int, Mapping[str, Any]],
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
) -> list[int]:
    """Recover the complete header band, even after a weak first inference.

    POPS sheets frequently place a PowerPoint banner on row 1, scenario groups
    on row 3, and the actual leaf headers on row 5.  Treating the detector's
    proposed rows as final therefore produces the generic/hash labels seen in
    the old report.  This pass validates the proposal and augments it from the
    leading worksheet rows until a convincing business row is reached.
    """

    proposed: set[int] = set()
    for value in structure.get("header_rows", ()) or ():
        try:
            proposed.add(int(value))
        except (TypeError, ValueError):
            continue
    for column in columns:
        for source in column.get("header_sources", ()) or ():
            if (row := _source_row_from_coordinate(source)) is not None:
                proposed.add(row)

    scan_end = min(bounds.max_row, bounds.min_row + 11)
    valid: set[int] = set()
    for row in sorted(proposed):
        if not (bounds.min_row <= row <= bounds.max_row):
            continue
        if row > scan_end:
            continue
        texts = _row_header_texts(row, bounds, snapshots, merges)
        role = str(row_map.get(row, {}).get("role", "unknown"))
        measures = _row_measure_count(row, bounds, snapshots, merges)
        reason = _row_reason(role, texts, measures)
        if reason in {
            "banner",
            "source",
            "instruction",
            "note",
            "section",
            "group",
            "spacer",
            "title",
        }:
            continue
        semantic_headers = sum(bool(_HEADER_TOKEN_RE.search(text)) for text in texts)
        if texts and not (measures >= 2 and semantic_headers < 2):
            valid.add(row)

    # Always inspect the leading band, even if some proposed rows survived.
    # This is what recovers a missing leaf row after a blank visual spacer.
    for row in range(bounds.min_row, scan_end + 1):
        texts = _row_header_texts(row, bounds, snapshots, merges)
        role = str(row_map.get(row, {}).get("role", "unknown"))
        measures = _row_measure_count(row, bounds, snapshots, merges)
        reason = _row_reason(role, texts, measures)
        if reason in {"section", "group"}:
            if valid:
                break
            continue
        if reason in {
            "banner",
            "source",
            "instruction",
            "note",
            "spacer",
            "title",
        }:
            continue
        if not texts:
            continue
        semantic_headers = sum(bool(_HEADER_TOKEN_RE.search(text)) for text in texts)
        folded_role = role.casefold()
        body_like = (
            semantic_headers < 2
            and (
                measures >= 2
                or folded_role in {"data", "total"}
            )
        )
        if body_like:
            break
        if semantic_headers >= 2 or (
            len(texts) >= 2
            and measures == 0
            and folded_role not in {"data", "total"}
        ):
            valid.add(row)
    return sorted(valid)


def _clean_header_path(path: object, title: str | None) -> list[str]:
    values = path if isinstance(path, Sequence) and not isinstance(path, (str, bytes)) else [path]
    result: list[str] = []
    for value in values:
        text = normalize_whitespace(value or "")
        if (
            not text
            or _BANNER_RE.search(text)
            or _NOTE_RE.search(text)
            or _INSTRUCTION_RE.search(text)
            or _RANGE_LIKE_RE.fullmatch(text)
        ):
            continue
        if title and text.casefold() == title.casefold():
            continue
        if not result or result[-1].casefold() != text.casefold():
            result.append(text)
    return result


def _column_header_path(
    col: int,
    descriptor: Mapping[str, Any],
    header_rows: Sequence[int],
    title: str | None,
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
) -> list[str]:
    actual_values: list[str] = []
    for row in sorted(header_rows):
        snapshot, _, _ = _resolved_snapshot(row, col, snapshots, merges)
        text = _header_value(snapshot)
        if text:
            actual_values.append(text)
    actual = _clean_header_path(actual_values, title)
    inferred = _clean_header_path(descriptor.get("header_path", []), title)
    if not actual:
        return inferred
    if not inferred:
        return actual

    # Keep a richer inferred parent path only when it demonstrably terminates
    # in the same physical leaf.  Otherwise the direct worksheet headers win.
    if (
        len(inferred) > len(actual)
        and inferred[-1].casefold() == actual[-1].casefold()
    ):
        return inferred
    return actual


def _annotation_column(
    descriptor: Mapping[str, Any],
    col: int,
    preliminary_rows: Sequence[tuple[int, str, list[str], Mapping[str, Any]]],
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    merges: Sequence[tuple[Bounds, tuple[int, int]]],
) -> bool:
    role = str(descriptor.get("role", "unknown")).casefold()
    if role == "annotation":
        return True
    values: list[str] = []
    for row, _, _, _ in preliminary_rows:
        snapshot, _, _ = _resolved_snapshot(
            row,
            col,
            snapshots,
            merges,
            propagate_horizontal=False,
        )
        if _formula(snapshot):
            return False
        value = snapshot.get("literal_value")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return False
        if isinstance(value, str) and normalize_whitespace(value):
            values.append(normalize_whitespace(value))
    if not values:
        return False
    if all(_noise_kind(value) is not None for value in values):
        return True
    return (
        role not in _STRUCTURAL_ROLES
        and len(values) <= 3
        and all(len(value) >= 100 for value in values)
    )


def _natural_structural_labels(
    descriptors: Sequence[dict[str, Any]],
) -> dict[int, str]:
    structural = [item for item in descriptors if item.get("role") in _STRUCTURAL_ROLES]
    result: dict[int, str] = {}
    total = len(structural)
    for position, item in enumerate(structural):
        col = int(item["column"])
        role = str(item.get("role", "unknown"))
        if total == 1:
            label = "Metric"
        elif role == "row_group":
            label = "Entity" if position == 0 else "Group"
        elif role == "row_label" and position == total - 1:
            label = "Metric"
        elif role == "dimension":
            label = "Activity" if position < total - 1 else "Group"
        else:
            label = "Activity" if position < total - 1 else "Metric"
        result[col] = label
    return result


def _unique_labels(labels: Sequence[str]) -> list[str]:
    counts: Counter[str] = Counter()
    result: list[str] = []
    for raw in labels:
        base = normalize_whitespace(raw) or "Value"
        folded = base.casefold()
        counts[folded] += 1
        result.append(base if counts[folded] == 1 else f"{base} ({counts[folded]})")
    return result


def _key_for_label(label: str, used: Counter[str]) -> str:
    base = re.sub(r"[^a-z0-9]+", "_", label.casefold()).strip("_") or "value"
    used[base] += 1
    return base if used[base] == 1 else f"{base}_{used[base]}"


def _section_for_row(
    row: int,
    row_info: Mapping[str, Any],
    sections: Sequence[Mapping[str, Any]],
    current: Sequence[str],
) -> list[str]:
    explicit = row_info.get("section_path") or []
    if explicit:
        return [normalize_whitespace(value) for value in explicit if normalize_whitespace(value)]
    containing = [
        section
        for section in sections
        if int(section.get("start_row", -1)) <= row <= int(section.get("end_row", -1))
    ]
    if containing:
        containing.sort(key=lambda item: (int(item.get("start_row", 0)), str(item.get("title", ""))))
        title = normalize_whitespace(containing[-1].get("title", ""))
        return [title] if title else list(current)
    return list(current)


def _label_value(
    row_info: Mapping[str, Any],
    descriptor: Mapping[str, Any],
    structural_position: int,
) -> Any:
    path = list(row_info.get("label_path") or [])
    if not path or structural_position >= len(path):
        return None
    if (
        descriptor.get("role") == "row_label"
        and structural_position == len(path) - 1
    ):
        return row_info.get("metric_label") or path[-1]
    return path[structural_position]


def _clean_cell(
    row: int,
    col: int,
    snapshot: Mapping[str, Any],
    value_source: str | None,
    inherited: bool,
    synthetic_value: Any = None,
) -> dict[str, Any]:
    coordinate = f"{get_column_letter(col)}{row}"
    formula = _formula(snapshot)
    literal = snapshot.get("literal_value")
    cached = snapshot.get("cached_value")
    value = _display_value(snapshot)
    synthetic = False
    if value is None and synthetic_value is not None:
        value = synthetic_value
        literal = synthetic_value
        synthetic = True
    style_ref = snapshot.get("style_ref")
    if style_ref is None and snapshot.get("style_hash"):
        style_ref = f"style:{snapshot['style_hash']}"
    result = {
        "coordinate": coordinate,
        "value": value,
        "formula": formula,
        "cached": cached,
        "literal": literal,
        "number_format": snapshot.get("number_format", "General"),
        "style_ref": style_ref,
        "style": snapshot.get("style"),
        "cached_value_status": _cache_status(snapshot),
    }
    if inherited and value_source and value_source != coordinate:
        result["value_source_coordinate"] = value_source
        result["merged_inherited"] = True
    if synthetic:
        result["synthetic"] = True
    return result


def build_clean_table(
    table: object,
    structure: Mapping[str, Any] | object,
    snapshots: object,
) -> dict[str, Any]:
    """Return a deterministic, business-readable view of a raw table region.

    Parameters
    ----------
    table:
        A ``TableCandidate`` or mapping/object exposing ``bounds`` or an A1
        ``range``/``source_range``.
    structure:
        Semantic inference output.  Both ``column_descriptors``/``columns`` and
        ``row_descriptors``/``rows`` aliases are accepted.
    snapshots:
        Lossless cell records keyed by ``(row, column)``, coordinate, or supplied
        as an iterable.  Literal, formula, cached, format, style, and merge fields
        are consumed without evaluating formulas.

    Returns
    -------
    dict
        JSON-compatible clean table with the exact top-level contract
        ``status``, ``title``, ``source_range``, ``header_rows``, ``columns``,
        ``rows``, ``dropped_rows``, ``dropped_columns``, and ``counts``.
        ``status`` is ``"empty"`` when no defensible business body remains.
    """

    bounds = _coerce_bounds(table)
    structure_map = _plain(structure)
    snapshot_map = _snapshot_index(snapshots)
    merges = _merge_catalog(snapshot_map)
    column_descriptors = [
        _plain(value)
        for value in (
            structure_map.get("column_descriptors")
            or structure_map.get("columns")
            or []
        )
    ]
    by_column = {
        int(item["column"]): item
        for item in column_descriptors
        if item.get("column") is not None
    }
    for col in range(bounds.min_col, bounds.max_col + 1):
        by_column.setdefault(
            col,
            {
                "column": col,
                "letter": get_column_letter(col),
                "role": "unknown",
                "header_path": [],
                "header_sources": [],
            },
        )
    descriptors = [by_column[col] for col in sorted(by_column) if bounds.min_col <= col <= bounds.max_col]

    row_descriptors = [
        _plain(value)
        for value in (
            structure_map.get("row_descriptors")
            or structure_map.get("rows")
            or []
        )
    ]
    row_map = {
        int(item["row"]): item
        for item in row_descriptors
        if item.get("row") is not None
    }
    sections = [_plain(value) for value in structure_map.get("sections", []) or []]
    header_rows = _real_header_rows(
        bounds,
        structure_map,
        descriptors,
        row_map,
        snapshot_map,
        merges,
    )

    title = _usable_title(structure_map.get("title"))
    if title is None and isinstance(table, Mapping):
        fallback = table.get("title") or table.get("source_name")
        title = _usable_title(fallback)
    if title is None:
        title = _infer_title_from_headers(
            bounds,
            header_rows,
            snapshot_map,
            merges,
        )

    # Classify source rows first; pure section/group rows update context but do
    # not become observations.
    dropped_rows: list[dict[str, Any]] = []
    preliminary_rows: list[tuple[int, str, list[str], Mapping[str, Any]]] = []
    current_section: list[str] = []
    for row in range(bounds.min_row, bounds.max_row + 1):
        row_info = row_map.get(row, {})
        role = str(row_info.get("role", "unknown"))
        texts = _row_text(row, bounds, snapshot_map, merges)
        measure_count = _row_measure_count(row, bounds, snapshot_map, merges)
        if row in header_rows:
            dropped_rows.append({"source_row": row, "reason": "header"})
            continue
        reason = _row_reason(role, texts, measure_count)
        if reason in {"section", "group"}:
            label = normalize_whitespace(
                row_info.get("metric_label")
                or (row_info.get("label_path") or [""])[-1]
                or (texts[0] if texts else "")
            )
            if label:
                current_section = [label]
        if reason is not None:
            dropped_rows.append({"source_row": row, "reason": reason})
            continue
        section_path = _section_for_row(row, row_info, sections, current_section)
        preliminary_rows.append((row, role, section_path, row_info))

    structural_labels = _natural_structural_labels(descriptors)
    structural_order = {
        int(item["column"]): position
        for position, item in enumerate(
            [item for item in descriptors if item.get("role") in _STRUCTURAL_ROLES]
        )
    }

    # Determine body-empty columns using merged values, formulas, and semantic
    # row-label paths.  Uncached formulas always count as populated.
    kept_descriptors: list[dict[str, Any]] = []
    dropped_columns: list[dict[str, Any]] = []
    for descriptor in descriptors:
        col = int(descriptor["column"])
        if _annotation_column(
            descriptor,
            col,
            preliminary_rows,
            snapshot_map,
            merges,
        ):
            dropped_columns.append(
                {
                    "source_column": col,
                    "source_letter": get_column_letter(col),
                    "reason": "annotation",
                }
            )
            continue
        populated = False
        for row, _, _, row_info in preliminary_rows:
            snapshot, _, _ = _resolved_snapshot(
                row,
                col,
                snapshot_map,
                merges,
                propagate_horizontal=False,
            )
            if _meaningful(snapshot):
                populated = True
                break
            if descriptor.get("role") in _STRUCTURAL_ROLES and _label_value(
                row_info, descriptor, structural_order.get(col, 0)
            ) is not None:
                populated = True
                break
        if populated:
            kept_descriptors.append(descriptor)
        else:
            dropped_columns.append(
                {
                    "source_column": col,
                    "source_letter": get_column_letter(col),
                    "reason": "body_empty",
                }
            )

    # Build readable, collision-free labels and keys without leaking inference
    # hashes into the business table.
    raw_labels: list[str] = []
    cleaned_paths: list[list[str]] = []
    synthetic_flags: list[bool] = []
    for descriptor in kept_descriptors:
        col = int(descriptor["column"])
        path = _column_header_path(
            col,
            descriptor,
            header_rows,
            title,
            snapshot_map,
            merges,
        )
        cleaned_paths.append(path)
        if path:
            label = " | ".join(path)
            synthetic = False
        elif descriptor.get("role") in _STRUCTURAL_ROLES:
            label = structural_labels.get(col, "Group")
            synthetic = True
        elif descriptor.get("role") in _VALUE_ROLES:
            label = "Value"
            synthetic = True
        else:
            label = "Group" if col == bounds.min_col else "Value"
            synthetic = True
        raw_labels.append(label)
        synthetic_flags.append(synthetic)

    labels = _unique_labels(raw_labels)
    key_counts: Counter[str] = Counter()
    clean_columns: list[dict[str, Any]] = []
    for descriptor, path, label, synthetic in zip(
        kept_descriptors, cleaned_paths, labels, synthetic_flags
    ):
        col = int(descriptor["column"])
        output = {
            "key": _key_for_label(label, key_counts),
            "label": label,
            "source_column": col,
            "source_letter": get_column_letter(col),
            "role": str(descriptor.get("role", "unknown")),
            "header_path": path,
        }
        if synthetic:
            output["synthetic"] = True
        clean_columns.append(output)

    clean_rows: list[dict[str, Any]] = []
    for row, role, section_path, row_info in preliminary_rows:
        cells: list[dict[str, Any]] = []
        meaningful = False
        for descriptor in kept_descriptors:
            col = int(descriptor["column"])
            snapshot, value_source, inherited = _resolved_snapshot(
                row,
                col,
                snapshot_map,
                merges,
                propagate_horizontal=False,
            )
            synthetic_value = None
            if descriptor.get("role") in _STRUCTURAL_ROLES:
                synthetic_value = _label_value(
                    row_info, descriptor, structural_order.get(col, 0)
                )
            cell = _clean_cell(
                row,
                col,
                snapshot,
                value_source,
                inherited,
                synthetic_value,
            )
            if cell["value"] is not None and not (
                isinstance(cell["value"], str)
                and not normalize_whitespace(cell["value"])
            ):
                meaningful = True
            cells.append(cell)
        if not meaningful:
            dropped_rows.append({"source_row": row, "reason": "body_empty"})
            continue
        clean_rows.append(
            {
                "source_row": row,
                "role": role,
                "section_path": section_path,
                "cells": cells,
            }
        )

    dropped_rows.sort(key=lambda item: (int(item["source_row"]), str(item["reason"])))
    dropped_columns.sort(
        key=lambda item: (int(item["source_column"]), str(item["reason"]))
    )
    formula_cells = sum(
        bool(cell.get("formula"))
        for row in clean_rows
        for cell in row["cells"]
    )
    inherited_cells = sum(
        bool(cell.get("merged_inherited"))
        for row in clean_rows
        for cell in row["cells"]
    )
    synthetic_cells = sum(
        bool(cell.get("synthetic"))
        for row in clean_rows
        for cell in row["cells"]
    )
    status = "ready" if clean_rows and clean_columns else "empty"
    return {
        "status": status,
        "title": title,
        "source_range": bounds.ref,
        "header_rows": sorted(header_rows),
        "columns": clean_columns,
        "rows": clean_rows,
        "dropped_rows": dropped_rows,
        "dropped_columns": dropped_columns,
        "counts": {
            "source_rows": bounds.height,
            "source_columns": bounds.width,
            "header_rows": len(header_rows),
            "rows": len(clean_rows),
            "columns": len(clean_columns),
            "cells": len(clean_rows) * len(clean_columns),
            "dropped_rows": len(dropped_rows),
            "dropped_columns": len(dropped_columns),
            "formula_cells": formula_cells,
            "merged_inherited_cells": inherited_cells,
            "synthetic_cells": synthetic_cells,
        },
    }


__all__ = ["build_clean_table"]
