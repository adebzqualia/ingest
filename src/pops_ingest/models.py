"""Small, dependency-free models shared by scanner, detector, and exporters."""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries


@dataclass(frozen=True, order=True, slots=True)
class Bounds:
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    def __post_init__(self) -> None:
        if min(self.min_row, self.min_col) < 1:
            raise ValueError("Excel bounds are one-based")
        if self.max_row < self.min_row or self.max_col < self.min_col:
            raise ValueError("Invalid bounds")

    @classmethod
    def from_a1(cls, ref: str) -> "Bounds":
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        return cls(min_row, min_col, max_row, max_col)

    @property
    def width(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def height(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def area(self) -> int:
        return self.width * self.height

    @property
    def ref(self) -> str:
        start = f"{get_column_letter(self.min_col)}{self.min_row}"
        end = f"{get_column_letter(self.max_col)}{self.max_row}"
        return start if start == end else f"{start}:{end}"

    def contains(self, row: int, col: int) -> bool:
        return (
            self.min_row <= row <= self.max_row
            and self.min_col <= col <= self.max_col
        )

    def contains_bounds(self, other: "Bounds") -> bool:
        return (
            self.min_row <= other.min_row
            and self.min_col <= other.min_col
            and self.max_row >= other.max_row
            and self.max_col >= other.max_col
        )

    def intersects(self, other: "Bounds") -> bool:
        return not (
            self.max_row < other.min_row
            or other.max_row < self.min_row
            or self.max_col < other.min_col
            or other.max_col < self.min_col
        )

    def intersection_area(self, other: "Bounds") -> int:
        if not self.intersects(other):
            return 0
        return (
            min(self.max_row, other.max_row)
            - max(self.min_row, other.min_row)
            + 1
        ) * (
            min(self.max_col, other.max_col)
            - max(self.min_col, other.min_col)
            + 1
        )

    def iou(self, other: "Bounds") -> float:
        intersection = self.intersection_area(other)
        union = self.area + other.area - intersection
        return intersection / union if union else 0.0

    def union(self, other: "Bounds") -> "Bounds":
        return Bounds(
            min(self.min_row, other.min_row),
            min(self.min_col, other.min_col),
            max(self.max_row, other.max_row),
            max(self.max_col, other.max_col),
        )


@dataclass(slots=True)
class CellFeature:
    row: int
    col: int
    coordinate: str
    has_value: bool = False
    has_formula: bool = False
    value_kind: str = "blank"
    text: str | None = None
    numeric: bool = False
    border_edges: int = 0
    has_fill: bool = False
    bold: bool = False
    italic: bool = False
    non_general_format: bool = False
    indent: float = 0.0
    merged_range: str | None = None
    merge_anchor: str | None = None
    validation_refs: tuple[str, ...] = ()
    conditional_format_refs: tuple[str, ...] = ()
    has_comment: bool = False
    has_hyperlink: bool = False
    style_hash: str | None = None
    row_hidden: bool = False
    column_hidden: bool = False

    @property
    def structural(self) -> bool:
        return bool(
            self.has_value
            or self.has_formula
            or self.border_edges
            or self.merged_range
            or self.validation_refs
            or self.has_comment
            or self.has_hyperlink
        )


@dataclass(slots=True)
class TableCandidate:
    bounds: Bounds
    methods: set[str] = field(default_factory=set)
    confidence: float = 0.0
    features: dict[str, float] = field(default_factory=dict)
    reasons: list[str] = field(default_factory=list)
    source_name: str | None = None
    forced: bool = False
    uncertain: bool = False

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["bounds"] = self.bounds.ref
        payload["methods"] = sorted(self.methods)
        return payload


@dataclass(slots=True)
class ColumnDescriptor:
    column: int
    letter: str
    role: str
    header_path: list[str]
    header_sources: list[str]
    key: str
    scenario: str | None = None
    year: int | None = None
    comparison_kind: str | None = None
    unit: str | None = None
    measure: str | None = None
    group: str | None = None
    confidence: float = 0.0


@dataclass(slots=True)
class RowDescriptor:
    row: int
    role: str
    label_path: list[str]
    label_sources: list[str]
    metric_label: str | None = None
    section_path: list[str] = field(default_factory=list)
    is_total: bool = False
    is_placeholder: bool = False
    confidence: float = 0.0


@dataclass(slots=True)
class AnnotationBlock:
    bounds: Bounds
    text: list[str]
    kind: str
    nearest_table_id: str | None = None
    distance: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "range": self.bounds.ref,
            "text": self.text,
            "kind": self.kind,
            "nearest_table_id": self.nearest_table_id,
            "distance": self.distance,
        }


@dataclass(slots=True)
class WarningRecord:
    code: str
    message: str
    severity: str = "warning"
    sheet: str | None = None
    coordinate: str | None = None
    table_id: str | None = None
    details: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)
