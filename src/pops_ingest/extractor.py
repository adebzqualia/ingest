"""End-to-end POPS workbook extraction orchestrator."""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
from datetime import datetime, timezone
from hashlib import sha256
import json
import platform
from pathlib import Path
import re
from typing import Any, Iterable, Iterator, Mapping

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import __version__
from .config import ExtractionConfig
from .detection import detect_table_candidates
from .metadata import (
    defined_names_metadata,
    static_named_ranges_for_sheet,
    workbook_metadata,
)
from .models import AnnotationBlock, Bounds, TableCandidate, WarningRecord
from .ooxml import OOXMLIndex
from .output import AtomicOutput
from .report import build_html_report
from .schema import BUNDLE_SCHEMA
from .semantics import infer_annotation_blocks, infer_table_structure
from .sheet_scan import SheetProfile, StyleRegistry, cell_snapshot, scan_sheet
from .utils import (
    canonical_json,
    normalize_whitespace,
    safe_csv_text,
    sha256_file,
    slugify,
    stable_hash,
)


SCHEMA_VERSION = "1.0.0"
_SHEET_REF_RE = re.compile(r"(?:'((?:[^']|'')+)'|([A-Za-z_][^!\[\]]*))!")


class ExtractionError(RuntimeError):
    """Raised for a controlled extraction failure."""


class _DigestSet:
    """Incremental deterministic fingerprints for a stream of records."""

    def __init__(self) -> None:
        self._digests = {
            "layout": sha256(),
            "formula": sha256(),
            "style": sha256(),
            "content": sha256(),
        }

    def add(self, kind: str, value: Any) -> None:
        digest = self._digests[kind]
        digest.update(canonical_json(value).encode("utf-8"))
        digest.update(b"\n")

    def values(self) -> dict[str, str]:
        return {f"{name}_sha256": digest.hexdigest() for name, digest in self._digests.items()}


def _warning_dict(value: WarningRecord | Mapping[str, Any] | str) -> dict[str, Any]:
    if isinstance(value, WarningRecord):
        return value.to_dict()
    if isinstance(value, Mapping):
        return dict(value)
    return {"code": "EXTRACTION_WARNING", "message": str(value), "severity": "warning"}


def _deduplicate_warnings(values: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """Keep the first deterministic summary for the same warning locus."""

    result: list[dict[str, Any]] = []
    seen: set[tuple[object, ...]] = set()
    for value in values:
        warning = dict(value)
        key = (
            warning.get("code"),
            warning.get("sheet"),
            warning.get("coordinate"),
            warning.get("table_id"),
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(warning)
    return result


def _range_text(value: object) -> str:
    return getattr(value, "ref", str(value))


def _load_openpyxl_workbook(path: Path, *, data_only: bool) -> Any:
    suffix = path.suffix.casefold()
    keep_vba = suffix in {".xlsm", ".xltm"}
    try:
        return load_workbook(
            filename=path,
            read_only=False,
            data_only=data_only,
            keep_vba=keep_vba,
            keep_links=True,
            rich_text=True,
        )
    except TypeError:
        # `rich_text` was introduced before the supported minimum, but this
        # fallback produces a clearer experience with vendor-patched builds.
        return load_workbook(
            filename=path,
            read_only=False,
            data_only=data_only,
            keep_vba=keep_vba,
            keep_links=True,
        )


def _profile_value(profile: SheetProfile, name: str, default: Any) -> Any:
    return getattr(profile, name, default)


def _serialize_candidate(candidate: TableCandidate) -> dict[str, Any]:
    return {
        "range": candidate.bounds.ref,
        "methods": sorted(candidate.methods),
        "confidence": round(float(candidate.confidence), 6),
        "confidence_features": dict(sorted(candidate.features.items())),
        "detection_reasons": list(candidate.reasons),
        "source_name": candidate.source_name,
        "forced": candidate.forced,
        "uncertain": candidate.uncertain,
    }


def _descriptor_dict(value: Any) -> dict[str, Any]:
    if is_dataclass(value):
        return asdict(value)
    return dict(value)


def _candidate_table_id(sheet: str, candidate: TableCandidate, structure: Mapping[str, Any]) -> str:
    identity = {
        "sheet": sheet,
        "range": candidate.bounds.ref,
        "title": structure.get("title"),
        "layout_kind": structure.get("layout_kind"),
        "source_name": candidate.source_name,
    }
    return "tbl_" + stable_hash(identity)[:16]


def _bounds_distance(left: Bounds, right: Bounds) -> int:
    row_gap = max(0, left.min_row - right.max_row, right.min_row - left.max_row)
    col_gap = max(0, left.min_col - right.max_col, right.min_col - left.max_col)
    return row_gap + col_gap


def _cell_id(sheet: str, coordinate: str) -> str:
    return "cell_" + stable_hash({"sheet": sheet, "coordinate": coordinate})[:18]


def _normalize_snapshot(
    snapshot: Mapping[str, Any], *, sheet_name: str, sheet_xml: str | None
) -> dict[str, Any]:
    """Normalize the scanner's rich record into the versioned public schema."""

    result = dict(snapshot)
    formula_metadata = snapshot.get("formula")
    if isinstance(formula_metadata, Mapping):
        formula_text = (
            formula_metadata.get("exact")
            or formula_metadata.get("resolved")
            or formula_metadata.get("raw")
        )
        result["formula_metadata"] = dict(formula_metadata)
        result["formula"] = formula_text
        result["formula_signature"] = formula_metadata.get(
            "normalized_relative_signature"
        )
        result["formula_dependencies"] = formula_metadata.get("dependencies", [])
    elif formula_metadata:
        result["formula"] = str(formula_metadata)
        result.setdefault("formula_metadata", {"exact": str(formula_metadata)})
    else:
        result["formula"] = None
        result.setdefault("formula_metadata", None)
    result["formula_sha256"] = (
        stable_hash(result["formula"]) if result.get("formula") else None
    )
    result["formula_signature_sha256"] = (
        stable_hash(result.get("formula_signature"))
        if result.get("formula_signature")
        else None
    )
    result["data_type"] = snapshot.get("excel_data_type")
    result["literal_value_type"] = snapshot.get("literal_type", "blank")
    result["cached_value_type"] = snapshot.get("cached_type", "blank")
    cache_status = str(snapshot.get("cache_status", "not_applicable"))
    result["cached_value_status"] = {
        "disabled": "not_requested",
        "not_loaded": "not_requested",
        "not_applicable": "not_applicable",
    }.get(cache_status, cache_status)
    style_hash = snapshot.get("style_hash")
    result["style_ref"] = f"style:{style_hash}" if style_hash else None
    result["merge_ref"] = snapshot.get("merge_range")
    result["visibility"] = {
        "row_hidden": bool(snapshot.get("row_hidden")),
        "column_hidden": bool(snapshot.get("column_hidden")),
    }
    row_dimension = snapshot.get("row_dimension") or {}
    column_dimension = snapshot.get("column_dimension") or {}
    result["outline"] = {
        "row_level": int(row_dimension.get("outline_level", 0) or 0),
        "column_level": int(column_dimension.get("outline_level", 0) or 0),
        "row_collapsed": bool(row_dimension.get("collapsed")),
        "column_collapsed": bool(column_dimension.get("collapsed")),
    }
    result["source"] = {
        "sheet_xml": sheet_xml,
        "coordinate": snapshot.get("coordinate"),
        "sheet": sheet_name,
    }
    return result


def _referenced_sheet_names(formula: str | None) -> set[str]:
    if not formula:
        return set()
    result: set[str] = set()
    for match in _SHEET_REF_RE.finditer(formula):
        quoted, bare = match.groups()
        name = quoted.replace("''", "'") if quoted is not None else (bare or "").strip()
        if name:
            result.add(name)
    return result


def _effective_value(snapshot: Mapping[str, Any]) -> Any:
    if snapshot.get("formula"):
        if snapshot.get("cached_value_status") == "present":
            return snapshot.get("cached_value")
        return None
    return snapshot.get("literal_value")


def _value_state(snapshot: Mapping[str, Any]) -> str:
    if snapshot.get("formula") and snapshot.get("cached_value_status") != "present":
        return "formula_cache_missing"
    value = _effective_value(snapshot)
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "zero" if value == 0 else "number"
    text = normalize_whitespace(value)
    folded = text.casefold()
    if text in {"-", "–", "—"}:
        return "dash"
    if folded in {"ns", "n/s", "n.a.", "n/a", "na"}:
        return "ns"
    if folded in {"xxx", "xxxx", "tbd", "to be completed"}:
        return "placeholder"
    if str(snapshot.get("data_type", "")) == "e" or text.startswith("#"):
        return "excel_error"
    return "text"


def _structure_columns(structure: Mapping[str, Any]) -> list[dict[str, Any]]:
    columns = structure.get("column_descriptors") or structure.get("columns") or []
    return [_descriptor_dict(value) for value in columns]


def _structure_rows(structure: Mapping[str, Any]) -> list[dict[str, Any]]:
    rows = structure.get("row_descriptors") or structure.get("rows") or []
    return [_descriptor_dict(value) for value in rows]


def _semantic_column_role(column: Mapping[str, Any]) -> bool:
    return str(column.get("role", "unknown")) in {
        "measure",
        "formula_measure",
        "variance",
        "unknown_measure",
    }


def _table_long_records(
    *,
    sheet_name: str,
    table_id: str,
    structure: Mapping[str, Any],
    bounds: Bounds,
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
) -> Iterator[dict[str, Any]]:
    columns = _structure_columns(structure)
    rows = _structure_rows(structure)
    row_map = {int(item["row"]): item for item in rows if item.get("row")}
    value_columns = [column for column in columns if _semantic_column_role(column)]
    if not value_columns:
        value_columns = [
            {
                "column": col,
                "letter": get_column_letter(col),
                "role": "unknown_measure",
                "key": f"column_{get_column_letter(col).casefold()}",
                "header_path": [],
                "header_sources": [],
            }
            for col in range(bounds.min_col + 1, bounds.max_col + 1)
        ]
    excluded_roles = {"title", "header", "section_header", "spacer", "note", "footnote"}
    for row in range(bounds.min_row, bounds.max_row + 1):
        if row not in row_map:
            # Rows omitted by semantic inference are title/header context, not
            # observations. Their raw cells remain in cells.jsonl.
            continue
        row_info = row_map.get(
            row,
            {
                "row": row,
                "role": "unknown",
                "label_path": [],
                "label_sources": [],
                "section_path": [],
                "metric_label": None,
                "is_total": False,
            },
        )
        if row_info.get("role") in excluded_roles:
            continue
        for column_info in value_columns:
            col = int(column_info["column"])
            snapshot = snapshots.get((row, col))
            if snapshot is None:
                continue
            yield {
                "table_id": table_id,
                "cell_id": snapshot["cell_id"],
                "source_sheet": sheet_name,
                "source_cell": snapshot["coordinate"],
                "source_row": row,
                "source_column": col,
                "row_role": row_info.get("role"),
                "section_path": row_info.get("section_path", []),
                "row_label_path": row_info.get("label_path", []),
                "row_label_sources": row_info.get("label_sources", []),
                "metric_label": row_info.get("metric_label"),
                "is_total": bool(row_info.get("is_total")),
                "is_placeholder_row": bool(row_info.get("is_placeholder")),
                "column_role": column_info.get("role"),
                "column_key": column_info.get("key"),
                "column_header_path": column_info.get("header_path", []),
                "column_header_sources": column_info.get("header_sources", []),
                "measure_group": column_info.get("group"),
                "measure": column_info.get("measure"),
                "unit": column_info.get("unit"),
                "scenario": column_info.get("scenario"),
                "year": column_info.get("year"),
                "comparison_kind": column_info.get("comparison_kind"),
                "value": _effective_value(snapshot),
                "value_type": snapshot.get("cached_value_type")
                if snapshot.get("formula")
                else snapshot.get("literal_value_type"),
                "value_state": _value_state(snapshot),
                "literal_value": snapshot.get("literal_value"),
                "formula": snapshot.get("formula"),
                "cached_value": snapshot.get("cached_value"),
                "cached_value_status": snapshot.get("cached_value_status"),
                "number_format": snapshot.get("number_format"),
                "style_ref": snapshot.get("style_ref"),
                "row_hidden": snapshot.get("visibility", {}).get("row_hidden"),
                "column_hidden": snapshot.get("visibility", {}).get("column_hidden"),
                "interpretation_confidence": min(
                    float(row_info.get("confidence", 0.0) or 0.0),
                    float(column_info.get("confidence", 0.0) or 0.0),
                ),
            }


def _csv_record(record: Mapping[str, Any], safe: bool) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in record.items():
        if isinstance(value, (list, dict)):
            value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        if safe:
            value = safe_csv_text(value)
        result[key] = value
    return result


def _snapshot_preview_value(snapshot: Mapping[str, Any]) -> Any:
    value = _effective_value(snapshot)
    if value is None and snapshot.get("formula"):
        return "∅ cache"
    return value


def _table_preview(
    table: Mapping[str, Any],
    structure: Mapping[str, Any],
    snapshots: Mapping[tuple[int, int], dict[str, Any]],
    config: ExtractionConfig,
) -> dict[str, Any]:
    bounds = Bounds.from_a1(str(table["range"]))
    columns = _structure_columns(structure)
    column_map = {int(item["column"]): item for item in columns if item.get("column")}
    shown_cols = list(
        range(bounds.min_col, min(bounds.max_col, bounds.min_col + config.preview_columns - 1) + 1)
    )
    preview_columns: list[dict[str, Any]] = []
    for col in shown_cols:
        descriptor = column_map.get(col, {})
        header_path = list(descriptor.get("header_path", []))
        if (
            len(header_path) > 1
            and normalize_whitespace(header_path[0]).casefold()
            == normalize_whitespace(table.get("title", "")).casefold()
        ):
            header_path = header_path[1:]
        preview_columns.append(
            {
                "column": col,
                "letter": get_column_letter(col),
                "key": descriptor.get("key", ""),
                "header_path": descriptor.get("header_path", []),
                "label": " / ".join(header_path) or descriptor.get("key", ""),
            }
        )
    row_map = {int(item["row"]): item for item in _structure_rows(structure) if item.get("row")}
    rows: list[dict[str, Any]] = []
    for row in range(
        bounds.min_row,
        min(bounds.max_row, bounds.min_row + config.preview_rows - 1) + 1,
    ):
        cells: list[dict[str, Any]] = []
        for col in shown_cols:
            snapshot = snapshots.get((row, col), {})
            cells.append(
                {
                    "coordinate": snapshot.get(
                        "coordinate", f"{get_column_letter(col)}{row}"
                    ),
                    "value": _snapshot_preview_value(snapshot),
                    "formula": snapshot.get("formula"),
                    "cached": snapshot.get("cached_value"),
                }
            )
        rows.append(
            {
                "row": row,
                "role": row_map.get(row, {}).get("role", "unknown"),
                "cells": cells,
            }
        )
    truncated = bounds.height > config.preview_rows or bounds.width > config.preview_columns
    return {
        **{key: value for key, value in table.items() if key != "structure"},
        "columns": preview_columns,
        "preview": rows,
        "preview_note": (
            f"Preview limited to {config.preview_rows} rows × {config.preview_columns} columns; "
            "the JSON/CSV artifacts contain the complete extraction."
            if truncated
            else "Complete detected range shown."
        ),
    }


def _sheet_metadata(
    ws: Any,
    profile: SheetProfile,
    ooxml_sheet: Mapping[str, Any],
    tables: list[dict[str, Any]],
    annotations: list[AnnotationBlock],
    rejected: list[TableCandidate],
    sheet_warnings: list[dict[str, Any]],
) -> dict[str, Any]:
    row_dimensions = _profile_value(profile, "row_dimensions", [])
    if isinstance(row_dimensions, Mapping):
        row_dimensions = [
            row_dimensions[key]
            for key in sorted(row_dimensions, key=lambda value: int(value))
        ]
    column_dimensions = _profile_value(profile, "column_dimensions", [])
    if isinstance(column_dimensions, Mapping):
        column_dimensions = [
            column_dimensions[key]
            for key in sorted(column_dimensions, key=lambda value: str(value))
        ]
    return {
        "name": ws.title,
        "index": ooxml_sheet.get("index"),
        "state": ws.sheet_state,
        "xml_part": ooxml_sheet.get("part"),
        "stored_dimension": ooxml_sheet.get("stored_dimension"),
        "content_bounds": getattr(_profile_value(profile, "content_bounds", None), "ref", None),
        "style_bounds": getattr(_profile_value(profile, "style_bounds", None), "ref", None),
        "evidence_bounds": getattr(_profile_value(profile, "evidence_bounds", None), "ref", None),
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else None,
        "print_area": str(ws.print_area) if ws.print_area else None,
        "print_title_rows": ws.print_title_rows,
        "print_title_cols": ws.print_title_cols,
        "sheet_protection": {
            "enabled": bool(ws.protection.sheet),
            "format_cells": bool(ws.protection.formatCells),
            "insert_rows": bool(ws.protection.insertRows),
            "delete_rows": bool(ws.protection.deleteRows),
        },
        "merged_ranges": sorted(_range_text(value) for value in ws.merged_cells.ranges),
        "row_dimensions": row_dimensions,
        "column_dimensions": column_dimensions,
        "explicit_excel_tables": _profile_value(profile, "explicit_tables", []),
        "data_validations": _profile_value(profile, "validations", []),
        "conditional_formats": _profile_value(profile, "conditional_formats", []),
        "drawings": [],
        "detected_tables": tables,
        "annotations": [item.to_dict() for item in annotations],
        "rejected_regions": [_serialize_candidate(item) for item in rejected],
        "counts": dict(_profile_value(profile, "counts", {})),
        "warnings": sheet_warnings,
    }


def extract_workbook(
    source: str | Path,
    selected_sheets: list[str],
    output_dir: str | Path,
    *,
    config: ExtractionConfig | None = None,
    ooxml_index: OOXMLIndex | None = None,
    selection_method: str = "cli",
) -> Path:
    """Extract selected workbook sheets into an atomic audit bundle."""

    config = config or ExtractionConfig()
    source_path = Path(source).expanduser().resolve()
    output_path = Path(output_dir).expanduser().resolve()
    ooxml = ooxml_index or OOXMLIndex.open(source_path, config)
    catalog_by_name = {str(item["name"]): item for item in ooxml.sheet_catalog}
    unknown = [name for name in selected_sheets if name not in catalog_by_name]
    if unknown:
        raise ExtractionError(f"Selected sheets are absent from workbook: {unknown}")
    if not selected_sheets:
        raise ExtractionError("At least one sheet must be selected.")

    workbook_formula = _load_openpyxl_workbook(source_path, data_only=False)
    workbook_cached = (
        _load_openpyxl_workbook(source_path, data_only=True)
        if config.include_cached_values
        else None
    )
    defined_names = defined_names_metadata(workbook_formula)
    styles = StyleRegistry()
    extraction_time = datetime.now(timezone.utc).isoformat()
    source_stat = source_path.stat()
    source_info = {
        "path": str(source_path),
        "filename": source_path.name,
        "size_bytes": source_stat.st_size,
        "modified_at": datetime.fromtimestamp(
            source_stat.st_mtime, timezone.utc
        ).isoformat(),
        "sha256": sha256_file(source_path),
        "format": source_path.suffix.casefold().lstrip("."),
    }
    all_warnings: list[dict[str, Any]] = [_warning_dict(item) for item in ooxml.warnings]
    report_data: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "source": source_info,
        "sheets": [],
        "table_lookup": {},
        "warnings": all_warnings,
        "totals": {"sheets": 0, "tables": 0, "cells": 0, "formulas": 0},
    }
    sheet_manifests: list[dict[str, Any]] = []
    workbook_fingerprints: list[dict[str, Any]] = []
    selected_dependencies: set[str] = set()
    csv_fields = [
        "table_id", "cell_id", "source_sheet", "source_cell", "source_row",
        "source_column", "row_role", "section_path", "row_label_path",
        "row_label_sources", "metric_label", "is_total", "is_placeholder_row",
        "column_role", "column_key", "column_header_path", "column_header_sources",
        "measure_group", "measure", "unit", "scenario", "year", "comparison_kind",
        "value", "value_type", "value_state", "literal_value", "formula",
        "cached_value", "cached_value_status", "number_format", "style_ref",
        "row_hidden", "column_hidden", "interpretation_confidence",
    ]
    table_index_fields = [
        "table_id", "sheet", "title", "range", "layout_kind", "confidence",
        "uncertain", "methods", "rows", "columns", "long_records", "relative_path",
    ]

    try:
        with AtomicOutput(output_path) as bundle:
            long_jsonl = bundle.open_text("records_long.jsonl", newline="\n")
            long_csv = bundle.csv_writer("records_long.csv", csv_fields)
            table_index = bundle.csv_writer("tables_index.csv", table_index_fields)

            for sheet_index, sheet_name in enumerate(selected_sheets, 1):
                ws = workbook_formula[sheet_name]
                cached_ws = workbook_cached[sheet_name] if workbook_cached else None
                raw_formulas = ooxml.formula_records.get(sheet_name, {})
                profile = scan_sheet(
                    ws,
                    cached_ws,
                    raw_formulas,
                    config,
                    styles,
                )
                explicit_ranges = list(_profile_value(profile, "explicit_tables", []))
                named_ranges = static_named_ranges_for_sheet(defined_names, sheet_name)
                merged_ranges = [Bounds.from_a1(_range_text(value)) for value in ws.merged_cells.ranges]
                accepted, rejected = detect_table_candidates(
                    _profile_value(profile, "features", {}),
                    explicit_ranges,
                    named_ranges,
                    merged_ranges,
                    config,
                )
                structures: list[tuple[TableCandidate, dict[str, Any], str]] = []
                for candidate in accepted:
                    structure = infer_table_structure(
                        ws,
                        candidate,
                        _profile_value(profile, "features", {}),
                        _profile_value(profile, "merge_index", {}),
                    )
                    if "explicit-range" in candidate.methods and candidate.source_name:
                        # A native ListObject name is authoritative and more stable
                        # than guessing a title from one of its header cells.
                        structure = dict(structure)
                        structure["title"] = candidate.source_name
                    table_id = _candidate_table_id(sheet_name, candidate, structure)
                    structures.append((candidate, structure, table_id))
                annotations = infer_annotation_blocks(
                    _profile_value(profile, "features", {}), accepted, rejected
                )
                if structures:
                    for annotation in annotations:
                        nearest_candidate, _, nearest_id = min(
                            structures,
                            key=lambda item: (
                                _bounds_distance(annotation.bounds, item[0].bounds),
                                item[0].bounds.min_row,
                                item[0].bounds.min_col,
                            ),
                        )
                        annotation.nearest_table_id = nearest_id
                        annotation.distance = _bounds_distance(
                            annotation.bounds, nearest_candidate.bounds
                        )
                accepted_bounds = [candidate.bounds for candidate in accepted]
                annotation_bounds = [item.bounds for item in annotations]
                material: set[tuple[int, int]] = set(
                    _profile_value(profile, "material_coordinates", set())
                )
                dense_table_cells = sum(candidate.bounds.area for candidate in accepted)
                sheet_warnings = [
                    _warning_dict(item) for item in _profile_value(profile, "warnings", [])
                ]
                if dense_table_cells <= config.max_material_cells_per_sheet:
                    for candidate in accepted:
                        if candidate.bounds.area > config.max_table_area:
                            sheet_warnings.append(
                                WarningRecord(
                                    code="TABLE_DENSE_EXPORT_SKIPPED",
                                    message=(
                                        f"{candidate.bounds.ref} contains {candidate.bounds.area:,} cells; "
                                        "only stored/material cells were exported."
                                    ),
                                    sheet=sheet_name,
                                    coordinate=candidate.bounds.ref,
                                ).to_dict()
                            )
                            continue
                        material.update(
                            (row, col)
                            for row in range(candidate.bounds.min_row, candidate.bounds.max_row + 1)
                            for col in range(candidate.bounds.min_col, candidate.bounds.max_col + 1)
                        )
                else:
                    sheet_warnings.append(
                        WarningRecord(
                            code="SHEET_DENSE_EXPORT_SKIPPED",
                            message=(
                                f"Detected table rectangles cover {dense_table_cells:,} cells, above "
                                f"the configured {config.max_material_cells_per_sheet:,} limit."
                            ),
                            sheet=sheet_name,
                        ).to_dict()
                    )
                if len(material) > config.max_material_cells_per_sheet:
                    raise ExtractionError(
                        f"Sheet {sheet_name!r} has {len(material):,} material cells, exceeding "
                        f"the configured limit of {config.max_material_cells_per_sheet:,}."
                    )

                original_position = int(
                    catalog_by_name[sheet_name].get("position", sheet_index) or sheet_index
                )
                slug = f"{original_position:02d}_{slugify(sheet_name, 'sheet')}"
                sheet_root = Path("sheets") / slug
                snapshots: dict[tuple[int, int], dict[str, Any]] = {}
                digest = _DigestSet()
                formula_cells = 0
                missing_cache_cells: list[str] = []
                cells_handle = bundle.open_text(sheet_root / "cells.jsonl", newline="\n")
                unassigned_handle = bundle.open_text(sheet_root / "unassigned_cells.jsonl", newline="\n")
                for row, col in sorted(material):
                    snapshot = cell_snapshot(
                        ws,
                        cached_ws,
                        row,
                        col,
                        profile,
                        raw_formulas,
                        styles,
                    )
                    snapshot = _normalize_snapshot(
                        snapshot,
                        sheet_name=sheet_name,
                        sheet_xml=catalog_by_name[sheet_name].get("part"),
                    )
                    snapshot.setdefault("cell_id", _cell_id(sheet_name, snapshot["coordinate"]))
                    snapshot.setdefault("sheet", sheet_name)
                    snapshots[(row, col)] = snapshot
                    cells_handle.write(canonical_json(snapshot) + "\n")
                    formula = snapshot.get("formula")
                    if formula:
                        formula_cells += 1
                        selected_dependencies.update(_referenced_sheet_names(str(formula)))
                        if snapshot.get("cached_value_status") != "present":
                            missing_cache_cells.append(str(snapshot["coordinate"]))
                    digest.add(
                        "layout",
                        {
                            "coordinate": snapshot.get("coordinate"),
                            "data_type": snapshot.get("data_type"),
                            "merge_ref": snapshot.get("merge_ref"),
                            "visibility": snapshot.get("visibility"),
                            "outline": snapshot.get("outline"),
                        },
                    )
                    digest.add(
                        "formula",
                        {
                            "coordinate": snapshot.get("coordinate"),
                            "formula": snapshot.get("formula"),
                            "formula_signature": snapshot.get("formula_signature"),
                        },
                    )
                    digest.add(
                        "style",
                        {"coordinate": snapshot.get("coordinate"), "style_ref": snapshot.get("style_ref")},
                    )
                    digest.add(
                        "content",
                        {
                            "coordinate": snapshot.get("coordinate"),
                            "literal": snapshot.get("literal_value"),
                            "cached": snapshot.get("cached_value"),
                        },
                    )
                    in_table = any(bounds.contains(row, col) for bounds in accepted_bounds)
                    in_annotation = any(bounds.contains(row, col) for bounds in annotation_bounds)
                    if not in_table and not in_annotation:
                        unassigned_handle.write(
                            canonical_json(
                                {
                                    **snapshot,
                                    "unassigned_reason": "not_assigned_to_table_or_annotation",
                                }
                            )
                            + "\n"
                        )

                raw_formula_count = int(
                    catalog_by_name[sheet_name].get("formula_count", len(raw_formulas)) or 0
                )
                if raw_formula_count != formula_cells:
                    sheet_warnings.append(
                        WarningRecord(
                            code="FORMULA_COUNT_MISMATCH",
                            message=(
                                f"Raw OOXML has {raw_formula_count} formula cells; the exported cell layer "
                                f"has {formula_cells}."
                            ),
                            severity="error",
                            sheet=sheet_name,
                        ).to_dict()
                    )
                if missing_cache_cells:
                    sheet_warnings.append(
                        WarningRecord(
                            code="FORMULA_CACHE_MISSING",
                            message=(
                                f"{len(missing_cache_cells)} formulas have no stored workbook result. "
                                "The extractor did not calculate or substitute values."
                            ),
                            sheet=sheet_name,
                            details={"examples": missing_cache_cells[:20]},
                        ).to_dict()
                    )
                sheet_warnings = _deduplicate_warnings(sheet_warnings)

                table_entries: list[dict[str, Any]] = []
                report_sheet_tables: list[dict[str, Any]] = []
                for ordinal, (candidate, structure, table_id) in enumerate(structures, 1):
                    columns = _structure_columns(structure)
                    rows = _structure_rows(structure)
                    table_relative = sheet_root / "tables" / f"{ordinal:02d}_{table_id}"
                    table_entry = {
                        "table_id": table_id,
                        "template_key": structure.get("template_key")
                        or stable_hash(
                            {
                                "title": structure.get("title"),
                                "headers": [item.get("header_path", []) for item in columns],
                                "row_skeleton": [item.get("metric_label") for item in rows],
                            }
                        )[:24],
                        "sheet": sheet_name,
                        **_serialize_candidate(candidate),
                        "title": structure.get("title") or candidate.source_name or candidate.bounds.ref,
                        "layout_kind": structure.get("layout_kind", "matrix"),
                        "header_rows": structure.get("header_rows", []),
                        "columns": columns,
                        "rows": rows,
                        "sections": structure.get("sections", []),
                        "column_groups": structure.get("column_groups", []),
                        "annotations": [
                            item.to_dict()
                            for item in annotations
                            if item.nearest_table_id == table_id
                        ],
                        "fingerprints": structure.get("fingerprints", {}),
                        "warnings": structure.get("warnings", []),
                        "relative_path": str(table_relative).replace("\\", "/"),
                    }
                    long_count = 0
                    table_jsonl = bundle.open_text(table_relative / "records_long.jsonl", newline="\n")
                    table_csv = bundle.csv_writer(table_relative / "records_long.csv", csv_fields)
                    table_records = _table_long_records(
                        sheet_name=sheet_name,
                        table_id=table_id,
                        structure=structure,
                        bounds=candidate.bounds,
                        snapshots=snapshots,
                    )
                    for record in table_records:
                        encoded = canonical_json(record)
                        table_jsonl.write(encoded + "\n")
                        long_jsonl.write(encoded + "\n")
                        csv_row = _csv_record(record, config.csv_formula_injection_safe)
                        table_csv.writerow(csv_row)
                        long_csv.writerow(csv_row)
                        long_count += 1
                    table_entry["counts"] = {
                        "physical_rows": candidate.bounds.height,
                        "physical_columns": candidate.bounds.width,
                        "long_records": long_count,
                    }
                    bundle.write_json(table_relative / "table.json", table_entry)

                    # Wide review export retains a row-role and row-path prefix.
                    semantic_columns = [item for item in columns if _semantic_column_role(item)]
                    wide_fieldnames = ["source_row", "row_role", "section_path", "row_label_path"]
                    for item in semantic_columns:
                        key = str(item.get("key") or f"column_{item.get('letter')}")
                        if key not in wide_fieldnames:
                            wide_fieldnames.append(key)
                    wide_writer = bundle.csv_writer(table_relative / "values_wide.csv", wide_fieldnames)
                    row_map = {int(item["row"]): item for item in rows if item.get("row")}
                    for row in range(candidate.bounds.min_row, candidate.bounds.max_row + 1):
                        if row not in row_map:
                            continue
                        row_info = row_map.get(row, {})
                        if row_info.get("role") in {
                            "title",
                            "header",
                            "section_header",
                            "spacer",
                            "note",
                            "footnote",
                        }:
                            continue
                        wide_row: dict[str, Any] = {
                            "source_row": row,
                            "row_role": row_info.get("role", "unknown"),
                            "section_path": json.dumps(row_info.get("section_path", []), ensure_ascii=False),
                            "row_label_path": json.dumps(row_info.get("label_path", []), ensure_ascii=False),
                        }
                        for item in semantic_columns:
                            col = int(item["column"])
                            key = str(item.get("key") or f"column_{item.get('letter')}")
                            value = _effective_value(snapshots.get((row, col), {}))
                            wide_row[key] = safe_csv_text(value) if config.csv_formula_injection_safe else value
                        wide_writer.writerow(wide_row)

                    table_entries.append(table_entry)
                    table_index.writerow(
                        {
                            "table_id": table_id,
                            "sheet": sheet_name,
                            "title": table_entry["title"],
                            "range": candidate.bounds.ref,
                            "layout_kind": table_entry["layout_kind"],
                            "confidence": table_entry["confidence"],
                            "uncertain": table_entry["uncertain"],
                            "methods": "|".join(table_entry["methods"]),
                            "rows": candidate.bounds.height,
                            "columns": candidate.bounds.width,
                            "long_records": long_count,
                            "relative_path": table_entry["relative_path"],
                        }
                    )
                    report_table = _table_preview(table_entry, structure, snapshots, config)
                    report_data["table_lookup"][table_id] = report_table
                    report_sheet_tables.append(
                        {
                            "table_id": table_id,
                            "title": table_entry["title"],
                            "range": candidate.bounds.ref,
                            "confidence": candidate.confidence,
                            "uncertain": candidate.uncertain,
                        }
                    )

                sheet_meta = _sheet_metadata(
                    ws,
                    profile,
                    catalog_by_name[sheet_name],
                    table_entries,
                    annotations,
                    rejected,
                    sheet_warnings,
                )
                sheet_meta["drawings"] = ooxml.drawings_by_sheet.get(sheet_name, [])
                sheet_meta["fingerprints"] = digest.values()
                sheet_meta["counts"].update(
                    {
                        "exported_cells": len(snapshots),
                        "exported_formulas": formula_cells,
                        "detected_tables": len(table_entries),
                        "annotations": len(annotations),
                    }
                )
                bundle.write_json(sheet_root / "sheet.json", sheet_meta)
                sheet_manifests.append(
                    {
                        "name": sheet_name,
                        "state": ws.sheet_state,
                        "relative_path": str(sheet_root).replace("\\", "/"),
                        "counts": sheet_meta["counts"],
                        "fingerprints": sheet_meta["fingerprints"],
                        "tables": [
                            {
                                "table_id": item["table_id"],
                                "title": item["title"],
                                "range": item["range"],
                                "confidence": item["confidence"],
                            }
                            for item in table_entries
                        ],
                    }
                )
                workbook_fingerprints.append(
                    {"sheet": sheet_name, **sheet_meta["fingerprints"]}
                )
                all_warnings.extend(sheet_warnings)
                report_data["sheets"].append(
                    {
                        "name": sheet_name,
                        "state": ws.sheet_state,
                        "tables": report_sheet_tables,
                    }
                )
                report_data["totals"]["sheets"] += 1
                report_data["totals"]["tables"] += len(table_entries)
                report_data["totals"]["cells"] += len(snapshots)
                report_data["totals"]["formulas"] += formula_cells

            missing_dependency_sheets = sorted(
                name
                for name in selected_dependencies
                if name in catalog_by_name and name not in set(selected_sheets)
            )
            if missing_dependency_sheets:
                all_warnings.append(
                    WarningRecord(
                        code="DEPENDENCY_SHEET_NOT_SELECTED",
                        message=(
                            "Selected formulas reference excluded sheets: "
                            + ", ".join(missing_dependency_sheets)
                        ),
                        details={"sheets": missing_dependency_sheets},
                    ).to_dict()
                )
            all_warnings = _deduplicate_warnings(all_warnings)

            manifest = {
                "schema_version": SCHEMA_VERSION,
                "extracted_at": extraction_time,
                "source": source_info,
                "selection": {
                    "method": selection_method,
                    "selected_sheets": selected_sheets,
                    "unselected_sheets": [
                        str(item["name"])
                        for item in ooxml.sheet_catalog
                        if str(item["name"]) not in set(selected_sheets)
                    ],
                    "referenced_but_unselected_sheets": missing_dependency_sheets,
                },
                "engine": {
                    "name": "pops-ingest",
                    "version": __version__,
                    "python": platform.python_version(),
                    "openpyxl": openpyxl.__version__,
                    "configuration": config.to_dict(),
                    "configuration_sha256": config.fingerprint,
                },
                "workbook": {
                    **workbook_metadata(workbook_formula),
                    "sheet_catalog": ooxml.sheet_catalog,
                    "defined_names": defined_names,
                    "external_relationships": ooxml.external_relationships,
                    "package": ooxml.package_manifest,
                },
                "sheets": sheet_manifests,
                "styles_file": "styles.json",
                "schema_file": "schema.json",
                "records_long_jsonl": "records_long.jsonl",
                "records_long_csv": "records_long.csv",
                "tables_index_csv": "tables_index.csv",
                "report_html": "report.html",
                "fingerprints": {
                    "structure_sha256": stable_hash(
                        [
                            {
                                "sheet": item["sheet"],
                                "layout": item["layout_sha256"],
                                "formula": item["formula_sha256"],
                                "style": item["style_sha256"],
                            }
                            for item in workbook_fingerprints
                        ]
                    ),
                    "content_sha256": stable_hash(
                        [
                            {"sheet": item["sheet"], "content": item["content_sha256"]}
                            for item in workbook_fingerprints
                        ]
                    ),
                },
                "warnings": all_warnings,
            }
            report_data["warnings"] = all_warnings
            bundle.write_json("styles.json", styles.to_dict())
            bundle.write_json("schema.json", BUNDLE_SCHEMA)
            bundle.write_json("manifest.json", manifest)
            build_html_report(report_data, bundle.path("report.html"))
            final = bundle.commit()
    finally:
        try:
            workbook_formula.close()
        finally:
            if workbook_cached is not None:
                workbook_cached.close()
    return final
