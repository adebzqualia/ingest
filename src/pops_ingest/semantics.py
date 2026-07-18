"""Infer logical POPS table semantics from detected worksheet regions.

This module never mutates the source worksheet.  It creates a conservative,
coordinate-backed interpretation on top of the lossless cell feature layer:
multi-row header paths, row and column roles, vertical sections, horizontal
column groups, normalized period/scenario metadata, and structural
fingerprints.  Raw labels and source coordinates remain present whenever a
normalization is inferred.
"""

from __future__ import annotations

from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import asdict
import re
from typing import Any, TypeAlias

from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    AnnotationBlock,
    Bounds,
    CellFeature,
    ColumnDescriptor,
    RowDescriptor,
    TableCandidate,
)
from .utils import (
    normalize_whitespace,
    slugify,
    stable_hash,
    unique_preserving_order,
)


FeatureCollection: TypeAlias = (
    Mapping[object, CellFeature] | Iterable[CellFeature]
)


_YEAR_RE = re.compile(r"(?<!\d)(20\d{2})(?!\d)")
_COMPARISON_RE = re.compile(
    r"(?:\bvs\.?\b|\bversus\b|\bvariance\b|\bvar\.?\b|"
    r"\bdelta\b|[Δ∆])",
    re.IGNORECASE,
)
_PERIOD_OR_SCENARIO_RE = re.compile(
    r"(?:\b20\d{2}\b|\bactual\b|\bbaseline\b|\bbudget\b|"
    r"\bforecast\b|\bplan\b|\btarget\b|\bf\d{1,2}\b|"
    r"\bb(?=\s*20\d{2})|\bdeep[ -]?dive\b|\bvs\.?\b|"
    r"\bversus\b|\bvariance\b|\bdelta\b|[Δ∆])",
    re.IGNORECASE,
)
_HEADER_SEMANTIC_RE = re.compile(
    r"(?:\b20\d{2}\b|\b(?:actual|baseline|budget|forecast|plan|target)\b|"
    r"\bf\d{1,2}\b|\bb\s*20\d{2}\b|\b(?:fte|opex|kpi)\b|"
    r"(?:\bvs\.?\b|\bversus\b|\bvariance\b|\bdelta\b|[Δ∆])|"
    r"(?:m|k)?€|%|\b(?:nb|number|count)\b)",
    re.IGNORECASE,
)
_TOTAL_RE = re.compile(
    r"(?:^|\b)(?:grand\s+total|sub\s*total|subtotal|total)(?:\b|$)",
    re.IGNORECASE,
)
_NOTE_RE = re.compile(
    r"(?:^\s*(?:note|source|sourcing|definition|where|please|\*|\(\*)|"
    r"copy\s+table|do\s+not\s+copy|manual(?:ly)?\s+enter)",
    re.IGNORECASE,
)
_GENERIC_BANNER_RE = re.compile(
    r"(?:copy\s+table|copy\s+.*powerpoint|please\s+ensure)",
    re.IGNORECASE,
)
_INSTRUCTION_RE = re.compile(
    r"(?:please|ensure|must|enter|complete|copy|do\s+not|manually)",
    re.IGNORECASE,
)
_SOURCE_RE = re.compile(
    r"(?:source|sourcing|ilon|perseus|https?://|equal\s+to\s+figures)",
    re.IGNORECASE,
)


def _feature_map(features: FeatureCollection) -> dict[tuple[int, int], CellFeature]:
    values: Iterable[CellFeature]
    if isinstance(features, Mapping):
        values = features.values()
    else:
        values = features
    result: dict[tuple[int, int], CellFeature] = {}
    for feature in values:
        if not isinstance(feature, CellFeature):
            raise TypeError(
                "features must contain CellFeature instances; "
                f"got {type(feature)!r}"
            )
        result[(feature.row, feature.col)] = feature
    return result


def _coerce_bounds(value: object) -> Bounds:
    if isinstance(value, Bounds):
        return value
    if isinstance(value, str):
        ref = value.rsplit("!", 1)[-1].replace("$", "").strip().strip("'")
        return Bounds.from_a1(ref)
    if isinstance(value, Mapping):
        for key in ("bounds", "range", "ref", "sqref"):
            if key in value:
                return _coerce_bounds(value[key])
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes)):
        if len(value) == 4 and all(isinstance(item, int) for item in value):
            return Bounds(*value)
    if all(
        hasattr(value, attr)
        for attr in ("min_row", "min_col", "max_row", "max_col")
    ):
        return Bounds(
            int(getattr(value, "min_row")),
            int(getattr(value, "min_col")),
            int(getattr(value, "max_row")),
            int(getattr(value, "max_col")),
        )
    if hasattr(value, "ref"):
        return _coerce_bounds(getattr(value, "ref"))
    raise TypeError(f"Unsupported range representation: {value!r}")


class _MergeLookup:
    """Candidate-local merged-range lookup with deterministic precedence."""

    def __init__(self, ranges: Iterable[Bounds]) -> None:
        unique = {item.ref: item for item in ranges}
        self.ranges = sorted(unique.values(), key=lambda item: (item.area, item))

    def at(self, row: int, col: int) -> Bounds | None:
        for bounds in self.ranges:
            if bounds.contains(row, col):
                return bounds
        return None


def _merge_lookup(
    ws: Worksheet,
    merge_index: object,
    candidate_bounds: Bounds,
) -> _MergeLookup:
    values: list[object] = []
    if isinstance(merge_index, Mapping):
        values.extend(merge_index.values())
        # Some scanners map a merge reference to its anchor rather than the
        # inverse.  Range-like keys are therefore useful fallback inputs.
        values.extend(
            key
            for key in merge_index
            if isinstance(key, str) and ":" in key
        )
    elif isinstance(merge_index, Iterable) and not isinstance(
        merge_index, (str, bytes)
    ):
        values.extend(merge_index)
    elif merge_index is not None:
        values.append(merge_index)

    try:
        values.extend(ws.merged_cells.ranges)
    except AttributeError:
        pass

    ranges: list[Bounds] = []
    for value in values:
        try:
            bounds = _coerce_bounds(value)
        except (TypeError, ValueError):
            continue
        if bounds.intersects(candidate_bounds):
            ranges.append(bounds)
    return _MergeLookup(ranges)


def _direct_text(
    ws: Worksheet,
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    row: int,
    col: int,
) -> str:
    feature = feature_by_coord.get((row, col))
    if feature and feature.text is not None:
        text = normalize_whitespace(feature.text)
        if text:
            return text
    try:
        value = ws.cell(row=row, column=col).value
    except Exception:
        return ""
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith("="):
        return ""
    return normalize_whitespace(value)


def _effective_text(
    ws: Worksheet,
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    merges: _MergeLookup,
    row: int,
    col: int,
) -> tuple[str, str | None]:
    merge = merges.at(row, col)
    source_row, source_col = (merge.min_row, merge.min_col) if merge else (row, col)
    text = _direct_text(ws, feature_by_coord, source_row, source_col)
    if not text:
        return "", None
    source = f"{get_column_letter(source_col)}{source_row}"
    return text, source


def _row_features(
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    bounds: Bounds,
) -> dict[int, list[CellFeature]]:
    result: dict[int, list[CellFeature]] = defaultdict(list)
    for (row, col), feature in feature_by_coord.items():
        if bounds.contains(row, col):
            result[row].append(feature)
    for features in result.values():
        features.sort(key=lambda item: item.col)
    return dict(result)


def _header_row_score(
    row: int,
    bounds: Bounds,
    row_features: Sequence[CellFeature],
    texts: Sequence[str],
    merges: _MergeLookup,
) -> float:
    material = [
        feature
        for feature in row_features
        if feature.structural or feature.has_fill or feature.non_general_format
    ]
    numeric = sum(feature.numeric for feature in row_features)
    formulas = sum(feature.has_formula for feature in row_features)
    semantic = sum(bool(_HEADER_SEMANTIC_RE.search(text)) for text in texts)
    styled = sum(
        feature.bold or feature.has_fill or bool(feature.merged_range)
        for feature in row_features
    )
    merged_spans = sum(
        1
        for col in range(bounds.min_col, bounds.max_col + 1)
        if (merge := merges.at(row, col)) is not None and merge.width > 1
    )

    score = 0.0
    if texts:
        score += 0.18
    score += 0.42 * semantic / max(1, len(texts))
    score += 0.20 * styled / max(1, len(material))
    if merged_spans:
        score += 0.14
    if row == bounds.min_row:
        score += 0.08
    if numeric + formulas >= 2 and texts:
        # A metric label containing "%", "KPI", or a unit is still a body
        # row when followed by a numeric/formula series.  Header semantics must
        # not keep such rows in a growing multi-row header band.
        score -= 0.38
    if not texts and numeric + formulas:
        score -= 0.30
    return max(0.0, min(1.0, score))


def _infer_header_rows(
    ws: Worksheet,
    bounds: Bounds,
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    rows: Mapping[int, Sequence[CellFeature]],
    merges: _MergeLookup,
) -> tuple[list[int], list[str]]:
    warnings: list[str] = []
    max_scan = min(bounds.max_row, bounds.min_row + min(9, max(2, bounds.height // 3)))
    scores: dict[int, float] = {}
    has_content: dict[int, bool] = {}
    texts_by_row: dict[int, list[str]] = {}
    full_width_band: dict[int, bool] = {}
    for row in range(bounds.min_row, max_scan + 1):
        texts = unique_preserving_order(
            text
            for col in range(bounds.min_col, bounds.max_col + 1)
            if (text := _effective_text(
                ws, feature_by_coord, merges, row, col
            )[0])
        )
        texts_by_row[row] = texts
        full_width_band[row] = bool(
            len(texts) == 1
            and any(
                merge.min_col <= bounds.min_col
                and merge.max_col >= bounds.max_col
                and merge.min_row <= row <= merge.max_row
                for merge in merges.ranges
            )
            and not any(
                feature.numeric or feature.has_formula
                for feature in rows.get(row, ())
            )
        )
        scores[row] = _header_row_score(
            row, bounds, rows.get(row, ()), texts, merges
        )
        has_content[row] = bool(texts or rows.get(row))

    header_rows: list[int] = []
    started = False
    for row in range(bounds.min_row, max_scan + 1):
        score = scores[row]
        previous_texts = texts_by_row.get(row - 1, [])
        previous_is_leaf_header = bool(
            len(previous_texts) >= 2
            or sum(
                bool(_HEADER_SEMANTIC_RE.search(text))
                for text in previous_texts
            )
            >= 2
        )
        if (
            started
            and header_rows
            and full_width_band.get(row, False)
            and previous_is_leaf_header
        ):
            # A full-width, single-label band immediately below the leaf
            # headers starts the first table section; it is not another header
            # level.  Keeping it in the header path would incorrectly append
            # labels such as "TRANSFORMATION KPI'S" to every value column.
            break
        if score >= 0.36:
            if not started and row > bounds.min_row + 1:
                break
            started = True
            header_rows.append(row)
            continue
        if started:
            next_score = scores.get(row + 1, 0.0)
            if not has_content[row] and next_score >= 0.36:
                continue
            break
        if row >= bounds.min_row + 1:
            break

    if not header_rows:
        first_texts = [
            _effective_text(ws, feature_by_coord, merges, bounds.min_row, col)[0]
            for col in range(bounds.min_col, bounds.max_col + 1)
        ]
        if any(first_texts) and bounds.height > 1:
            header_rows = [bounds.min_row]
            warnings.append("Header row inferred weakly from the first table row.")
        else:
            warnings.append("No reliable header row was detected.")
    return header_rows, warnings


def _title_score(
    text: str,
    feature: CellFeature | None,
    merge: Bounds | None,
    bounds: Bounds,
    row: int,
) -> float:
    if not text or _GENERIC_BANNER_RE.search(text):
        return -1.0
    score = 0.15
    if not _PERIOD_OR_SCENARIO_RE.fullmatch(text):
        score += 0.20
    if feature and (feature.bold or feature.has_fill):
        score += 0.22
    if merge:
        score += 0.28 * min(1.0, merge.width / max(1, bounds.width))
    if 3 <= len(text) <= 90:
        score += 0.10
    score += 0.08 / (1 + row - bounds.min_row)
    if _HEADER_SEMANTIC_RE.fullmatch(text):
        score -= 0.25
    return score


def _infer_title(
    ws: Worksheet,
    bounds: Bounds,
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    merges: _MergeLookup,
    header_rows: Sequence[int],
) -> str | None:
    scan_rows = list(header_rows) or list(
        range(bounds.min_row, min(bounds.max_row, bounds.min_row + 4) + 1)
    )
    candidates: list[tuple[float, int, int, str]] = []
    seen_sources: set[str] = set()
    for row in scan_rows:
        for col in range(bounds.min_col, bounds.max_col + 1):
            text, source = _effective_text(
                ws, feature_by_coord, merges, row, col
            )
            if not text or not source or source in seen_sources:
                continue
            seen_sources.add(source)
            merge = merges.at(row, col)
            feature = feature_by_coord.get((row, col))
            candidates.append(
                (_title_score(text, feature, merge, bounds, row), row, col, text)
            )
    if not candidates:
        return None

    def is_period_only(text: str) -> bool:
        remainder = _PERIOD_OR_SCENARIO_RE.sub(" ", text)
        remainder = re.sub(
            r"(?:\(?(?:m|k)?\s*(?:€|eur)\)?|%|[Δ∆]|[-–—:|()/\[\]])",
            " ",
            remainder,
            flags=re.IGNORECASE,
        )
        return not normalize_whitespace(remainder)

    plausible = [
        item for item in candidates if item[0] > 0 and not is_period_only(item[3])
    ]
    if not plausible:
        return None
    # POPS section bands are often wider and visually stronger than the table
    # title.  The title is nevertheless the earliest, leftmost plausible
    # merged/header label (for example "OBS KPIs EVOLUTION" or "BUY & SHARE").
    # Row/column precedence therefore comes before minor style-score gains.
    earliest_row = min(item[1] for item in plausible)
    earliest = [item for item in plausible if item[1] == earliest_row]
    earliest.sort(key=lambda item: (item[2], -item[0], item[3]))
    return earliest[0][3]


def _is_parent_header(text: str) -> bool:
    """Return whether a styled header may span following blank cells."""

    if not text or _YEAR_RE.search(text) or _COMPARISON_RE.search(text):
        return False
    if re.search(r"\b(?:actual|budget|forecast|plan|f\d{1,2})\b", text, re.I):
        return False
    return len(text) <= 70


def _build_header_grid(
    ws: Worksheet,
    bounds: Bounds,
    header_rows: Sequence[int],
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    merges: _MergeLookup,
) -> dict[int, dict[int, tuple[str, str | None]]]:
    grid: dict[int, dict[int, tuple[str, str | None]]] = {}
    for row in header_rows:
        row_grid = {
            col: _effective_text(ws, feature_by_coord, merges, row, col)
            for col in range(bounds.min_col, bounds.max_col + 1)
        }
        nonblank = [col for col, (text, _) in row_grid.items() if text]
        for left, right in zip(nonblank, nonblank[1:]):
            if right - left <= 1 or right - left > 7:
                continue
            text, source = row_grid[left]
            source_feature = feature_by_coord.get((row, left))
            if (
                not source_feature
                or not _is_parent_header(text)
                or not (source_feature.bold or source_feature.has_fill)
                or not source_feature.style_hash
            ):
                continue
            blank_cols = list(range(left + 1, right))
            blank_features = [feature_by_coord.get((row, col)) for col in blank_cols]
            if blank_features and all(
                feature is not None
                and feature.style_hash == source_feature.style_hash
                and not row_grid[col][0]
                for col, feature in zip(blank_cols, blank_features)
            ):
                for col in blank_cols:
                    row_grid[col] = (text, source)
        grid[row] = row_grid
    return grid


def _parse_scenario(path: Sequence[str]) -> str | None:
    joined = " | ".join(path)
    leaf = path[-1] if path else ""
    if _COMPARISON_RE.search(joined):
        return "comparison"
    patterns = (
        ("actual", r"\b(?:actual|act\.?|réel|reel)\b"),
        ("forecast", r"\b(?:forecast|f\d{1,2})\b"),
        ("plan", r"\bplan\b"),
        ("budget", r"\b(?:budget|b(?=\s*20\d{2}))\b"),
        ("baseline", r"\bbaseline\b"),
        ("target", r"\btarget\b"),
    )
    for search_text in (leaf, joined):
        for scenario, pattern in patterns:
            if re.search(pattern, search_text, re.IGNORECASE):
                return scenario
    return None


def _parse_year(path: Sequence[str]) -> int | None:
    if not path:
        return None
    leaf_years = _YEAR_RE.findall(path[-1])
    if leaf_years:
        return int(leaf_years[0])
    all_years = _YEAR_RE.findall(" | ".join(path))
    return int(all_years[-1]) if all_years else None


def _parse_comparison(path: Sequence[str]) -> str | None:
    joined = " | ".join(path)
    if not _COMPARISON_RE.search(joined):
        return None
    return "percentage_variance" if "%" in joined else "variance"


def _parse_unit(path: Sequence[str], title: str | None) -> str | None:
    text = " | ".join([*path, title or ""])
    patterns = (
        ("%", r"%|\bpercent(?:age)?\b"),
        ("K€", r"(?:k\s*€|k\s*eur|keur)"),
        ("M€", r"(?:m\s*€|m\s*eur|meur|\(\s*mi\s*\))"),
        ("€", r"(?:€|\beur\b)"),
        ("FTE", r"\bfte\b"),
        ("nb", r"\b(?:nb|number|count)\b"),
    )
    for unit, pattern in patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return unit
    return None


def _parse_measure(path: Sequence[str], title: str | None) -> str | None:
    for part in reversed(path):
        cleaned = _PERIOD_OR_SCENARIO_RE.sub(" ", part)
        cleaned = re.sub(
            r"(?:\(?(?:m|k)?\s*(?:€|eur)\)?|%|\b(?:nb|number|count)\b)",
            " ",
            cleaned,
            flags=re.IGNORECASE,
        )
        cleaned = normalize_whitespace(cleaned).strip("-–—:|()[]")
        if len(cleaned) >= 2:
            return cleaned
    return title


def _normalized_parts(values: Sequence[str]) -> list[str]:
    return [normalize_whitespace(value).casefold() for value in values if value]


def _column_descriptors(
    bounds: Bounds,
    header_rows: Sequence[int],
    header_grid: Mapping[int, Mapping[int, tuple[str, str | None]]],
    rows: Mapping[int, Sequence[CellFeature]],
    merges: _MergeLookup,
    title: str | None,
) -> list[dict[str, Any]]:
    body_rows = [row for row in range(bounds.min_row, bounds.max_row + 1) if row not in header_rows]
    stats: dict[int, Counter[str]] = {
        col: Counter() for col in range(bounds.min_col, bounds.max_col + 1)
    }
    for row in body_rows:
        for feature in rows.get(row, ()):
            if not bounds.contains(feature.row, feature.col):
                continue
            if feature.text and not feature.numeric:
                stats[feature.col]["text"] += 1
                if len(normalize_whitespace(feature.text)) >= 80:
                    stats[feature.col]["long_text"] += 1
            if feature.numeric:
                stats[feature.col]["numeric"] += 1
            if feature.has_formula:
                stats[feature.col]["formula"] += 1
            if feature.border_edges:
                stats[feature.col]["border"] += 1
            if feature.validation_refs:
                stats[feature.col]["validation"] += 1
            merge = merges.at(feature.row, feature.col)
            if merge and merge.height > 1:
                stats[feature.col]["vertical_merge"] += 1

    paths: dict[int, list[str]] = {}
    sources: dict[int, list[str]] = {}
    parsed: dict[int, dict[str, Any]] = {}
    for col in range(bounds.min_col, bounds.max_col + 1):
        raw_path: list[str] = []
        raw_sources: list[str] = []
        for row in header_rows:
            text, source = header_grid.get(row, {}).get(col, ("", None))
            if text and (not raw_path or normalize_whitespace(raw_path[-1]).casefold() != normalize_whitespace(text).casefold()):
                raw_path.append(text)
            if source and source not in raw_sources:
                raw_sources.append(source)
        paths[col] = raw_path
        sources[col] = raw_sources
        parsed[col] = {
            "scenario": _parse_scenario(raw_path),
            "year": _parse_year(raw_path),
            "comparison_kind": _parse_comparison(raw_path),
            "unit": _parse_unit(raw_path, title),
            "measure": _parse_measure(raw_path, title),
        }

    value_columns = [
        col
        for col in range(bounds.min_col, bounds.max_col + 1)
        if stats[col]["numeric"]
        or stats[col]["formula"]
        or parsed[col]["year"] is not None
        or parsed[col]["scenario"] is not None
        or parsed[col]["comparison_kind"] is not None
    ]
    first_value_col = min(value_columns) if value_columns else None

    descriptors: list[dict[str, Any]] = []
    signature_occurrences: Counter[str] = Counter()
    for col in range(bounds.min_col, bounds.max_col + 1):
        column_stats = stats[col]
        comparison_kind = parsed[col]["comparison_kind"]
        before_values = first_value_col is not None and col < first_value_col
        if comparison_kind:
            role = "variance"
        elif before_values and column_stats["vertical_merge"]:
            role = "row_group"
        elif before_values and column_stats["text"] >= column_stats["numeric"]:
            role = "row_label"
        elif column_stats["long_text"] > column_stats["numeric"] + column_stats["formula"]:
            role = "annotation"
        elif column_stats["formula"] > column_stats["numeric"] and column_stats["formula"]:
            role = "formula_measure"
        elif col in value_columns:
            role = "measure"
        elif column_stats["text"]:
            role = "dimension"
        elif column_stats["border"] or column_stats["validation"]:
            role = "measure" if first_value_col is not None and col >= first_value_col else "unknown"
        else:
            role = "unknown"

        path = paths[col]
        semantic_signature = {
            "path": _normalized_parts(path),
            "role": role,
            "scenario": parsed[col]["scenario"],
            "year": parsed[col]["year"],
            "comparison_kind": comparison_kind,
            "unit": parsed[col]["unit"],
        }
        signature_text = stable_hash(semantic_signature)
        occurrence = signature_occurrences[signature_text]
        signature_occurrences[signature_text] += 1
        readable = "-".join(path) if path else f"column-{col - bounds.min_col + 1}"
        key = (
            f"{slugify(readable, fallback='column')}-"
            f"{stable_hash({'signature': semantic_signature, 'occurrence': occurrence})[:8]}"
        )

        evidence = bool(path) + bool(column_stats["numeric"] or column_stats["formula"])
        confidence = min(
            1.0,
            0.25
            + 0.25 * bool(path)
            + 0.20 * bool(column_stats["numeric"] or column_stats["formula"])
            + 0.20 * bool(parsed[col]["year"] or parsed[col]["scenario"] or comparison_kind)
            + 0.10 * bool(evidence and role != "unknown"),
        )
        descriptor = ColumnDescriptor(
            column=col,
            letter=get_column_letter(col),
            role=role,
            header_path=path,
            header_sources=sources[col],
            key=key,
            scenario=parsed[col]["scenario"],
            year=parsed[col]["year"],
            comparison_kind=comparison_kind,
            unit=parsed[col]["unit"],
            measure=parsed[col]["measure"],
            group=path[0] if len(path) > 1 else None,
            confidence=round(confidence, 6),
        )
        descriptors.append(asdict(descriptor))
    return descriptors


def _row_role(
    row: int,
    bounds: Bounds,
    row_features: Sequence[CellFeature],
    direct_texts: Sequence[str],
    labels: Sequence[str],
    merges: _MergeLookup,
) -> tuple[str, bool, bool, float]:
    numeric = sum(feature.numeric for feature in row_features)
    formulas = sum(feature.has_formula for feature in row_features)
    structural = sum(feature.structural for feature in row_features)
    bordered = sum(bool(feature.border_edges) for feature in row_features)
    styled = sum(feature.bold or feature.has_fill for feature in row_features)
    text = " | ".join(unique_preserving_order([*labels, *direct_texts]))
    primary = labels[-1] if labels else (direct_texts[0] if direct_texts else "")
    long_prose = bool(direct_texts) and (
        max(len(item) for item in direct_texts) >= 90
        or sum(len(item) for item in direct_texts) >= 150
    )
    is_total = bool(_TOTAL_RE.search(text))
    note_like = bool(_NOTE_RE.search(text)) or (
        long_prose and numeric + formulas == 0
    )
    merged_wide = False
    for feature in row_features:
        merge = merges.at(row, feature.col)
        if merge and merge.width >= max(2, int(bounds.width * 0.55)):
            merged_wide = True
            break
    uppercase_band = bool(primary) and len(primary) <= 70 and primary.isupper()
    section_like = (
        numeric + formulas == 0
        and 0 < len(unique_preserving_order(direct_texts)) <= 2
        and (merged_wide or uppercase_band or styled >= max(1, len(row_features) // 2))
    )
    group_like = (
        numeric + formulas == 0
        and bool(direct_texts)
        and styled > 0
        and len(direct_texts) <= 3
    )

    if note_like:
        return "note", False, False, 0.90
    if is_total:
        return "total", True, False, 0.92
    if section_like:
        return "section_header", False, False, 0.88
    if group_like:
        return "group_header", False, False, 0.72
    if numeric + formulas or labels:
        placeholder = numeric + formulas == 0 and structural > 0
        return "data", False, placeholder, 0.82 if numeric + formulas else 0.65
    if structural and bordered >= max(1, len(row_features) // 2):
        return "data", False, True, 0.58
    if not structural and not direct_texts:
        return "spacer", False, False, 0.95
    return "unknown", False, False, 0.35


def _row_descriptors(
    ws: Worksheet,
    bounds: Bounds,
    header_rows: Sequence[int],
    columns: Sequence[Mapping[str, Any]],
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    rows: Mapping[int, Sequence[CellFeature]],
    merges: _MergeLookup,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    label_columns = [
        int(column["column"])
        for column in columns
        if column["role"] in {"row_group", "row_label", "dimension"}
    ]
    if not label_columns:
        text_columns = sorted(
            {
                feature.col
                for feature in feature_by_coord.values()
                if bounds.contains(feature.row, feature.col)
                and feature.text
                and not feature.numeric
            }
        )
        label_columns = text_columns[:1]

    descriptors: list[dict[str, Any]] = []
    sections: list[dict[str, Any]] = []
    current_section: list[str] = []
    open_section: dict[str, Any] | None = None
    occurrences: Counter[str] = Counter()

    for row in range(bounds.min_row, bounds.max_row + 1):
        if row in header_rows:
            continue
        labels: list[str] = []
        label_sources: list[str] = []
        for col in label_columns:
            text, source = _effective_text(
                ws, feature_by_coord, merges, row, col
            )
            if text and (not labels or labels[-1].casefold() != text.casefold()):
                labels.append(text)
            if source and source not in label_sources:
                label_sources.append(source)

        direct_texts = unique_preserving_order(
            normalize_whitespace(feature.text)
            for feature in rows.get(row, ())
            if feature.text and normalize_whitespace(feature.text)
        )
        role, is_total, is_placeholder, confidence = _row_role(
            row,
            bounds,
            rows.get(row, ()),
            direct_texts,
            labels,
            merges,
        )
        metric_label = labels[-1] if labels else (direct_texts[0] if direct_texts else None)

        if role == "section_header" and metric_label:
            if open_section is not None:
                open_section["end_row"] = row - 1
                open_section["range"] = (
                    f"{get_column_letter(bounds.min_col)}{open_section['start_row']}:"
                    f"{get_column_letter(bounds.max_col)}{open_section['end_row']}"
                )
                sections.append(open_section)
            current_section = [metric_label]
            section_occurrence = sum(
                section["title"].casefold() == metric_label.casefold()
                for section in sections
            )
            open_section = {
                "key": (
                    f"{slugify(metric_label, fallback='section')}-"
                    f"{stable_hash({'title': metric_label.casefold(), 'occurrence': section_occurrence})[:8]}"
                ),
                "title": metric_label,
                "path": list(current_section),
                "start_row": row,
                "end_row": row,
                "range": None,
                "source": label_sources[0] if label_sources else None,
                "kind": "section",
            }

        row_section = list(current_section)
        signature = {
            "labels": _normalized_parts(labels),
            "section": _normalized_parts(row_section),
            "role": role,
        }
        signature_hash = stable_hash(signature)
        occurrence = occurrences[signature_hash]
        occurrences[signature_hash] += 1
        row_key = (
            f"{slugify(metric_label or role, fallback='row')}-"
            f"{stable_hash({'signature': signature, 'occurrence': occurrence})[:8]}"
        )

        descriptor = RowDescriptor(
            row=row,
            role=role,
            label_path=labels,
            label_sources=label_sources,
            metric_label=metric_label,
            section_path=row_section,
            is_total=is_total,
            is_placeholder=is_placeholder,
            confidence=round(confidence, 6),
        )
        payload = asdict(descriptor)
        payload["key"] = row_key
        descriptors.append(payload)

    if open_section is not None:
        open_section["end_row"] = bounds.max_row
        open_section["range"] = (
            f"{get_column_letter(bounds.min_col)}{open_section['start_row']}:"
            f"{get_column_letter(bounds.max_col)}{open_section['end_row']}"
        )
        sections.append(open_section)
    return descriptors, sections


def _column_groups(
    ws: Worksheet,
    bounds: Bounds,
    header_rows: Sequence[int],
    columns: Sequence[Mapping[str, Any]],
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    merges: _MergeLookup,
    title: str | None,
) -> list[dict[str, Any]]:
    groups: dict[tuple[int, int, str], dict[str, Any]] = {}
    header_set = set(header_rows)
    for merge in merges.ranges:
        if merge.width < 2 or not any(
            row in header_set for row in range(merge.min_row, merge.max_row + 1)
        ):
            continue
        start = max(bounds.min_col, merge.min_col)
        end = min(bounds.max_col, merge.max_col)
        if end <= start:
            continue
        text, source = _effective_text(
            ws, feature_by_coord, merges, merge.min_row, merge.min_col
        )
        if not text or (title and text.casefold() == title.casefold()):
            continue
        key_tuple = (start, end, text.casefold())
        groups[key_tuple] = {
            "key": (
                f"{slugify(text, fallback='column-group')}-"
                f"{stable_hash({'label': text.casefold(), 'width': end - start + 1})[:8]}"
            ),
            "label": text,
            "header_path": [text],
            "start_column": start,
            "end_column": end,
            "start_letter": get_column_letter(start),
            "end_letter": get_column_letter(end),
            "range": (
                f"{get_column_letter(start)}{merge.min_row}:"
                f"{get_column_letter(end)}{merge.max_row}"
            ),
            "source": source,
        }

    # Fallback for non-merged but repeated top-level header labels.
    value_columns = [
        column
        for column in columns
        if column["role"] in {"measure", "formula_measure", "variance"}
    ]
    run: list[Mapping[str, Any]] = []
    run_label: str | None = None

    def flush() -> None:
        nonlocal run, run_label
        if run_label and len(run) >= 2 and not (
            title and run_label.casefold() == title.casefold()
        ):
            start = int(run[0]["column"])
            end = int(run[-1]["column"])
            key_tuple = (start, end, run_label.casefold())
            groups.setdefault(
                key_tuple,
                {
                    "key": (
                        f"{slugify(run_label, fallback='column-group')}-"
                        f"{stable_hash({'label': run_label.casefold(), 'width': end - start + 1})[:8]}"
                    ),
                    "label": run_label,
                    "header_path": [run_label],
                    "start_column": start,
                    "end_column": end,
                    "start_letter": get_column_letter(start),
                    "end_letter": get_column_letter(end),
                    "range": (
                        f"{get_column_letter(start)}{min(header_rows) if header_rows else bounds.min_row}:"
                        f"{get_column_letter(end)}{max(header_rows) if header_rows else bounds.min_row}"
                    ),
                    "source": (run[0].get("header_sources") or [None])[0],
                },
            )
        run = []
        run_label = None

    for column in value_columns:
        path = list(column.get("header_path") or [])
        label = path[0] if len(path) > 1 else None
        contiguous = bool(run) and int(column["column"]) == int(run[-1]["column"]) + 1
        if label and label == run_label and contiguous:
            run.append(column)
        else:
            flush()
            if label:
                run_label = label
                run = [column]
    flush()
    return [groups[key] for key in sorted(groups)]


def _formula_signature(formula: str, coordinate: str) -> str:
    try:
        return Translator(formula, origin=coordinate).translate_formula("A1")
    except Exception:
        return normalize_whitespace(formula).casefold()


def _fingerprints(
    ws: Worksheet,
    candidate: TableCandidate,
    title: str | None,
    header_rows: Sequence[int],
    columns: Sequence[Mapping[str, Any]],
    rows: Sequence[Mapping[str, Any]],
    sections: Sequence[Mapping[str, Any]],
    column_groups: Sequence[Mapping[str, Any]],
    feature_by_coord: Mapping[tuple[int, int], CellFeature],
    merges: _MergeLookup,
) -> dict[str, str]:
    bounds = candidate.bounds
    schema_payload = {
        "title": normalize_whitespace(title or "").casefold(),
        "columns": [
            {
                "header": _normalized_parts(column.get("header_path") or []),
                "role": column.get("role"),
                "scenario": column.get("scenario"),
                "year": column.get("year"),
                "comparison_kind": column.get("comparison_kind"),
                "unit": column.get("unit"),
                "measure": normalize_whitespace(column.get("measure") or "").casefold(),
            }
            for column in columns
        ],
        "rows": [
            {
                "labels": _normalized_parts(row.get("label_path") or []),
                "section": _normalized_parts(row.get("section_path") or []),
                "role": row.get("role"),
                "is_total": row.get("is_total"),
            }
            for row in rows
            if row.get("role") != "spacer"
        ],
    }
    layout_payload = {
        "width": bounds.width,
        "height": bounds.height,
        "headers": [row - bounds.min_row for row in header_rows],
        "merges": [
            (
                merge.min_row - bounds.min_row,
                merge.min_col - bounds.min_col,
                merge.max_row - bounds.min_row,
                merge.max_col - bounds.min_col,
            )
            for merge in merges.ranges
            if bounds.contains_bounds(merge)
        ],
        "sections": [
            (
                section["start_row"] - bounds.min_row,
                section["end_row"] - bounds.min_row,
                normalize_whitespace(section["title"]).casefold(),
            )
            for section in sections
        ],
        "column_groups": [
            (
                group["start_column"] - bounds.min_col,
                group["end_column"] - bounds.min_col,
                normalize_whitespace(group["label"]).casefold(),
            )
            for group in column_groups
        ],
    }

    formula_payload: list[tuple[int, int, str]] = []
    style_payload: list[tuple[int, int, str]] = []
    for (row, col), feature in sorted(feature_by_coord.items()):
        if not bounds.contains(row, col):
            continue
        if feature.style_hash:
            style_payload.append(
                (row - bounds.min_row, col - bounds.min_col, feature.style_hash)
            )
        if feature.has_formula:
            try:
                formula = ws.cell(row=row, column=col).value
            except Exception:
                formula = None
            if isinstance(formula, str) and formula.startswith("="):
                coordinate = f"{get_column_letter(col)}{row}"
                formula_payload.append(
                    (
                        row - bounds.min_row,
                        col - bounds.min_col,
                        _formula_signature(formula, coordinate),
                    )
                )

    semantic_payload = [
        (
            column.get("role"),
            column.get("scenario"),
            column.get("year"),
            column.get("comparison_kind"),
            column.get("unit"),
            normalize_whitespace(column.get("measure") or "").casefold(),
        )
        for column in columns
    ]
    schema_hash = stable_hash(schema_payload)
    return {
        "table_key": f"table-{stable_hash({'title': schema_payload['title'], 'schema': schema_hash})[:16]}",
        "schema": schema_hash,
        "layout": stable_hash(layout_payload),
        "formula": stable_hash(formula_payload),
        "style": stable_hash(style_payload),
        "semantics": stable_hash(semantic_payload),
    }


def _layout_kind(
    candidate: TableCandidate,
    columns: Sequence[Mapping[str, Any]],
    rows: Sequence[Mapping[str, Any]],
    sections: Sequence[Mapping[str, Any]],
    column_groups: Sequence[Mapping[str, Any]],
) -> str:
    methods = {method.casefold() for method in candidate.methods}
    if any("listobject" in method or "native-table" in method for method in methods):
        return "list_table"
    value_columns = sum(
        column["role"] in {"measure", "formula_measure", "variance"}
        for column in columns
    )
    label_columns = sum(
        column["role"] in {"row_group", "row_label", "dimension"}
        for column in columns
    )
    placeholder_rows = sum(bool(row.get("is_placeholder")) for row in rows)
    if column_groups and sections:
        return "compound_matrix"
    if column_groups:
        return "horizontal_compound_matrix"
    if sections and value_columns >= 2:
        return "sectioned_matrix"
    if value_columns >= 2 and label_columns:
        return "matrix"
    if placeholder_rows and not any(
        row.get("role") == "data" and not row.get("is_placeholder") for row in rows
    ):
        return "input_grid"
    if value_columns == 1 and label_columns:
        return "list"
    return "region"


def infer_table_structure(
    ws: Worksheet,
    candidate: TableCandidate,
    features: FeatureCollection,
    merge_index: object,
) -> dict[str, Any]:
    """Infer a coordinate-backed logical structure for one table candidate.

    The returned mapping contains only JSON-serializable primitives.  Header
    propagation is limited to exact merged ranges and conservative same-style
    parent spans; blank cells are never globally forward-filled.
    """

    feature_by_coord = _feature_map(features)
    bounds = candidate.bounds
    merges = _merge_lookup(ws, merge_index, bounds)
    rows_by_number = _row_features(feature_by_coord, bounds)
    header_rows, warnings = _infer_header_rows(
        ws, bounds, feature_by_coord, rows_by_number, merges
    )
    title = _infer_title(ws, bounds, feature_by_coord, merges, header_rows)
    header_grid = _build_header_grid(
        ws, bounds, header_rows, feature_by_coord, merges
    )
    columns = _column_descriptors(
        bounds,
        header_rows,
        header_grid,
        rows_by_number,
        merges,
        title,
    )
    row_descriptors, sections = _row_descriptors(
        ws,
        bounds,
        header_rows,
        columns,
        feature_by_coord,
        rows_by_number,
        merges,
    )
    column_groups = _column_groups(
        ws,
        bounds,
        header_rows,
        columns,
        feature_by_coord,
        merges,
        title,
    )
    layout_kind = _layout_kind(
        candidate, columns, row_descriptors, sections, column_groups
    )

    if title is None:
        warnings.append("No reliable table title was detected.")
    if not any(
        column["role"] in {"row_group", "row_label", "dimension"}
        for column in columns
    ):
        warnings.append("No reliable row-label column was detected.")
    if not any(
        column["role"] in {"measure", "formula_measure", "variance"}
        for column in columns
    ):
        warnings.append("No reliable value column was detected.")
    if candidate.uncertain:
        warnings.append(
            "Candidate confidence is below the automatic table threshold."
        )
    for merge in merges.ranges:
        if bounds.intersects(merge) and not bounds.contains_bounds(merge):
            warnings.append(
                f"Candidate boundary cuts through merged range {merge.ref}."
            )

    fingerprints = _fingerprints(
        ws,
        candidate,
        title,
        header_rows,
        columns,
        row_descriptors,
        sections,
        column_groups,
        feature_by_coord,
        merges,
    )
    return {
        "title": title,
        "layout_kind": layout_kind,
        "header_rows": sorted(header_rows),
        "column_descriptors": sorted(
            columns, key=lambda item: int(item["column"])
        ),
        "row_descriptors": sorted(
            row_descriptors, key=lambda item: int(item["row"])
        ),
        "sections": sorted(
            sections,
            key=lambda item: (
                int(item["start_row"]),
                int(item["end_row"]),
                str(item["title"]),
            ),
        ),
        "column_groups": sorted(
            column_groups,
            key=lambda item: (
                int(item["start_column"]),
                int(item["end_column"]),
                str(item["label"]),
            ),
        ),
        "fingerprints": fingerprints,
        "warnings": unique_preserving_order(warnings),
    }


def _candidate_identifier(candidate: TableCandidate) -> str:
    if candidate.source_name:
        return candidate.source_name
    return (
        "table-"
        + stable_hash(
            {
                "range": candidate.bounds.ref,
                "methods": sorted(candidate.methods),
            }
        )[:12]
    )


def _distance(left: Bounds, right: Bounds) -> int:
    row_gap = max(
        0,
        left.min_row - right.max_row,
        right.min_row - left.max_row,
    )
    col_gap = max(
        0,
        left.min_col - right.max_col,
        right.min_col - left.max_col,
    )
    return row_gap + col_gap


def _annotation_kind(
    texts: Sequence[str], features: Sequence[CellFeature]
) -> str:
    joined = " | ".join(texts)
    if _SOURCE_RE.search(joined):
        return "source_note"
    if _INSTRUCTION_RE.search(joined):
        return "instruction"
    if _NOTE_RE.search(joined) or any(len(text) >= 80 for text in texts):
        return "note"
    if len(texts) == 1 and len(texts[0]) <= 80 and any(
        feature.bold or feature.has_fill for feature in features
    ):
        return "title"
    if any(feature.has_comment for feature in features) and not texts:
        return "comment"
    return "annotation"


def _text_for_feature(feature: CellFeature) -> str | None:
    text = normalize_whitespace(feature.text or "")
    if text:
        return text
    if feature.has_comment:
        return f"[comment at {feature.coordinate}]"
    if feature.has_hyperlink:
        return f"[hyperlink at {feature.coordinate}]"
    return None


def _nearest_table(
    bounds: Bounds, accepted_candidates: Sequence[TableCandidate]
) -> tuple[str | None, int | None]:
    if not accepted_candidates:
        return None, None
    ranked = sorted(
        (
            _distance(bounds, candidate.bounds),
            candidate.bounds.min_row,
            candidate.bounds.min_col,
            _candidate_identifier(candidate),
        )
        for candidate in accepted_candidates
    )
    distance, _, _, identifier = ranked[0]
    return identifier, distance


def _annotation_components(
    features: Mapping[tuple[int, int], CellFeature]
) -> list[list[CellFeature]]:
    if not features:
        return []
    parent = {coord: coord for coord in features}

    def find(coord: tuple[int, int]) -> tuple[int, int]:
        while parent[coord] != coord:
            parent[coord] = parent[parent[coord]]
            coord = parent[coord]
        return coord

    def union(left: tuple[int, int], right: tuple[int, int]) -> None:
        root_left = find(left)
        root_right = find(right)
        if root_left == root_right:
            return
        if root_right < root_left:
            root_left, root_right = root_right, root_left
        parent[root_right] = root_left

    by_row: dict[int, list[CellFeature]] = defaultdict(list)
    by_col: dict[int, list[CellFeature]] = defaultdict(list)
    for feature in features.values():
        by_row[feature.row].append(feature)
        by_col[feature.col].append(feature)
    for row_features in by_row.values():
        ordered = sorted(row_features, key=lambda item: item.col)
        for left, right in zip(ordered, ordered[1:]):
            if right.col - left.col - 1 <= 6:
                union((left.row, left.col), (right.row, right.col))
    for col_features in by_col.values():
        ordered = sorted(col_features, key=lambda item: item.row)
        for upper, lower in zip(ordered, ordered[1:]):
            if lower.row - upper.row - 1 <= 1:
                union((upper.row, upper.col), (lower.row, lower.col))

    groups: dict[tuple[int, int], list[CellFeature]] = defaultdict(list)
    for coord, feature in sorted(features.items()):
        groups[find(coord)].append(feature)
    return [
        sorted(group, key=lambda item: (item.row, item.col))
        for _, group in sorted(groups.items())
    ]


def infer_annotation_blocks(
    features: FeatureCollection,
    accepted_candidates: Sequence[TableCandidate],
    rejected_candidates: Sequence[TableCandidate],
) -> list[AnnotationBlock]:
    """Group annotation-like cells outside accepted table regions.

    Rejected detector regions are considered first so their original bounds are
    retained.  Remaining unassigned text/comment/hyperlink cells are grouped by
    sparse proximity.  Blocks are classified and linked to the nearest accepted
    table using a deterministic Manhattan distance between rectangles.
    """

    feature_by_coord = _feature_map(features)
    accepted = sorted(
        accepted_candidates,
        key=lambda item: (item.bounds, item.source_name or ""),
    )
    rejected = sorted(
        rejected_candidates,
        key=lambda item: (item.bounds, item.source_name or ""),
    )

    def inside_accepted(row: int, col: int) -> bool:
        return any(candidate.bounds.contains(row, col) for candidate in accepted)

    blocks: list[AnnotationBlock] = []
    covered: set[tuple[int, int]] = set()
    seen_rejected_bounds: set[Bounds] = set()
    for candidate in rejected:
        if candidate.bounds in seen_rejected_bounds:
            continue
        seen_rejected_bounds.add(candidate.bounds)
        block_features = [
            feature
            for (row, col), feature in feature_by_coord.items()
            if candidate.bounds.contains(row, col) and not inside_accepted(row, col)
        ]
        texts = unique_preserving_order(
            text
            for feature in sorted(block_features, key=lambda item: (item.row, item.col))
            if (text := _text_for_feature(feature))
        )
        if not texts:
            continue
        covered.update((feature.row, feature.col) for feature in block_features)
        nearest_id, distance = _nearest_table(candidate.bounds, accepted)
        blocks.append(
            AnnotationBlock(
                bounds=candidate.bounds,
                text=texts,
                kind=_annotation_kind(texts, block_features),
                nearest_table_id=nearest_id,
                distance=distance,
            )
        )

    remaining = {
        coord: feature
        for coord, feature in feature_by_coord.items()
        if coord not in covered
        and not inside_accepted(*coord)
        and _text_for_feature(feature) is not None
    }
    for component in _annotation_components(remaining):
        bounds = Bounds(
            min(feature.row for feature in component),
            min(feature.col for feature in component),
            max(feature.row for feature in component),
            max(feature.col for feature in component),
        )
        texts = unique_preserving_order(
            text
            for feature in component
            if (text := _text_for_feature(feature))
        )
        if not texts:
            continue
        nearest_id, distance = _nearest_table(bounds, accepted)
        blocks.append(
            AnnotationBlock(
                bounds=bounds,
                text=texts,
                kind=_annotation_kind(texts, component),
                nearest_table_id=nearest_id,
                distance=distance,
            )
        )

    blocks.sort(
        key=lambda item: (
            item.bounds.min_row,
            item.bounds.min_col,
            item.bounds.max_row,
            item.bounds.max_col,
            item.kind,
            item.text,
        )
    )
    return blocks


__all__ = ["infer_annotation_blocks", "infer_table_structure"]
